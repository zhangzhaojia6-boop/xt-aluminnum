from __future__ import annotations

import json
from copy import deepcopy
from datetime import UTC, date, datetime, time
from pathlib import Path

import pytest
from sqlalchemy import create_engine
from sqlalchemy.orm import Session

from app.database import Base
from app.models.consumable import DailyConsumableLog
from app.models.energy import MachineEnergyRecord
from app.models.executive import CostDailyResult, MachineDailyCostSnapshot
from app.models.master import Equipment, Team, Workshop
from app.models.mes import MesDailyWipSnapshot, MesStockRecord, MesWipTotalSnapshot, MesWorkshopProcessRecord
from app.models.production import MobileShiftReport, ShiftProductionData, WorkOrder, WorkOrderEntry
from app.models.shift import ShiftConfig
from app.models.system import User
from app.services.mapping_reconciliation_service import (
    MappingFieldSpec,
    build_system_mapping_rows,
    compare_mapping_rows,
    parse_output_skill_reference_file,
    propose_rules,
)


FIXTURE_PATH = Path(__file__).parent / 'fixtures' / 'output_skill_mapping_sample.json'
RECONCILIATION_TABLES = [
    Workshop.__table__,
    Team.__table__,
    User.__table__,
    ShiftConfig.__table__,
    Equipment.__table__,
    ShiftProductionData.__table__,
    WorkOrder.__table__,
    WorkOrderEntry.__table__,
    MobileShiftReport.__table__,
    MachineEnergyRecord.__table__,
    CostDailyResult.__table__,
    MachineDailyCostSnapshot.__table__,
    DailyConsumableLog.__table__,
    MesWorkshopProcessRecord.__table__,
    MesStockRecord.__table__,
    MesDailyWipSnapshot.__table__,
    MesWipTotalSnapshot.__table__,
]


def test_compare_mapping_rows_converts_kg_to_tons_and_reports_match_rate() -> None:
    reference_rows = [
        {
            'business_date': '2026-06-13',
            'workshop': '精整',
            'shift': '长白班',
            'output_tons': 12.5,
        }
    ]
    system_rows = [
        {
            'business_date': '2026-06-13',
            'workshop': '精整车间',
            'shift': '白班',
            'output_kg': 12500,
        }
    ]

    result = compare_mapping_rows(
        reference_rows=reference_rows,
        system_rows=system_rows,
        fields=[
            MappingFieldSpec(
                metric='output',
                reference_field='output_tons',
                system_field='output_kg',
                reference_unit='ton',
                system_unit='kg',
                tolerance=0.001,
                weight=30,
            )
        ],
        dimension_aliases={'workshop': {'精整车间': '精整'}, 'shift': {'白班': '长白班'}},
    )

    assert result.total_fields == 1
    assert result.matched_fields == 1
    assert result.overall_match_rate == 100
    assert result.differences == []


def test_output_skill_mapping_fixture_matches_after_alias_and_unit_normalization() -> None:
    payload = json.loads(FIXTURE_PATH.read_text(encoding='utf-8'))

    result = compare_mapping_rows(
        reference_rows=payload['reference_rows'],
        system_rows=payload['system_rows'],
        fields=[MappingFieldSpec(**item) for item in payload['fields']],
        dimension_aliases=payload['dimension_aliases'],
    )

    assert result.total_fields == 2
    assert result.matched_fields == 2
    assert result.field_match_rates == {'output': 100, 'energy': 100}
    assert result.differences == []


def test_compare_mapping_rows_explains_value_diff_and_missing_rows() -> None:
    reference_rows = [
        {
            'business_date': '2026-06-13',
            'workshop': '拉矫',
            'shift': '小夜班',
            'energy_kwh': 1800,
        },
        {
            'business_date': '2026-06-13',
            'workshop': '园区剪切',
            'shift': '小夜班',
            'energy_kwh': 900,
        },
    ]
    system_rows = [
        {
            'business_date': '2026-06-13',
            'workshop': '拉矫车间',
            'shift': '小夜',
            'electricity_kwh': 1760,
        }
    ]

    result = compare_mapping_rows(
        reference_rows=reference_rows,
        system_rows=system_rows,
        fields=[
            MappingFieldSpec(
                metric='energy',
                reference_field='energy_kwh',
                system_field='electricity_kwh',
                reference_unit='kwh',
                system_unit='kwh',
                tolerance=5,
                weight=15,
            )
        ],
        dimension_aliases={'workshop': {'拉矫车间': '拉矫'}, 'shift': {'小夜': '小夜班'}},
    )

    assert result.overall_match_rate == 0
    assert [item.reason_code for item in result.differences] == ['value_diff', 'missing_system_row']
    assert result.differences[0].reference_value == 1800
    assert result.differences[0].system_value == 1760
    assert result.differences[0].suggested_rule == '检查 energy 字段口径、单位或时间范围。'
    assert result.differences[1].dimension['workshop'] == '园区剪切'


def test_compare_mapping_rows_merges_duplicate_dimensions_without_overwriting_fields() -> None:
    reference_rows = [
        {
            'business_date': '2026-06-16',
            'workshop': '全厂',
            'shift': '',
            'wip_total': 879,
        },
        {
            'business_date': '2026-06-16',
            'workshop': '全厂',
            'shift': '',
            'total_electricity_kwh': 168000,
        },
    ]
    system_rows = [
        {
            'business_date': '2026-06-16',
            'workshop': '全厂',
            'shift': '',
            'wip_total': 879,
            'total_electricity_kwh': 168000,
        }
    ]

    result = compare_mapping_rows(
        reference_rows=reference_rows,
        system_rows=system_rows,
        fields=[
            MappingFieldSpec(
                metric='wip_total',
                reference_field='wip_total',
                system_field='wip_total',
                reference_unit='ton',
                system_unit='ton',
                tolerance=0.001,
                weight=8,
            ),
            MappingFieldSpec(
                metric='total_electricity',
                reference_field='total_electricity_kwh',
                system_field='total_electricity_kwh',
                reference_unit='kwh',
                system_unit='kwh',
                tolerance=0.1,
                weight=8,
            ),
        ],
        dimensions=['business_date', 'workshop'],
    )

    assert result.total_fields == 2
    assert result.matched_fields == 2
    assert result.overall_match_rate == 100
    assert result.differences == []


def test_propose_rules_is_dry_run_and_does_not_mutate_source_rows() -> None:
    reference_rows = [
        {
            'business_date': '2026-06-13',
            'workshop': '在线退火',
            'shift': '大夜班',
            'output_tons': 20,
        }
    ]
    system_rows = [
        {
            'business_date': '2026-06-13',
            'workshop': '新厂在线退火',
            'shift': '第一班',
            'output_tons': 20,
        }
    ]
    before_reference = deepcopy(reference_rows)
    before_system = deepcopy(system_rows)

    result = compare_mapping_rows(
        reference_rows=reference_rows,
        system_rows=system_rows,
        fields=[
            MappingFieldSpec(
                metric='output',
                reference_field='output_tons',
                system_field='output_tons',
                reference_unit='ton',
                system_unit='ton',
                tolerance=0.001,
                weight=30,
            )
        ],
        dimension_aliases={},
    )
    proposals = propose_rules(result.differences)

    assert [item.reason_code for item in result.differences] == ['missing_system_row', 'extra_system_row']
    assert proposals == [
        {
            'rule_type': 'alias_candidate',
            'field': 'workshop',
            'reference_value': '在线退火',
            'system_value': '新厂在线退火',
            'confidence': 'manual_review',
            'dry_run': True,
        },
        {
            'rule_type': 'alias_candidate',
            'field': 'shift',
            'reference_value': '大夜班',
            'system_value': '第一班',
            'confidence': 'manual_review',
            'dry_run': True,
        },
    ]
    assert reference_rows == before_reference
    assert system_rows == before_system


def test_parse_output_skill_text_file_extracts_business_metrics(tmp_path) -> None:
    report = tmp_path / '2026-06-13-daily.txt'
    report.write_text(
        '2026年6月13日 生产日报\n'
        '精整 长白班 投料 13 吨 产量 12.5 吨 能耗 1800 度 燃气 32 m3 用电月累计 131500 用电指标 130000 用气月累计 53433 用气指标 53000 全厂用电 131500 新厂用电 6124 园区用电 9056 铸轧用气 2599 熔炼炉用气 5348 加热炉用气 876 锅炉用气 1087 天然气总量 53433 地下水 3860 自来水 4295 废料 0.2 吨 停机 30 分钟 质量异常 2 项 成材率 96.15% 轧制油吨耗 1.25 液化气吨耗 0.05 钛丝吨耗 0.01 镁吨耗 0.02 锰吨耗 0.03 铁吨耗 0.04 铜吨耗 0.005 滤布日耗 3 高温胶带日耗 2 再生油出库 0.8 再生油入库 0.6 液压油日耗 1.2 液压油月累计 10 液压油指标 0.5 齿轮油日耗 0.7 齿轮油月累计 6 齿轮油指标 0.4 铸锭块数 12 铸锭投料量 20 铸锭下机量 19 当日接合同 143 月累计合同 2422 余合同量 400 余热轧合同 125 余合同较昨日 77 坯料总量 120.5 当日投料 70 月累计投料 905 总成本 12800.5 元 单吨成本 867 元/吨 过站吨成本 280.25 元/吨\n'
        '拉矫 小夜班 下机量 8000 kg 能耗 950 kWh\n',
        encoding='utf-8',
    )

    result = parse_output_skill_reference_file(report)

    assert result['status'] == 'parsed'
    assert result['source_type'] == 'output_skill_text'
    assert result['rows'] == [
        {
            'business_date': '2026-06-13',
            'workshop': '精整',
            'shift': '长白班',
            'input_tons': 13.0,
            'output_tons': 12.5,
            'energy_kwh': 1800.0,
            'gas_m3': 32.0,
            'electricity_monthly': 131500.0,
            'electricity_target': 130000.0,
            'gas_monthly': 53433.0,
            'gas_target': 53000.0,
            'total_electricity_kwh': 131500.0,
            'new_plant_electricity_kwh': 6124.0,
            'park_electricity_kwh': 9056.0,
            'cast_roll_gas_m3': 2599.0,
            'smelting_gas_m3': 5348.0,
            'heating_furnace_gas_m3': 876.0,
            'boiler_gas_m3': 1087.0,
            'total_gas_m3': 53433.0,
            'groundwater_ton': 3860.0,
            'tap_water_ton': 4295.0,
            'scrap_tons': 0.2,
            'downtime_minutes': 30.0,
            'quality_issue_count': 2.0,
            'yield_rate': 96.15,
            'rolling_oil_per_ton': 1.25,
            'liquefied_gas_per_ton': 0.05,
            'titanium_wire_per_ton': 0.01,
            'magnesium_per_ton': 0.02,
            'manganese_per_ton': 0.03,
            'iron_per_ton': 0.04,
            'copper_per_ton': 0.005,
            'filter_cloth_daily': 3.0,
            'high_temp_tape_daily': 2.0,
            'regen_oil_out': 0.8,
            'regen_oil_in': 0.6,
            'hydraulic_oil_daily': 1.2,
            'hydraulic_oil_monthly': 10.0,
            'hydraulic_oil_target': 0.5,
            'gear_oil_daily': 0.7,
            'gear_oil_monthly': 6.0,
            'gear_oil_target': 0.4,
            'ingot_block_count': 12.0,
            'ingot_input_tons': 20.0,
            'ingot_output_tons': 19.0,
            'daily_contract_weight': 143.0,
            'month_to_date_contract_weight': 2422.0,
            'remaining_contract_weight': 400.0,
            'remaining_hot_roll_contract_weight': 125.0,
            'remaining_contract_delta_weight': 77.0,
            'billet_inventory_weight': 120.5,
            'daily_input_weight': 70.0,
            'month_to_date_input_weight': 905.0,
            'total_cost': 12800.5,
            'cost_per_ton': 867.0,
            'throughput_cost_per_ton': 280.25,
            'source_file': str(report),
            'source_type': 'output_skill_text',
        },
        {
            'business_date': '2026-06-13',
            'workshop': '拉矫',
            'shift': '小夜班',
            'output_tons': 8.0,
            'energy_kwh': 950.0,
            'source_file': str(report),
            'source_type': 'output_skill_text',
        },
    ]


def test_parse_output_skill_text_file_extracts_factory_narrative_rows(tmp_path) -> None:
    report = tmp_path / '2026-6-14_日报正文.txt'
    report.write_text(
        '6月14日，车间总产量日合计221吨。\n'
        '当天在制料1205吨；全厂高压总用电量126500度（分项用电124874度）；'
        '铸轧用气11977m³、铸锭熔炼炉用气25991m³、热轧加热炉用气8430m³、热轧锅炉用气867m³，共计59141m³。\n'
        '入库成品日合计221吨（寄存117吨），月累计4380吨。当天接合同192吨（含热轧158吨）；'
        '冷轧日投料336吨（2050投307吨、1850投0吨、外加工29吨），中厚板0吨，总余合同量2632吨，比昨日↑6吨。\n'
        '成本核算方面，电费约10.12万元、气费约21.29万元，已核合计约31.41万元，按220.671吨折算约1423元/吨。\n',
        encoding='utf-8',
    )

    result = parse_output_skill_reference_file(report)

    assert result['status'] == 'parsed'
    assert result['source_type'] == 'output_skill_text'
    assert result['rows'] == [
        {
            'business_date': '2026-06-14',
            'workshop': '全厂',
            'shift': '',
            'total_electricity_kwh': 126500.0,
            'cast_roll_gas_m3': 11977.0,
            'smelting_gas_m3': 25991.0,
            'heating_furnace_gas_m3': 8430.0,
            'boiler_gas_m3': 867.0,
            'wip_total': 1205.0,
            'total_gas_m3': 59141.0,
            'source_file': str(report),
            'source_type': 'output_skill_text',
        },
        {
            'business_date': '2026-06-14',
            'workshop': '全厂',
            'shift': '',
            'daily_contract_weight': 192.0,
            'daily_hot_roll_contract_weight': 158.0,
            'remaining_contract_weight': 2632.0,
            'daily_input_weight': 336.0,
            'source_file': str(report),
            'source_type': 'output_skill_text',
        },
        {
            'business_date': '2026-06-14',
            'workshop': '全厂',
            'shift': '',
            'electricity_cost': 101200.0,
            'natural_gas_cost': 212900.0,
            'total_cost': 314100.0,
            'cost_per_ton': 1423.0,
            'source_file': str(report),
            'source_type': 'output_skill_text',
        },
    ]


def test_parse_output_skill_text_file_uses_date_from_filename_when_body_has_no_date(tmp_path) -> None:
    report = tmp_path / '2026-06-13-daily.txt'
    report.write_text(
        '精整 长白班 产量 12.5 吨 能耗 1800 度\n',
        encoding='utf-8',
    )

    result = parse_output_skill_reference_file(report)

    assert result['status'] == 'parsed'
    assert result['rows'] == [
        {
            'business_date': '2026-06-13',
            'workshop': '精整',
            'shift': '长白班',
            'output_tons': 12.5,
            'energy_kwh': 1800.0,
            'source_file': str(report),
            'source_type': 'output_skill_text',
        }
    ]


def test_parse_output_skill_xlsx_file_normalizes_common_columns(tmp_path) -> None:
    from openpyxl import Workbook

    report = tmp_path / 'mapping.xlsx'
    workbook = Workbook()
    sheet = workbook.active
    sheet.append([
        '日期',
        '车间',
        '班次',
        '机台',
        '工序',
        '卷号',
        '合同号',
        '客户',
        '产量(吨)',
        '能耗(kWh)',
        '用电月累计',
        '用电指标',
        '废料(吨)',
        '停机(分钟)',
        '质量异常数',
        '良品率(%)',
        '轧制油吨耗',
        '热轧乳液吨耗',
        '硅藻土吨耗',
        '白土吨耗',
        '滤布日耗',
        '高温胶带日耗',
        '再生油出库',
        '再生油入库',
        '用气月累计',
        '用气指标',
        '全厂用电',
        '新厂用电',
        '园区用电',
        '铸轧用气',
        '熔炼炉用气',
        '加热炉用气',
        '锅炉用气',
        '天然气总量',
        '地下水',
        '自来水',
        '液压油日耗',
        '齿轮油日耗',
        '铸锭投料量',
        '铸锭下机量',
        '当日接合同',
        '月累计合同',
        '余合同量',
        '余热轧合同',
        '余合同较昨日',
        '坯料总量',
        '当日投料',
        '月累计投料',
        'D40吨耗',
        '综合成本(元/吨)',
        '过站吨成本',
    ])
    sheet.append(['2026-06-13', '园区剪切', '长白班', 'JQ-01', '包装', '26A04967', 'HT-001', '客户A', 9.75, 1200, 131500, 130000, 0.12, 25, 1, 0.942, 1.1, 0.6, 0.07, 0.09, 4, 2, 0.8, 0.6, 53433, 53000, 131500, 6124, 9056, 2599, 5348, 876, 1087, 53433, 3860, 4295, 1.2, 0.7, 20, 19, 143, 2422, 400, 125, 77, 120.5, 70, 905, 0.2, 331.16, 280.25])
    workbook.save(report)

    result = parse_output_skill_reference_file(report)

    assert result['status'] == 'parsed'
    assert result['source_type'] == 'output_skill_excel'
    assert result['rows'] == [
        {
            'business_date': '2026-06-13',
            'workshop': '园区剪切',
            'shift': '长白班',
            'machine': 'JQ-01',
            'process': '包装',
            'coil_no': '26A04967',
            'contract_no': 'HT-001',
            'customer': '客户A',
            'output_tons': 9.75,
            'energy_kwh': 1200.0,
            'electricity_monthly': 131500.0,
            'electricity_target': 130000.0,
            'scrap_tons': 0.12,
            'downtime_minutes': 25.0,
            'quality_issue_count': 1.0,
            'yield_rate': 94.2,
            'rolling_oil_per_ton': 1.1,
            'hot_roll_emulsion_per_ton': 0.6,
            'diatomite_per_ton': 0.07,
            'white_earth_per_ton': 0.09,
            'filter_cloth_daily': 4.0,
            'high_temp_tape_daily': 2.0,
            'regen_oil_out': 0.8,
            'regen_oil_in': 0.6,
            'gas_monthly': 53433.0,
            'gas_target': 53000.0,
            'total_electricity_kwh': 131500.0,
            'new_plant_electricity_kwh': 6124.0,
            'park_electricity_kwh': 9056.0,
            'cast_roll_gas_m3': 2599.0,
            'smelting_gas_m3': 5348.0,
            'heating_furnace_gas_m3': 876.0,
            'boiler_gas_m3': 1087.0,
            'total_gas_m3': 53433.0,
            'groundwater_ton': 3860.0,
            'tap_water_ton': 4295.0,
            'hydraulic_oil_daily': 1.2,
            'gear_oil_daily': 0.7,
            'ingot_input_tons': 20.0,
            'ingot_output_tons': 19.0,
            'daily_contract_weight': 143.0,
            'month_to_date_contract_weight': 2422.0,
            'remaining_contract_weight': 400.0,
            'remaining_hot_roll_contract_weight': 125.0,
            'remaining_contract_delta_weight': 77.0,
            'billet_inventory_weight': 120.5,
            'daily_input_weight': 70.0,
            'month_to_date_input_weight': 905.0,
            'd40_per_ton': 0.2,
            'cost_per_ton': 331.16,
            'throughput_cost_per_ton': 280.25,
            'source_file': str(report),
            'source_type': 'output_skill_excel',
        }
    ]


def test_parse_output_skill_xls_file_normalizes_common_columns(tmp_path) -> None:
    xlwt = pytest.importorskip('xlwt')

    report = tmp_path / 'mapping.xls'
    workbook = xlwt.Workbook()
    sheet = workbook.add_sheet('日报')
    headers = [
        '日期',
        '车间',
        '班次',
        '设备名称',
        '当前工艺',
        '随行卡号',
        '合同',
        '客户名',
        '产量(吨)',
        '能耗(kWh)',
        '废料(吨)',
        '停机分钟',
        '质量问题数',
        '成材率',
        '轧制油单吨消耗',
        '钢带吨耗',
        '钢板吨耗',
        '钢扣吨耗',
        '吨成本',
        '流转吨成本',
    ]
    values = ['2026-06-13', '热轧', '大夜班', 'RZ-02', '热轧', '26B00001', 'HT-002', '客户B', 21.5, 2600, 0.4, 40, 3, 0.928, 1.3, 0.08, 0.11, 0.12, 405.5, 333.3]
    for index, header in enumerate(headers):
        sheet.write(0, index, header)
    for index, value in enumerate(values):
        sheet.write(1, index, value)
    workbook.save(str(report))

    result = parse_output_skill_reference_file(report)

    assert result['status'] == 'parsed'
    assert result['source_type'] == 'output_skill_excel'
    assert result['rows'] == [
        {
            'business_date': '2026-06-13',
            'workshop': '热轧',
            'shift': '大夜班',
            'machine': 'RZ-02',
            'process': '热轧',
            'coil_no': '26B00001',
            'contract_no': 'HT-002',
            'customer': '客户B',
            'output_tons': 21.5,
            'energy_kwh': 2600.0,
            'scrap_tons': 0.4,
            'downtime_minutes': 40.0,
            'quality_issue_count': 3.0,
            'yield_rate': 92.8,
            'rolling_oil_per_ton': 1.3,
            'steel_strip_per_ton': 0.08,
            'steel_plate_per_ton': 0.11,
            'steel_buckle_per_ton': 0.12,
            'cost_per_ton': 405.5,
            'throughput_cost_per_ton': 333.3,
            'source_file': str(report),
            'source_type': 'output_skill_excel',
        }
    ]


def test_parse_output_skill_xls_summary_sheet_accepts_missing_shift_and_header_offset(tmp_path) -> None:
    xlwt = pytest.importorskip('xlwt')

    report = tmp_path / '2026-6-16_日均报表.xls'
    workbook = xlwt.Workbook()
    sheet = workbook.add_sheet('2026-6-16')
    for row_index in range(4):
        sheet.write(row_index, 0, '各工序产量报表' if row_index == 0 else '')
    headers = ['工序', '车间', '日投料量', '月累计投料', '日产量', '月累计产量', '日期', '日电度', '日然气']
    values = ['铸轧', '铸二', 25, 664, 24.31, 644.95, 16, 2869, 4678]
    for index, header in enumerate(headers):
        sheet.write(4, index, header)
    for index, value in enumerate(values):
        sheet.write(5, index, value)
    workbook.save(str(report))

    result = parse_output_skill_reference_file(report)

    assert result['status'] == 'parsed'
    assert result['rows'] == [
        {
            'business_date': '2026-06-16',
            'workshop': '铸二',
            'process': '铸轧',
            'daily_input_weight': 25.0,
            'month_to_date_input_weight': 664.0,
            'output_tons': 24.31,
            'energy_kwh': 2869.0,
            'gas_m3': 4678.0,
            'shift': '',
            'source_file': str(report),
            'source_type': 'output_skill_excel',
        }
    ]


def test_parse_delivery_override_json_extracts_summary_rows(tmp_path) -> None:
    report = tmp_path / 'delivery_override_2026-06-16.json'
    report.write_text(
        json.dumps(
            {
                'rows': {'40': {'out_day': 328.033}},
                'summaries': {
                    'contract': '当天接合同66吨（含热轧66吨），月累计合同5408吨（含热轧4449吨），总余合同量2569吨。',
                    'feed': '当天冷轧投料197吨；坯料总量458吨；入库成品328.033吨。',
                    'power': '全厂高压用电168000度，分项用电166533度；气耗共计57776m³。',
                },
            },
            ensure_ascii=False,
        ),
        encoding='utf-8',
    )

    result = parse_output_skill_reference_file(report)

    assert result['status'] == 'parsed'
    assert result['source_type'] == 'output_skill_json'
    assert result['rows'] == [
        {
            'business_date': '2026-06-16',
            'workshop': '全厂',
            'shift': '',
            'daily_contract_weight': 66.0,
            'daily_hot_roll_contract_weight': 66.0,
            'month_to_date_contract_weight': 5408.0,
            'remaining_contract_weight': 2569.0,
            'source_file': str(report),
            'source_type': 'output_skill_json',
        },
        {
            'business_date': '2026-06-16',
            'workshop': '全厂',
            'shift': '',
            'billet_inventory_weight': 458.0,
            'daily_input_weight': 197.0,
            'source_file': str(report),
            'source_type': 'output_skill_json',
        },
        {
            'business_date': '2026-06-16',
            'workshop': '全厂',
            'shift': '',
            'total_electricity_kwh': 168000.0,
            'total_gas_m3': 57776.0,
            'source_file': str(report),
            'source_type': 'output_skill_json',
        },
    ]


def test_parse_output_skill_json_file_normalizes_common_columns(tmp_path) -> None:
    report = tmp_path / 'mapping.json'
    report.write_text(
        json.dumps(
            {
                'rows': [
                    {
                        '日期': '2026-06-13',
                        '车间': '冷轧1650',
                        '班次': '小夜班',
                        '设备名称': '1650-01',
                        '当前工艺': '冷轧',
                        '随行卡号': '26C00001',
                        '合同号': 'HT-003',
                        '客户名': '客户C',
                        '投入量(吨)': 7.6,
                        '产量(吨)': 7.2,
                        '能耗(kWh)': 980,
                        '用气(m3)': 18.5,
                        '废料(吨)': 0.08,
                        '停机分钟': 90,
                        '质量异常数': 2,
                        '成品率': 0.935,
                        '轧制油吨耗': 1.05,
                        '飞滤剂吨耗': 0.12,
                        '油漆吨耗': 0.03,
                        '总成本(元)': 12800.5,
                        '吨成本': 386.2,
                        '过站吨成本': 280.25,
                    }
                ]
            },
            ensure_ascii=False,
        ),
        encoding='utf-8',
    )

    result = parse_output_skill_reference_file(report)

    assert result['status'] == 'parsed'
    assert result['source_type'] == 'output_skill_json'
    assert result['rows'] == [
        {
            'business_date': '2026-06-13',
            'workshop': '冷轧1650',
            'shift': '小夜班',
            'machine': '1650-01',
            'process': '冷轧',
            'coil_no': '26C00001',
            'contract_no': 'HT-003',
            'customer': '客户C',
            'input_tons': 7.6,
            'output_tons': 7.2,
            'energy_kwh': 980.0,
            'gas_m3': 18.5,
            'scrap_tons': 0.08,
            'downtime_minutes': 90.0,
            'quality_issue_count': 2.0,
            'yield_rate': 93.5,
            'rolling_oil_per_ton': 1.05,
            'filter_agent_per_ton': 0.12,
            'paint_per_ton': 0.03,
            'total_cost': 12800.5,
            'cost_per_ton': 386.2,
            'throughput_cost_per_ton': 280.25,
            'source_file': str(report),
            'source_type': 'output_skill_json',
        }
    ]


def test_parse_output_skill_ndjson_file_normalizes_each_record(tmp_path) -> None:
    report = tmp_path / 'mapping.ndjson'
    report.write_text(
        '\n'.join(
            [
                json.dumps(
                    {
                        '日期': '2026-06-13',
                        '车间': '精整',
                        '班次': '长白班',
                        '设备名称': 'JZ-01',
                        '当前工艺': '包装',
                        '随行卡号': '26D00001',
                        '产量(吨)': 10.25,
                        '能耗(kWh)': 1350,
                    },
                    ensure_ascii=False,
                ),
                json.dumps(
                    {
                        '日期': '2026-06-13',
                        '车间': '拉矫',
                        '班次': '小夜班',
                        '设备名称': 'LJ-02',
                        '当前工艺': '拉矫',
                        '随行卡号': '26D00002',
                        '产量(吨)': 8.5,
                        '成材率': 0.96,
                    },
                    ensure_ascii=False,
                ),
            ]
        ),
        encoding='utf-8',
    )

    result = parse_output_skill_reference_file(report)

    assert result['status'] == 'parsed'
    assert result['source_type'] == 'output_skill_json_lines'
    assert result['rows'] == [
        {
            'business_date': '2026-06-13',
            'workshop': '精整',
            'shift': '长白班',
            'machine': 'JZ-01',
            'process': '包装',
            'coil_no': '26D00001',
            'output_tons': 10.25,
            'energy_kwh': 1350.0,
            'source_file': str(report),
            'source_type': 'output_skill_json_lines',
        },
        {
            'business_date': '2026-06-13',
            'workshop': '拉矫',
            'shift': '小夜班',
            'machine': 'LJ-02',
            'process': '拉矫',
            'coil_no': '26D00002',
            'output_tons': 8.5,
            'yield_rate': 96.0,
            'source_file': str(report),
            'source_type': 'output_skill_json_lines',
        },
    ]


def test_build_system_mapping_rows_flattens_mes_process_records() -> None:
    engine = create_engine('sqlite:///:memory:')
    Base.metadata.create_all(engine, tables=RECONCILIATION_TABLES)

    with Session(engine) as db:
        db.add_all(
            [
                MesWorkshopProcessRecord(
                    source_id='mes-1',
                    source_path='ProcessRecord',
                    business_date=date(2026, 6, 13),
                    workshop_name='精整',
                    process_name='包装',
                    device_name='PC-01',
                    batch_no='26A04967',
                    output_weight_tons=12.5,
                    input_weight_tons=13.0,
                    yield_rate=96.15,
                ),
                MesWorkshopProcessRecord(
                    source_id='mes-2',
                    source_path='ProcessRecord',
                    business_date=date(2026, 6, 13),
                    workshop_name='精整',
                    process_name='包装',
                    device_name='PC-01',
                    batch_no='26A04968',
                    output_weight_kg=5000,
                    input_weight_kg=5200,
                ),
                MesWorkshopProcessRecord(
                    source_id='mes-other-date',
                    source_path='ProcessRecord',
                    business_date=date(2026, 6, 14),
                    workshop_name='精整',
                    process_name='包装',
                    device_name='PC-01',
                    output_weight_tons=99,
                ),
            ]
        )
        db.commit()

        rows = build_system_mapping_rows(db, business_date=date(2026, 6, 13))

    assert rows == [
        {
            'business_date': '2026-06-13',
            'workshop': '精整',
            'shift': '',
            'process': '包装',
            'machine': 'PC-01',
            'coil_no': '26A04967',
            'input_tons': 13.0,
            'output_tons': 12.5,
            'yield_rate': 96.15,
            'source_table': 'mes_workshop_process_records',
        },
        {
            'business_date': '2026-06-13',
            'workshop': '精整',
            'shift': '',
            'process': '包装',
            'machine': 'PC-01',
            'coil_no': '26A04968',
            'input_tons': 5.2,
            'output_tons': 5.0,
            'yield_rate': None,
            'source_table': 'mes_workshop_process_records',
        },
    ]


def test_build_system_mapping_rows_flattens_stock_energy_and_consumables() -> None:
    engine = create_engine('sqlite:///:memory:')
    Base.metadata.create_all(engine, tables=RECONCILIATION_TABLES)

    with Session(engine) as db:
        db.add_all(
            [
                Workshop(id=1, code='JZ', name='精整车间', workshop_type='finishing'),
                Team(id=1, workshop_id=1, code='A', name='甲班'),
                User(id=1, username='owner', password_hash='x', name='内勤', role='consumable_stat'),
                ShiftConfig(
                    id=1,
                    code='A',
                    name='长白班',
                    shift_type='day',
                    start_time=time(7, 30),
                    end_time=time(15, 30),
                ),
                Equipment(id=1, code='JZ-01', name='精整1#机', workshop_id=1),
                MesStockRecord(
                    source_id='stock-1',
                    source_path='StockRecord',
                    business_date=date(2026, 6, 13),
                    batch_no='26A05000',
                    contract_no='HT-001',
                    customer_alias='客户A',
                    net_weight_tons=7.25,
                    status_name='已入库',
                ),
                MobileShiftReport(
                    id=1,
                    business_date=date(2026, 6, 13),
                    shift_config_id=1,
                    workshop_id=1,
                    team_id=1,
                    report_status='submitted',
                ),
                MachineEnergyRecord(
                    shift_report_id=1,
                    machine_id=1,
                    machine_code='JZ-01',
                    machine_name='精整1#机',
                    energy_kwh=1800,
                    gas_m3=20,
                ),
                DailyConsumableLog(
                    workshop_id=1,
                    workshop_type='finishing',
                    business_date=date(2026, 6, 13),
                    payload={
                        'packaging_inbound_output_tons': 11.5,
                        'rolling_oil_per_ton': 1.25,
                        'electricity_daily': 2000,
                        'gas_daily': 32,
                    },
                    created_by_user_id=1,
                ),
            ]
        )
        db.commit()

        rows = build_system_mapping_rows(db, business_date=date(2026, 6, 13))

    assert {
        'business_date': '2026-06-13',
        'workshop': '成品库',
        'shift': '',
        'process': '入库',
        'machine': '',
        'coil_no': '26A05000',
        'contract_no': 'HT-001',
        'customer': '客户A',
        'output_tons': 7.25,
        'status': '已入库',
        'source_table': 'mes_stock_records',
    } in rows
    assert {
        'business_date': '2026-06-13',
        'workshop': '精整车间',
        'shift': '长白班',
        'process': '能耗',
        'machine': '精整1#机',
        'machine_code': 'JZ-01',
        'energy_kwh': 1800.0,
        'gas_m3': 20.0,
        'source_table': 'machine_energy_records',
    } in rows
    assert {
        'business_date': '2026-06-13',
        'workshop': '精整车间',
        'shift': '',
        'process': '内勤辅材',
        'machine': '',
        'output_tons': 11.5,
        'energy_kwh': 2000.0,
        'gas_m3': 32.0,
        'rolling_oil_per_ton': 1.25,
        'consumable_payload': {'rolling_oil_per_ton': 1.25},
        'source_table': 'daily_consumable_logs',
    } in rows


def test_build_system_mapping_rows_flattens_owner_daily_contract_payload() -> None:
    engine = create_engine('sqlite:///:memory:')
    Base.metadata.create_all(engine, tables=RECONCILIATION_TABLES)

    with Session(engine) as db:
        db.add_all(
            [
                Workshop(id=1, code='CPK', name='成品库', workshop_type='inventory'),
                WorkOrder(
                    id=1,
                    tracking_card_no='OWNER-DAILY-20260613',
                    process_route_code='owner_daily',
                    contract_no='HT-PLAN',
                    customer_name='计划科',
                    overall_status='created',
                ),
                WorkOrderEntry(
                    id=1,
                    work_order_id=1,
                    workshop_id=1,
                    business_date=date(2026, 6, 13),
                    entry_type='owner_daily',
                    entry_status='submitted',
                    extra_payload={
                        'daily_contract_weight': 143,
                        'month_to_date_contract_weight': 2422,
                        'remaining_contract_weight': 400,
                        'remaining_hot_roll_contract_weight': 125,
                        'remaining_contract_delta_weight': 77,
                        'billet_inventory_weight': 120.5,
                        'daily_input_weight': 70,
                        'month_to_date_input_weight': 905,
                        'total_electricity_kwh': 131500,
                        'new_plant_electricity_kwh': 6124,
                        'park_electricity_kwh': 9056,
                        'cast_roll_gas_m3': 2599,
                        'smelting_gas_m3': 5348,
                        'heating_furnace_gas_m3': 876,
                        'boiler_gas_m3': 1087,
                        'total_gas_m3': 53433,
                        'groundwater_ton': 3860,
                        'tap_water_ton': 4295,
                    },
                ),
            ]
        )
        db.commit()

        rows = build_system_mapping_rows(db, business_date=date(2026, 6, 13))

    assert {
        'business_date': '2026-06-13',
        'workshop': '成品库',
        'shift': '',
        'process': '每日一录',
        'machine': '',
        'machine_code': '',
        'contract_no': 'HT-PLAN',
        'customer': '计划科',
        'daily_contract_weight': 143.0,
        'month_to_date_contract_weight': 2422.0,
        'remaining_contract_weight': 400.0,
        'remaining_hot_roll_contract_weight': 125.0,
        'remaining_contract_delta_weight': 77.0,
        'billet_inventory_weight': 120.5,
        'daily_input_weight': 70.0,
        'month_to_date_input_weight': 905.0,
        'total_electricity_kwh': 131500.0,
        'new_plant_electricity_kwh': 6124.0,
        'park_electricity_kwh': 9056.0,
        'cast_roll_gas_m3': 2599.0,
        'smelting_gas_m3': 5348.0,
        'heating_furnace_gas_m3': 876.0,
        'boiler_gas_m3': 1087.0,
        'total_gas_m3': 53433.0,
        'groundwater_ton': 3860.0,
        'tap_water_ton': 4295.0,
        'owner_daily_payload': {
            'daily_contract_weight': 143.0,
            'month_to_date_contract_weight': 2422.0,
            'remaining_contract_weight': 400.0,
            'remaining_hot_roll_contract_weight': 125.0,
            'remaining_contract_delta_weight': 77.0,
            'billet_inventory_weight': 120.5,
            'daily_input_weight': 70.0,
            'month_to_date_input_weight': 905.0,
            'total_electricity_kwh': 131500.0,
            'new_plant_electricity_kwh': 6124.0,
            'park_electricity_kwh': 9056.0,
            'cast_roll_gas_m3': 2599.0,
            'smelting_gas_m3': 5348.0,
            'heating_furnace_gas_m3': 876.0,
            'boiler_gas_m3': 1087.0,
            'total_gas_m3': 53433.0,
            'groundwater_ton': 3860.0,
            'tap_water_ton': 4295.0,
        },
        'source_table': 'work_order_entries',
    } in rows


def test_build_system_mapping_rows_flattens_wip_with_total_snapshot_fallback() -> None:
    engine = create_engine('sqlite:///:memory:')
    Base.metadata.create_all(engine, tables=RECONCILIATION_TABLES)

    with Session(engine) as db:
        db.add_all(
            [
                MesDailyWipSnapshot(
                    business_date=date(2026, 6, 16),
                    workshop_name='新厂在线车间',
                    process_name='北线退火',
                    coil_count=3,
                    material_weight_tons=None,
                    feeding_weight_tons=28.5,
                    source='mes_coil_snapshot',
                ),
                MesWipTotalSnapshot(
                    source_id='新厂在线车间:北线退火',
                    workshop_name='新厂在线车间',
                    process_name='北线退火',
                    doing_count=588,
                    doing_weight_tons=4466.5,
                    snapshot_at=datetime(2026, 6, 16, 8, 0, tzinfo=UTC),
                ),
            ]
        )
        db.commit()

        rows = build_system_mapping_rows(db, business_date=date(2026, 6, 16))

    assert {
        'business_date': '2026-06-16',
        'workshop': '新厂在线车间',
        'shift': '',
        'process': '在制料',
        'machine': '',
        'wip_total': 4466.5,
        'wip_coil_count': 588,
        'wip_feeding_tons': 28.5,
        'source_table': 'mes_wip_total_snapshots',
    } in rows
    assert {
        'business_date': '2026-06-16',
        'workshop': '全厂',
        'shift': '',
        'process': '汇总',
        'machine': '',
        'wip_total': 4466.5,
        'source_table': 'mapping_reconciliation_summary',
    } in rows


def test_build_system_mapping_rows_flattens_cost_daily_results() -> None:
    engine = create_engine('sqlite:///:memory:')
    Base.metadata.create_all(engine, tables=RECONCILIATION_TABLES)

    with Session(engine) as db:
        db.add(Workshop(id=1, code='JZ', name='精整车间', workshop_type='finishing'))
        db.add_all(
            [
                CostDailyResult(
                    business_date=date(2026, 6, 13),
                    workshop_code='JZ',
                    strategy_code='finishing_parallel_process',
                    total_cost=12800.5,
                    output_ton_cost=331.16,
                    throughput_ton_cost=280.25,
                    caliber='output',
                    breakdown_count=4,
                    process_count=2,
                ),
                CostDailyResult(
                    business_date=date(2026, 6, 14),
                    workshop_code='JZ',
                    strategy_code='finishing_parallel_process',
                    total_cost=99999,
                    output_ton_cost=999,
                    throughput_ton_cost=999,
                    caliber='output',
                ),
            ]
        )
        db.commit()

        rows = build_system_mapping_rows(db, business_date=date(2026, 6, 13))

    assert {
        'business_date': '2026-06-13',
        'workshop': '精整车间',
        'shift': '',
        'process': '成本策略',
        'machine': '',
        'strategy_code': 'finishing_parallel_process',
        'cost_caliber': 'output',
        'total_cost': 12800.5,
        'cost_per_ton': 331.16,
        'throughput_cost_per_ton': 280.25,
        'breakdown_count': 4,
        'process_count': 2,
        'source_table': 'cost_daily_result',
    } in rows


def test_build_system_mapping_rows_flattens_machine_daily_cost_snapshots() -> None:
    engine = create_engine('sqlite:///:memory:')
    Base.metadata.create_all(engine, tables=RECONCILIATION_TABLES)

    with Session(engine) as db:
        db.add(Workshop(id=1, code='QC', name='全厂', workshop_type='factory'))
        db.add_all(
            [
                MachineDailyCostSnapshot(
                    business_date=date(2026, 6, 14),
                    workshop_id=1,
                    machine_line_id=None,
                    electricity_kwh=126500,
                    electricity_cost=101200,
                    natural_gas_m3=59141,
                    natural_gas_cost=212900,
                    total_cost=314100,
                    is_estimated=True,
                ),
                MachineDailyCostSnapshot(
                    business_date=date(2026, 6, 13),
                    workshop_id=1,
                    machine_line_id=None,
                    electricity_cost=999,
                    natural_gas_cost=999,
                    total_cost=1998,
                ),
            ]
        )
        db.commit()

        rows = build_system_mapping_rows(db, business_date=date(2026, 6, 14))

    assert {
        'business_date': '2026-06-14',
        'workshop': '全厂',
        'shift': '',
        'process': '机列日成本',
        'machine': '',
        'machine_code': '',
        'electricity_kwh': 126500.0,
        'electricity_cost': 101200.0,
        'natural_gas_m3': 59141.0,
        'natural_gas_cost': 212900.0,
        'total_cost': 314100.0,
        'is_estimated': True,
        'source_table': 'machine_daily_cost_snapshots',
    } in rows


def test_build_system_mapping_rows_flattens_shift_production_scrap() -> None:
    engine = create_engine('sqlite:///:memory:')
    Base.metadata.create_all(engine, tables=RECONCILIATION_TABLES)

    with Session(engine) as db:
        db.add_all(
            [
                Workshop(id=1, code='LJ', name='拉矫车间', workshop_type='finishing'),
                ShiftConfig(
                    id=1,
                    code='B',
                    name='小夜班',
                    shift_type='evening',
                    start_time=time(15, 30),
                    end_time=time(23, 30),
                ),
                Equipment(id=1, code='LJ-01', name='拉矫1#机', workshop_id=1),
                ShiftProductionData(
                    business_date=date(2026, 6, 13),
                    shift_config_id=1,
                    workshop_id=1,
                    equipment_id=1,
                    input_weight=10.0,
                    output_weight=9.5,
                    scrap_weight=0.3,
                    downtime_minutes=18,
                    issue_count=2,
                    electricity_kwh=1200,
                    data_source='mobile',
                    data_status='confirmed',
                ),
            ]
        )
        db.commit()

        rows = build_system_mapping_rows(db, business_date=date(2026, 6, 13))

    assert {
        'business_date': '2026-06-13',
        'workshop': '拉矫车间',
        'shift': '小夜班',
        'process': '班次产量',
        'machine': '拉矫1#机',
        'machine_code': 'LJ-01',
        'input_tons': 10.0,
        'output_tons': 9.5,
        'scrap_tons': 0.3,
        'downtime_minutes': 18.0,
        'quality_issue_count': 2.0,
        'energy_kwh': 1200.0,
        'source_table': 'shift_production_data',
    } in rows
