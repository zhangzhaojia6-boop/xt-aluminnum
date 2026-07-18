from __future__ import annotations

from datetime import date, datetime
import hashlib
from importlib.util import module_from_spec, spec_from_file_location
from io import BytesIO, StringIO
import json
import sys
from xml.etree import ElementTree
from zipfile import ZipFile

from openpyxl import Workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from app.models import Base
from app.models.agent_communication import MultimodalEvidence
from app.services import dingtalk_stream_gateway_service as gateway_service
from app.services.report.daily_fact_bundle import build_daily_fact_bundle
from tests.path_helpers import BACKEND_ROOT


SCRIPT_PATH = BACKEND_ROOT / 'scripts' / 'dingtalk_real_fact_backfill.py'


def _load_script_module():
    spec = spec_from_file_location('dingtalk_real_fact_backfill_script', SCRIPT_PATH)
    assert spec is not None and spec.loader is not None
    module = module_from_spec(spec)
    sys.modules[spec.name] = module
    spec.loader.exec_module(module)
    return module


def _db_sessionmaker():
    engine = create_engine('sqlite:///:memory:', future=True)
    Base.metadata.create_all(bind=engine)
    return sessionmaker(bind=engine, future=True)


def _write_jsonl(path, rows: list[dict]) -> None:
    path.write_text('\n'.join(json.dumps(row, ensure_ascii=False) for row in rows), encoding='utf-8')


def _sha256_text(value: str) -> str:
    return hashlib.sha256(value.encode('utf-8')).hexdigest()


def _set_workbook_modified(content: bytes, modified: datetime) -> bytes:
    output = BytesIO()
    with ZipFile(BytesIO(content), 'r') as source, ZipFile(output, 'w') as target:
        for item in source.infolist():
            value = source.read(item.filename)
            if item.filename == 'docProps/core.xml':
                root = ElementTree.fromstring(value)
                node = root.find('{http://purl.org/dc/terms/}modified')
                assert node is not None
                node.text = modified.isoformat(timespec='seconds') + 'Z'
                value = ElementTree.tostring(root, encoding='utf-8', xml_declaration=True)
            target.writestr(item, value)
    return output.getvalue()


def _owner_verified_text_row(
    *,
    text: str = '今日日报：产量 32 吨',
    message_id: str = 'msg-owner-verified-text-001',
) -> dict:
    return {
        'conversationId': 'group-001',
        'message_id': message_id,
        'senderStaffId': 'sender-001',
        'createTime': '2026-07-07T17:46:26+08:00',
        'message_text': text,
        'businessDate': '2026-07-07',
        'content_sha256': _sha256_text(text),
    }


WIP_SCREENSHOT_CONTENT = b'\xff\xd8\xffowner-verified-wip\xff\xd9'


def _owner_verified_visual_row(*, message_id: str = 'msg-owner-verified-wip-001') -> dict:
    return {
        'conversationId': 'group-001',
        'message_id': message_id,
        'senderStaffId': 'sender-wip-owner',
        'createTime': '2026-07-08T08:19:42+08:00',
        'fileName': 'wip.jpg',
        'fileId': f'dws-media-{message_id}',
        'localFilePath': 'wip.jpg',
        'businessDate': '2026-07-07',
        'content_sha256': hashlib.sha256(WIP_SCREENSHOT_CONTENT).hexdigest(),
        'owner_verified_visual_facts': {
            'wip_total': {
                'value': 1877,
                'unit': '吨',
                'reported_date': '2026-07-08',
                'row_label': '汇总',
                'column_label': '在制料',
            }
        },
    }


def _owner_verified_full_visual_row() -> dict:
    values = {
        'wip_total': 1166.5,
        'wip_1650_2050_cold': 345.5,
        'wip_1850_cold': 103.5,
        'wip_milling': 0,
        'wip_anneal_total': 323.5,
        'wip_new_north': 231,
        'wip_new_south': 37,
        'wip_park_anneal': 55.5,
        'wip_finishing_total': 394,
        'wip_straightening': 290,
        'wip_finishing': 99.5,
        'wip_park_finishing': 4.5,
        'wip_hot_plate_shearing': 0,
        'wip_coating': 0,
    }
    row_labels = {
        'wip_total': '汇总',
        'wip_1650_2050_cold': '1650+2050冷轧合计',
        'wip_1850_cold': '1850冷轧合计',
        'wip_milling': '铣床合计',
        'wip_anneal_total': '在线退火合计',
        'wip_new_north': '新厂北',
        'wip_new_south': '新厂南',
        'wip_park_anneal': '园区退火',
        'wip_finishing_total': '后工序合计',
        'wip_straightening': '拉矫合计',
        'wip_finishing': '精整合计',
        'wip_park_finishing': '园区剪切',
        'wip_hot_plate_shearing': '热轧（中厚板）',
        'wip_coating': '彩涂',
    }
    row = _owner_verified_visual_row(message_id='msg-owner-verified-wip-full')
    row['owner_verified_visual_facts'] = {
        field: {
            'value': value,
            'unit': '吨',
            'reported_date': '2026-07-08',
            'row_label': row_labels[field],
            'column_label': '在制料',
        }
        for field, value in values.items()
    }
    return row


def _prepare(monkeypatch):
    module = _load_script_module()
    Session = _db_sessionmaker()
    monkeypatch.setattr(module, 'get_sessionmaker', lambda: Session)
    monkeypatch.setattr(module, 'last_completed_production_business_date', lambda: date(2026, 7, 8))
    monkeypatch.setattr(gateway_service.settings, 'DINGTALK_AUTHORIZED_GROUP_IDS', 'group-001', raising=False)
    monkeypatch.setattr(gateway_service.settings, 'DINGTALK_FILE_TEXT_MAX_BYTES', 1_000_000, raising=False)
    return module, Session


def test_jsonl_text_row_writes_message_text(monkeypatch, tmp_path) -> None:
    module, Session = _prepare(monkeypatch)
    files_root = tmp_path / 'files'
    files_root.mkdir()
    input_jsonl = tmp_path / 'messages.jsonl'
    _write_jsonl(
        input_jsonl,
        [
            {
                'group_id': 'group-001',
                'message_id': 'msg-text-001',
                'message_text': '今日日报：产量 32 吨',
                'businessDate': '2026-07-07',
            }
        ],
    )
    captured = StringIO()

    exit_code = module.main(
        ['--input-jsonl', str(input_jsonl), '--files-root', str(files_root), '--days', '3'],
        stdout=captured,
    )

    assert exit_code == 0
    assert json.loads(captured.getvalue()) == {
        'accepted': 1,
        'duplicates': 0,
        'rejected': 0,
        'file_text': 0,
        'message_text': 1,
    }
    db = Session()
    try:
        evidence = db.query(MultimodalEvidence).one()
        assert evidence.payload['message_text'] == '今日日报：产量 32 吨'
        assert evidence.payload['source_transport'] == 'dws_history_backfill'
        assert evidence.confirmation_status == 'machine_only'
    finally:
        db.close()


def test_jsonl_file_row_writes_file_text(monkeypatch, tmp_path) -> None:
    module, Session = _prepare(monkeypatch)
    files_root = tmp_path / 'files'
    files_root.mkdir()
    (files_root / 'report.csv').write_text('日期,产量\n2026-07-07,32\n', encoding='utf-8')
    input_jsonl = tmp_path / 'messages.jsonl'
    _write_jsonl(
        input_jsonl,
        [
            {
                'group_id': 'group-001',
                'message_id': 'msg-file-001',
                'fileName': 'report.csv',
                'localFilePath': 'report.csv',
                'businessDate': '2026-07-07',
            }
        ],
    )
    captured = StringIO()

    exit_code = module.main(
        ['--input-jsonl', str(input_jsonl), '--files-root', str(files_root), '--days', '3'],
        stdout=captured,
    )

    assert exit_code == 0
    summary = json.loads(captured.getvalue())
    assert summary['accepted'] == 1
    assert summary['file_text'] == 1
    db = Session()
    try:
        evidence = db.query(MultimodalEvidence).one()
        assert evidence.payload['file_text'] == '日期\t产量\n2026-07-07\t32'
        assert evidence.payload['download_url_host'] == 'local-backfill'
    finally:
        db.close()


def test_path_traversal_is_rejected(monkeypatch, tmp_path) -> None:
    module, Session = _prepare(monkeypatch)
    files_root = tmp_path / 'files'
    files_root.mkdir()
    (tmp_path / 'secret.csv').write_text('不该读取', encoding='utf-8')
    input_jsonl = tmp_path / 'messages.jsonl'
    _write_jsonl(
        input_jsonl,
        [
            {
                'group_id': 'group-001',
                'message_id': 'msg-file-escape',
                'fileName': 'secret.csv',
                'localFilePath': '../secret.csv',
                'businessDate': '2026-07-07',
            }
        ],
    )
    captured = StringIO()

    exit_code = module.main(
        ['--input-jsonl', str(input_jsonl), '--files-root', str(files_root), '--days', '3'],
        stdout=captured,
    )

    assert exit_code == 0
    assert json.loads(captured.getvalue())['rejected'] == 1
    db = Session()
    try:
        assert db.query(MultimodalEvidence).count() == 0
    finally:
        db.close()


def test_unauthorized_group_is_rejected(monkeypatch, tmp_path) -> None:
    module, Session = _prepare(monkeypatch)
    files_root = tmp_path / 'files'
    files_root.mkdir()
    input_jsonl = tmp_path / 'messages.jsonl'
    _write_jsonl(
        input_jsonl,
        [
            {
                'group_id': 'group-999',
                'message_id': 'msg-unauthorized',
                'message_text': '今日日报：产量 32 吨',
                'businessDate': '2026-07-07',
            }
        ],
    )
    captured = StringIO()

    exit_code = module.main(
        ['--input-jsonl', str(input_jsonl), '--files-root', str(files_root), '--days', '3'],
        stdout=captured,
    )

    assert exit_code == 0
    assert json.loads(captured.getvalue())['rejected'] == 1
    db = Session()
    try:
        assert db.query(MultimodalEvidence).count() == 0
    finally:
        db.close()


def test_duplicate_file_is_counted_as_duplicate(monkeypatch, tmp_path) -> None:
    module, Session = _prepare(monkeypatch)
    files_root = tmp_path / 'files'
    files_root.mkdir()
    (files_root / 'report.csv').write_text('日期,产量\n2026-07-07,32\n', encoding='utf-8')
    input_jsonl = tmp_path / 'messages.jsonl'
    row = {
        'group_id': 'group-001',
        'message_id': 'msg-file-duplicate',
        'fileName': 'report.csv',
        'localFilePath': 'report.csv',
        'businessDate': '2026-07-07',
    }
    _write_jsonl(input_jsonl, [row, row])
    captured = StringIO()

    exit_code = module.main(
        ['--input-jsonl', str(input_jsonl), '--files-root', str(files_root), '--days', '3'],
        stdout=captured,
    )

    assert exit_code == 0
    summary = json.loads(captured.getvalue())
    assert summary['accepted'] == 1
    assert summary['duplicates'] == 1
    assert summary['file_text'] == 1
    db = Session()
    try:
        assert db.query(MultimodalEvidence).count() == 1
    finally:
        db.close()


def test_owner_verified_dws_text_is_confirmed_only_with_complete_lineage(monkeypatch, tmp_path) -> None:
    module, Session = _prepare(monkeypatch)
    files_root = tmp_path / 'files'
    files_root.mkdir()
    input_jsonl = tmp_path / 'messages.jsonl'
    row = _owner_verified_text_row()
    _write_jsonl(input_jsonl, [row])
    captured = StringIO()

    exit_code = module.main(
        [
            '--input-jsonl',
            str(input_jsonl),
            '--files-root',
            str(files_root),
            '--days',
            '3',
            '--confirmation-mode',
            'owner-verified-dws-history',
            '--confirmation-run-id',
            'test-run-001',
        ],
        stdout=captured,
    )

    assert exit_code == 0
    summary = json.loads(captured.getvalue())
    assert summary['confirmed'] == 1
    assert summary['confirmation_rejected'] == 0
    db = Session()
    try:
        evidence = db.query(MultimodalEvidence).one()
        assert evidence.confirmation_status == 'confirmed'
        assert evidence.payload['source_transport'] == 'dws_history_backfill'
        assert evidence.payload['messageId'] == row['message_id']
        assert evidence.payload['business_date'] == row['businessDate']
        assert evidence.payload['owner_verification'] == {
            'mode': 'owner_verified_dws_history',
            'run_id': 'test-run-001',
            'content_sha256': row['content_sha256'],
        }
    finally:
        db.close()


def test_owner_verified_plan_contract_text_adds_traceable_fact_updates(monkeypatch, tmp_path) -> None:
    module, Session = _prepare(monkeypatch)
    files_root = tmp_path / 'files'
    files_root.mkdir()
    input_jsonl = tmp_path / 'messages.jsonl'
    text = (
        '投料量：2050投料463吨 1850投料0吨 外加工62吨 中厚板0吨 '
        '当天合同443吨 热轧436吨 总余合同量2765吨'
    )
    row = _owner_verified_text_row(text=text, message_id='msg-plan-contract-001')
    _write_jsonl(input_jsonl, [row])

    summary = module.run_backfill(
        input_jsonl=input_jsonl,
        files_root=files_root,
        days=3,
        confirmation_mode='owner-verified-dws-history',
        confirmation_run_id='test-run-plan-contract',
    )

    assert summary['confirmed'] == 1
    assert summary['committed'] == 1
    db = Session()
    try:
        evidence = db.query(MultimodalEvidence).one()
        updates = evidence.payload['fact_updates']
        assert updates['daily_input_weight']['value'] == 525
        assert updates['remaining_contract_weight']['value'] == 2765
        assert updates['remaining_contract_weight']['source_ref']['parser'] == 'plan_contract_message_v1'
        assert updates['remaining_contract_weight']['source_ref']['content_sha256'] == row[
            'content_sha256'
        ]
    finally:
        db.close()


def test_owner_verified_dws_row_with_wrong_hash_is_rejected_before_persistence(monkeypatch, tmp_path) -> None:
    module, Session = _prepare(monkeypatch)
    files_root = tmp_path / 'files'
    files_root.mkdir()
    input_jsonl = tmp_path / 'messages.jsonl'
    row = _owner_verified_text_row()
    row['content_sha256'] = '0' * 64
    _write_jsonl(input_jsonl, [row])
    captured = StringIO()

    exit_code = module.main(
        [
            '--input-jsonl',
            str(input_jsonl),
            '--files-root',
            str(files_root),
            '--confirmation-mode',
            'owner-verified-dws-history',
            '--confirmation-run-id',
            'test-run-002',
        ],
        stdout=captured,
    )

    assert exit_code == 0
    summary = json.loads(captured.getvalue())
    assert summary['accepted'] == 0
    assert summary['rejected'] == 1
    assert summary['confirmation_rejected'] == 1
    assert summary['rejection_reasons'] == {'owner_verified_content_hash_mismatch': 1}
    db = Session()
    try:
        assert db.query(MultimodalEvidence).count() == 0
    finally:
        db.close()


def test_owner_verified_dws_row_missing_sender_is_rejected_before_persistence(monkeypatch, tmp_path) -> None:
    module, Session = _prepare(monkeypatch)
    files_root = tmp_path / 'files'
    files_root.mkdir()
    input_jsonl = tmp_path / 'messages.jsonl'
    row = _owner_verified_text_row()
    row.pop('senderStaffId')
    _write_jsonl(input_jsonl, [row])
    captured = StringIO()

    exit_code = module.main(
        [
            '--input-jsonl',
            str(input_jsonl),
            '--files-root',
            str(files_root),
            '--confirmation-mode',
            'owner-verified-dws-history',
            '--confirmation-run-id',
            'test-run-003',
        ],
        stdout=captured,
    )

    assert exit_code == 0
    summary = json.loads(captured.getvalue())
    assert summary['confirmation_rejected'] == 1
    assert summary['rejection_reasons'] == {'owner_verified_lineage_incomplete': 1}
    db = Session()
    try:
        assert db.query(MultimodalEvidence).count() == 0
    finally:
        db.close()


def test_owner_verified_dws_package_is_atomic_when_any_row_fails(monkeypatch, tmp_path) -> None:
    module, Session = _prepare(monkeypatch)
    files_root = tmp_path / 'files'
    files_root.mkdir()
    input_jsonl = tmp_path / 'messages.jsonl'
    valid = _owner_verified_text_row(message_id='msg-valid')
    invalid = _owner_verified_text_row(message_id='msg-invalid')
    invalid['content_sha256'] = '0' * 64
    _write_jsonl(input_jsonl, [valid, invalid])
    captured = StringIO()

    exit_code = module.main(
        [
            '--input-jsonl',
            str(input_jsonl),
            '--files-root',
            str(files_root),
            '--confirmation-mode',
            'owner-verified-dws-history',
            '--confirmation-run-id',
            'test-run-atomic',
        ],
        stdout=captured,
    )

    assert exit_code == 0
    summary = json.loads(captured.getvalue())
    assert summary['committed'] == 0
    assert summary['confirmed'] == 0
    assert summary['confirmation_rejected'] == 1
    assert summary['rejection_reasons'] == {'owner_verified_content_hash_mismatch': 1}
    db = Session()
    try:
        assert db.query(MultimodalEvidence).count() == 0
    finally:
        db.close()


def test_owner_verified_dws_retry_preserves_first_confirmation_audit(monkeypatch, tmp_path) -> None:
    module, Session = _prepare(monkeypatch)
    files_root = tmp_path / 'files'
    files_root.mkdir()
    input_jsonl = tmp_path / 'messages.jsonl'
    row = _owner_verified_text_row()
    _write_jsonl(input_jsonl, [row])

    first = module.run_backfill(
        input_jsonl=input_jsonl,
        files_root=files_root,
        days=3,
        confirmation_mode='owner-verified-dws-history',
        confirmation_run_id='test-run-first',
    )
    second = module.run_backfill(
        input_jsonl=input_jsonl,
        files_root=files_root,
        days=3,
        confirmation_mode='owner-verified-dws-history',
        confirmation_run_id='test-run-retry',
    )

    assert first['confirmed'] == 1
    assert first['committed'] == 1
    assert second['confirmed'] == 0
    assert second['already_confirmed'] == 1
    assert second['committed'] == 1
    db = Session()
    try:
        evidence = db.query(MultimodalEvidence).one()
        assert evidence.payload['owner_verification']['run_id'] == 'test-run-first'
    finally:
        db.close()


def test_owner_verified_dws_upgrades_machine_only_equivalent_event_time(monkeypatch, tmp_path) -> None:
    module, Session = _prepare(monkeypatch)
    files_root = tmp_path / 'files'
    files_root.mkdir()
    input_jsonl = tmp_path / 'messages.jsonl'
    row = _owner_verified_text_row()
    row['createTime'] = '2026-07-07 17:46:26'
    _write_jsonl(input_jsonl, [row])

    machine_only = module.run_backfill(
        input_jsonl=input_jsonl,
        files_root=files_root,
        days=3,
    )
    row['createTime'] = '2026-07-07T17:46:26+08:00'
    _write_jsonl(input_jsonl, [row])
    owner_verified = module.run_backfill(
        input_jsonl=input_jsonl,
        files_root=files_root,
        days=3,
        confirmation_mode='owner-verified-dws-history',
        confirmation_run_id='test-run-equivalent-event-time',
    )

    assert machine_only['accepted'] == 1
    assert owner_verified['duplicates'] == 1
    assert owner_verified['confirmed'] == 1
    assert owner_verified['confirmation_rejected'] == 0
    assert owner_verified['committed'] == 1
    db = Session()
    try:
        evidence = db.query(MultimodalEvidence).one()
        assert evidence.confirmation_status == 'confirmed'
        assert evidence.payload['event_time'] == '2026-07-07 17:46:26'
    finally:
        db.close()


def test_owner_verified_dws_rejects_different_event_time(monkeypatch, tmp_path) -> None:
    module, Session = _prepare(monkeypatch)
    files_root = tmp_path / 'files'
    files_root.mkdir()
    input_jsonl = tmp_path / 'messages.jsonl'
    row = _owner_verified_text_row()
    row['createTime'] = '2026-07-07 17:46:26'
    _write_jsonl(input_jsonl, [row])
    module.run_backfill(input_jsonl=input_jsonl, files_root=files_root, days=3)

    row['createTime'] = '2026-07-07T17:46:27+08:00'
    _write_jsonl(input_jsonl, [row])
    owner_verified = module.run_backfill(
        input_jsonl=input_jsonl,
        files_root=files_root,
        days=3,
        confirmation_mode='owner-verified-dws-history',
        confirmation_run_id='test-run-different-event-time',
    )

    assert owner_verified['confirmed'] == 0
    assert owner_verified['confirmation_rejected'] == 1
    assert owner_verified['committed'] == 0
    assert owner_verified['rejection_reasons'] == {'owner_verification_event_time_mismatch': 1}
    db = Session()
    try:
        evidence = db.query(MultimodalEvidence).one()
        assert evidence.confirmation_status != 'confirmed'
        assert 'owner_verification' not in evidence.payload
    finally:
        db.close()


def test_owner_verified_dws_reports_sender_identity_type_mismatch(monkeypatch, tmp_path) -> None:
    module, _Session = _prepare(monkeypatch)
    files_root = tmp_path / 'files'
    files_root.mkdir()
    input_jsonl = tmp_path / 'messages.jsonl'
    row = _owner_verified_text_row()
    _write_jsonl(input_jsonl, [row])
    module.run_backfill(input_jsonl=input_jsonl, files_root=files_root, days=3)

    row.pop('senderStaffId')
    row['senderOpenDingTalkId'] = 'sender-001'
    _write_jsonl(input_jsonl, [row])
    owner_verified = module.run_backfill(
        input_jsonl=input_jsonl,
        files_root=files_root,
        days=3,
        confirmation_mode='owner-verified-dws-history',
        confirmation_run_id='test-run-sender-identity-type-mismatch',
    )

    assert owner_verified['confirmed'] == 0
    assert owner_verified['confirmation_rejected'] == 1
    assert owner_verified['committed'] == 0
    assert owner_verified['rejection_reasons'] == {'owner_verification_sender_identity_type_mismatch': 1}


def test_owner_verified_dws_accepts_open_dingtalk_sender_identity(monkeypatch, tmp_path) -> None:
    module, Session = _prepare(monkeypatch)
    files_root = tmp_path / 'files'
    files_root.mkdir()
    input_jsonl = tmp_path / 'messages.jsonl'
    row = _owner_verified_text_row()
    row['senderOpenDingTalkId'] = row.pop('senderStaffId')
    _write_jsonl(input_jsonl, [row])

    summary = module.run_backfill(
        input_jsonl=input_jsonl,
        files_root=files_root,
        days=3,
        confirmation_mode='owner-verified-dws-history',
        confirmation_run_id='test-run-open-id',
    )

    assert summary['confirmed'] == 1
    assert summary['committed'] == 1
    db = Session()
    try:
        evidence = db.query(MultimodalEvidence).one()
        assert evidence.payload['dingtalk_sender_id'] == row['senderOpenDingTalkId']
        assert evidence.payload['sender_identity_type'] == 'open_dingtalk_id'
    finally:
        db.close()


def test_owner_verified_yield_workbook_adds_cell_traceable_fact_update(monkeypatch, tmp_path) -> None:
    from app.services.report import daily_fact_bundle

    module, Session = _prepare(monkeypatch)
    files_root = tmp_path / 'files'
    files_root.mkdir()
    workbook_path = files_root / 'yield.xlsx'
    workbook = Workbook()
    workbook.properties.modified = datetime(2026, 7, 12)
    sheet = workbook.active
    sheet.title = 'Sheet3'
    sheet['A1'] = '7月份各车间成品率车间指标'
    sheet['A2'] = '日期'
    sheet.merge_cells('A2:A5')
    sheet['AB2'] = '公司'
    sheet.merge_cells('AB2:AK2')
    sheet['AJ3'] = '总成品率92%'
    sheet.merge_cells('AJ3:AK4')
    sheet['AJ5'] = '日合计'
    sheet['AK5'] = '月累计'
    sheet['A6'] = 12
    sheet['AJ6'] = 0.838100137174211
    sheet['AK6'] = 0.851472683750808
    workbook.save(workbook_path)
    file_content = _set_workbook_modified(workbook_path.read_bytes(), datetime(2026, 7, 12))
    workbook_path.write_bytes(file_content)
    input_jsonl = tmp_path / 'messages.jsonl'
    _write_jsonl(
        input_jsonl,
        [
            {
                'conversationId': 'group-001',
                'message_id': 'msg-owner-verified-yield-001',
                'senderStaffId': 'sender-yield-owner',
                'createTime': '2026-07-13T07:41:00+08:00',
                'fileName': 'yield.xlsx',
                'fileId': 'dws-file-yield-001',
                'localFilePath': 'yield.xlsx',
                'businessDate': '2026-07-12',
                'content_sha256': hashlib.sha256(file_content).hexdigest(),
            }
        ],
    )
    captured = StringIO()

    exit_code = module.main(
        [
            '--input-jsonl',
            str(input_jsonl),
            '--files-root',
            str(files_root),
            '--confirmation-mode',
            'owner-verified-dws-history',
            '--confirmation-run-id',
            'test-run-yield-001',
        ],
        stdout=captured,
    )

    assert exit_code == 0
    assert json.loads(captured.getvalue())['confirmed'] == 1
    db = Session()
    try:
        evidence = db.query(MultimodalEvidence).one()
        update = evidence.payload['fact_updates']['daily_yield_rate']
        assert update['value'] == 83.81
        assert update['unit'] == '%'
        assert update['source_ref'] == {
            'parser': 'company_daily_yield_v1',
            'sheet_name': 'Sheet3',
            'workbook_modified_date': '2026-07-12',
            'date_cell': 'A6',
            'value_cell': 'AJ6',
            'header_cell': 'AJ3',
            'file_sha256': hashlib.sha256(file_content).hexdigest(),
        }
        monkeypatch.setattr(
            daily_fact_bundle.template_daily_report,
            'build_template_daily_report_facts',
            lambda db, *, target_date, wip_date=None: {
                'values': {'daily_yield_rate': 96.23},
                'sources': {'daily_yield_rate': 'computed'},
                'missing_fields': [],
                'conflicts': [],
            },
        )
        bundle = build_daily_fact_bundle(db, business_date=date(2026, 7, 12))
        assert bundle['facts']['daily_yield_rate']['value'] == 83.81
        assert bundle['facts']['daily_yield_rate']['source_type'] == 'dingtalk_supplement'
        assert bundle['facts']['daily_yield_rate']['source_ref']['evidence_id'] == evidence.id
        assert bundle['facts']['daily_yield_rate']['source_ref']['field_source_ref']['value_cell'] == 'AJ6'
        assert any(
            conflict.get('field') == 'daily_yield_rate'
            and conflict.get('previous_value') == 96.23
            and conflict.get('adopted_value') == 83.81
            for conflict in bundle['conflicts']
        )
    finally:
        db.close()


def test_owner_verified_wip_screenshot_adds_traceable_fact_update(monkeypatch, tmp_path) -> None:
    from app.services.report import daily_fact_bundle

    module, Session = _prepare(monkeypatch)
    files_root = tmp_path / 'files'
    files_root.mkdir()
    screenshot_path = files_root / 'wip.jpg'
    screenshot_content = WIP_SCREENSHOT_CONTENT
    screenshot_path.write_bytes(screenshot_content)
    input_jsonl = tmp_path / 'messages.jsonl'
    _write_jsonl(
        input_jsonl,
        [
            {
                'conversationId': 'group-001',
                'message_id': 'msg-owner-verified-wip-001',
                'senderStaffId': 'sender-wip-owner',
                'createTime': '2026-07-08T08:19:42+08:00',
                'fileName': 'wip.jpg',
                'fileId': 'dws-media-wip-001',
                'localFilePath': 'wip.jpg',
                'businessDate': '2026-07-07',
                'content_sha256': hashlib.sha256(screenshot_content).hexdigest(),
                'owner_verified_visual_facts': {
                    'wip_total': {
                        'value': 1877,
                        'unit': '吨',
                        'reported_date': '2026-07-08',
                        'row_label': '汇总',
                        'column_label': '在制料',
                    }
                },
            }
        ],
    )

    summary = module.run_backfill(
        input_jsonl=input_jsonl,
        files_root=files_root,
        days=3,
        confirmation_mode='owner-verified-dws-history',
        confirmation_run_id='test-run-wip-001',
    )

    assert summary['confirmed'] == 1
    assert summary['committed'] == 1
    db = Session()
    try:
        evidence = db.query(MultimodalEvidence).one()
        update = evidence.payload['fact_updates']['wip_total']
        assert update['value'] == 1877
        assert update['source_ref']['parser'] == 'owner_verified_wip_screenshot_v1'
        assert update['source_ref']['file_sha256'] == hashlib.sha256(screenshot_content).hexdigest()
        assert evidence.payload['parse_status'] == 'text_captured'
        assert evidence.payload['pre_confirmation_parse_status'] == 'unsupported_file_type'
        assert evidence.payload['attachment_text'] == '汇总/在制料：1877吨（报表日期 2026-07-08）'
        monkeypatch.setattr(
            daily_fact_bundle.template_daily_report,
            'build_template_daily_report_facts',
            lambda db, *, target_date, wip_date=None: {
                'values': {'wip_total': 1758.5},
                'sources': {'wip_total': 'mes_wip_distribution'},
                'missing_fields': [],
                'conflicts': [],
            },
        )
        bundle = build_daily_fact_bundle(
            db,
            business_date=date(2026, 7, 7),
            allow_output_skill_reference_adoption=False,
        )
        assert bundle['facts']['wip_total']['value'] == 1877
        assert bundle['facts']['wip_total']['source_type'] == 'dingtalk_supplement'
        assert bundle['facts']['wip_total']['source_ref']['evidence_id'] == evidence.id
        assert any(
            conflict.get('field') == 'wip_total'
            and conflict.get('previous_value') == 1758.5
            and conflict.get('adopted_value') == 1877
            for conflict in bundle['conflicts']
        )
    finally:
        db.close()


def test_owner_verified_full_wip_screenshot_adds_all_traceable_updates(monkeypatch, tmp_path) -> None:
    module, Session = _prepare(monkeypatch)
    files_root = tmp_path / 'files'
    files_root.mkdir()
    (files_root / 'wip.jpg').write_bytes(WIP_SCREENSHOT_CONTENT)
    input_jsonl = tmp_path / 'messages.jsonl'
    _write_jsonl(input_jsonl, [_owner_verified_full_visual_row()])

    summary = module.run_backfill(
        input_jsonl=input_jsonl,
        files_root=files_root,
        days=3,
        confirmation_mode='owner-verified-dws-history',
        confirmation_run_id='test-run-wip-full',
    )

    assert summary['confirmed'] == 1
    assert summary['committed'] == 1
    db = Session()
    try:
        evidence = db.query(MultimodalEvidence).one()
        updates = evidence.payload['fact_updates']
        assert len(updates) == 14
        assert updates['wip_total']['value'] == 1166.5
        assert updates['wip_finishing_total']['value'] == 394
        assert updates['wip_total']['source_ref']['parser'] == 'owner_verified_wip_screenshot_v2'
        assert '汇总/在制料：1166.5吨' in evidence.payload['attachment_text']
        assert '彩涂/在制料：0吨' in evidence.payload['attachment_text']
        assert evidence.payload['attachment_text'].endswith('报表日期：2026-07-08')
    finally:
        db.close()


def test_owner_verified_wip_screenshot_retry_preserves_first_confirmation_audit(monkeypatch, tmp_path) -> None:
    module, Session = _prepare(monkeypatch)
    files_root = tmp_path / 'files'
    files_root.mkdir()
    (files_root / 'wip.jpg').write_bytes(WIP_SCREENSHOT_CONTENT)
    input_jsonl = tmp_path / 'messages.jsonl'
    _write_jsonl(input_jsonl, [_owner_verified_visual_row()])

    first = module.run_backfill(
        input_jsonl=input_jsonl,
        files_root=files_root,
        days=3,
        confirmation_mode='owner-verified-dws-history',
        confirmation_run_id='test-run-wip-first',
    )
    db = Session()
    try:
        first_payload = json.loads(json.dumps(db.query(MultimodalEvidence).one().payload))
    finally:
        db.close()

    second = module.run_backfill(
        input_jsonl=input_jsonl,
        files_root=files_root,
        days=3,
        confirmation_mode='owner-verified-dws-history',
        confirmation_run_id='test-run-wip-retry',
    )

    assert first['confirmed'] == 1
    assert first['committed'] == 1
    assert second['confirmed'] == 0
    assert second['already_confirmed'] == 1
    assert second['committed'] == 1
    db = Session()
    try:
        evidence = db.query(MultimodalEvidence).one()
        assert evidence.payload == first_payload
        assert evidence.payload['owner_verification']['run_id'] == 'test-run-wip-first'
    finally:
        db.close()


def test_owner_verified_wip_screenshot_rolls_back_with_later_invalid_row(monkeypatch, tmp_path) -> None:
    module, Session = _prepare(monkeypatch)
    files_root = tmp_path / 'files'
    files_root.mkdir()
    (files_root / 'wip.jpg').write_bytes(WIP_SCREENSHOT_CONTENT)
    input_jsonl = tmp_path / 'messages.jsonl'
    invalid = _owner_verified_text_row(message_id='msg-invalid-after-wip')
    invalid['content_sha256'] = '0' * 64
    _write_jsonl(input_jsonl, [_owner_verified_visual_row(), invalid])

    summary = module.run_backfill(
        input_jsonl=input_jsonl,
        files_root=files_root,
        days=3,
        confirmation_mode='owner-verified-dws-history',
        confirmation_run_id='test-run-wip-atomic',
    )

    assert summary['committed'] == 0
    assert summary['confirmed'] == 0
    assert summary['confirmation_rejected'] == 1
    db = Session()
    try:
        assert db.query(MultimodalEvidence).count() == 0
    finally:
        db.close()


def test_owner_verified_wip_screenshot_rejects_invalid_visual_contract(monkeypatch, tmp_path) -> None:
    module, Session = _prepare(monkeypatch)
    files_root = tmp_path / 'files'
    files_root.mkdir()
    screenshot_path = files_root / 'wip.jpg'
    screenshot_content = WIP_SCREENSHOT_CONTENT
    screenshot_path.write_bytes(screenshot_content)
    input_jsonl = tmp_path / 'messages.jsonl'
    _write_jsonl(
        input_jsonl,
        [
            {
                'conversationId': 'group-001',
                'message_id': 'msg-owner-verified-wip-invalid',
                'senderStaffId': 'sender-wip-owner',
                'createTime': '2026-07-08T08:19:42+08:00',
                'fileName': 'wip.jpg',
                'fileId': 'dws-media-wip-invalid',
                'localFilePath': 'wip.jpg',
                'businessDate': '2026-07-08',
                'content_sha256': hashlib.sha256(screenshot_content).hexdigest(),
                'owner_verified_visual_facts': {
                    'wip_total': {
                        'value': 1877,
                        'unit': '吨',
                        'reported_date': '2026-07-08',
                        'row_label': '汇总',
                        'column_label': '在制料',
                    }
                },
            }
        ],
    )

    summary = module.run_backfill(
        input_jsonl=input_jsonl,
        files_root=files_root,
        days=3,
        confirmation_mode='owner-verified-dws-history',
        confirmation_run_id='test-run-wip-invalid',
    )

    assert summary['confirmation_rejected'] == 1
    assert summary['committed'] == 0
    db = Session()
    try:
        assert db.query(MultimodalEvidence).count() == 0
    finally:
        db.close()
