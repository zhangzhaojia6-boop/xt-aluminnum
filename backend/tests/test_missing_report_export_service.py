from io import BytesIO

from openpyxl import load_workbook

from app.services.missing_report_export_service import build_missing_report_workbook


def test_missing_report_workbook_contains_live_missing_rows_pending_rows_and_summary() -> None:
    content = build_missing_report_workbook(
        {
            'business_date': '2026-05-06',
            'live_aggregation': {
                'workshops': [
                    {
                        'workshop_id': 1,
                        'workshop_name': '铸轧三车间',
                        'machines': [
                            {
                                'machine_id': 9,
                                'machine_name': '9#机',
                                'shifts': [
                                    {
                                        'shift_id': 1,
                                        'shift_name': '大夜',
                                        'submission_status': 'not_started',
                                        'status_text': '缺报',
                                        'is_applicable': True,
                                    },
                                ],
                            },
                        ],
                    },
                ],
                'owner_daily_status': {
                    'items': [
                        {
                            'user_id': 30,
                            'workshop_name': '成品库',
                            'role_label': '总电工',
                            'person_name': '王电工',
                            'username': 'energy-chief',
                            'status': 'not_started',
                        },
                    ],
                },
            },
            'pending_assignment': {
                'summary': {
                    'entry_count': 1,
                    'missing_machine_count': 1,
                    'missing_shift_count': 0,
                },
                'items': [
                    {
                        'entry_id': 101,
                        'work_order_id': 201,
                        'tracking_card_no': 'RA260506001',
                        'workshop_name': '2050冷轧车间',
                        'shift_name': '夜班',
                        'created_by_user_name': '主操甲',
                        'created_by_username': 'op-a',
                        'created_at': '2026-05-06T09:30:00',
                        'missing_fields': ['machine_id'],
                        'mes_match_count': 1,
                        'mes_machine_name': '1#轧机',
                        'machine_candidate_names': ['1#轧机', '2#轧机'],
                        'input_weight': 100.0,
                        'output_weight': 96.0,
                        'scrap_weight': 4.0,
                        'entry_status': 'draft',
                        'entry_type': 'mobile_coil',
                    }
                ],
            },
            'summary': {
                'entry_count': 1,
                'missing_machine_count': 1,
                'missing_shift_count': 0,
            },
            'items': [
                {
                    'entry_id': 101,
                    'work_order_id': 201,
                    'tracking_card_no': 'RA260506001',
                    'workshop_name': '2050冷轧车间',
                    'shift_name': '夜班',
                    'created_by_user_name': '主操甲',
                    'created_by_username': 'op-a',
                    'created_at': '2026-05-06T09:30:00',
                    'missing_fields': ['machine_id'],
                    'mes_match_count': 1,
                    'mes_machine_name': '1#轧机',
                    'machine_candidate_names': ['1#轧机', '2#轧机'],
                    'input_weight': 100.0,
                    'output_weight': 96.0,
                    'scrap_weight': 4.0,
                    'entry_status': 'draft',
                    'entry_type': 'mobile_coil',
                }
            ],
        }
    )

    workbook = load_workbook(BytesIO(content), read_only=True)

    assert workbook.sheetnames == ['缺报明细', '待归属明细', '车间汇总']
    detail = workbook['缺报明细']
    assert detail['A1'].value == '缺报明细 2026-05-06'
    assert detail['B5'].value == '铸轧三车间'
    assert detail['C5'].value == '9#机'
    assert detail['D5'].value == '大夜班'
    assert detail['E5'].value == '主操'
    assert detail['H5'].value == '缺报'
    assert detail['B6'].value == '成品库'
    assert detail['C6'].value == '每日一录'
    assert detail['E6'].value == '总电工'
    assert detail['F6'].value == '王电工'

    pending = workbook['待归属明细']
    assert pending['H4'].value == '问题类型'
    assert pending['B5'].value == '2050冷轧车间'
    assert pending['D5'].value == '主操甲'
    assert pending['F5'].value == '2026-05-06 09:30:00'
    assert pending['G5'].value == 'RA260506001'
    assert pending['H5'].value == 'MES抓到但机列未填'
    assert pending['J5'].value == 1
    assert pending['K5'].value == '1#轧机'

    summary = workbook['车间汇总']
    assert summary['A3'].value == '车间'
    assert summary['A4'].value == '铸轧三车间'
    assert summary['B4'].value == 1
    assert summary['C4'].value == 1


def test_missing_report_workbook_appends_mes_gap_sheet() -> None:
    content = build_missing_report_workbook(
        {
            'business_date': '2026-05-06',
            'summary': {},
            'items': [],
            'mes_fill_gaps': {
                'summary': {'total': 1, 'status_counts': {'weight_mismatch': 1}},
                'items': [
                    {
                        'status': 'weight_mismatch',
                        'workshop_name': '2050冷轧车间',
                        'process_name': '冷轧',
                        'batch_no': 'BATCH-1',
                        'tracking_card_no': 'TRACK-1',
                        'local_entry_id': 101,
                        'mes_output_weight': 960.0,
                        'local_output_weight': 958.0,
                        'mes_machine_name': '1#轧机',
                        'local_machine_name': '1#轧机',
                    }
                ],
            },
        }
    )

    workbook = load_workbook(BytesIO(content), read_only=True)

    assert workbook.sheetnames == ['缺报明细', '待归属明细', '车间汇总', 'MES异常明细']
    mes_sheet = workbook['MES异常明细']
    assert mes_sheet['A1'].value == 'MES异常明细 2026-05-06'
    assert mes_sheet['B4'].value == '重量不一致'
    assert mes_sheet['C4'].value == '2050冷轧车间'
    assert mes_sheet['F4'].value == 'TRACK-1'
    assert mes_sheet['H4'].value == 960.0
    assert mes_sheet['I4'].value == 958.0
