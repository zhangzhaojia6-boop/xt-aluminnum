from datetime import date, datetime
from hashlib import sha256
from io import BytesIO
from xml.etree import ElementTree
from zipfile import ZipFile

from openpyxl import Workbook

from app.services.dingtalk_verified_fact_extractor import (
    extract_owner_verified_visual_fact_updates,
    extract_verified_file_fact_updates,
    extract_verified_text_fact_updates,
)


MODIFIED_TAG = '{http://purl.org/dc/terms/}modified'


def _set_workbook_modified(content: bytes, modified: datetime | None) -> bytes:
    output = BytesIO()
    with ZipFile(BytesIO(content), 'r') as source, ZipFile(output, 'w') as target:
        for item in source.infolist():
            value = source.read(item.filename)
            if item.filename == 'docProps/core.xml':
                root = ElementTree.fromstring(value)
                node = root.find(MODIFIED_TAG)
                if node is not None and modified is None:
                    root.remove(node)
                elif node is not None:
                    node.text = modified.isoformat(timespec='seconds') + 'Z'
                value = ElementTree.tostring(root, encoding='utf-8', xml_declaration=True)
            target.writestr(item, value)
    return output.getvalue()


def _workbook_bytes(
    *,
    month: int = 7,
    include_total_header: bool = True,
    value=0.8381,
    previous_total=0.8519,
    hot_roll=0.8354,
    previous_hot_roll=0.8446,
    modified: datetime | None = datetime(2026, 7, 12),
) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = '成品率'
    sheet['A1'] = f'{month}月份各车间成品率车间指标'
    sheet['A2'] = '日期'
    sheet.merge_cells('A2:A5')
    sheet['AB2'] = '公司'
    sheet.merge_cells('AB2:AK2')
    sheet['AD3'] = '铸轧普卷板'
    sheet.merge_cells('AD3:AE4')
    sheet['AF3'] = '铸轧（幕墙卷+普卷板）'
    sheet.merge_cells('AF3:AG4')
    sheet['AH3'] = '热轧普卷板'
    sheet.merge_cells('AH3:AI4')
    if include_total_header:
        sheet['AJ3'] = '总成品率92%'
        sheet.merge_cells('AJ3:AK4')
    for daily_cell, monthly_cell in (('AD5', 'AE5'), ('AF5', 'AG5'), ('AH5', 'AI5')):
        sheet[daily_cell] = '日合计'
        sheet[monthly_cell] = '月累计'
    sheet['AJ5'] = '日合计'
    sheet['AK5'] = '月累计'
    sheet['A6'] = 12
    sheet['AD6'] = 0.9301
    sheet['AE6'] = 0.9202
    sheet['AF6'] = 0.9311
    sheet['AG6'] = 0.9202
    sheet['AH6'] = hot_roll
    sheet['AI6'] = 0.8446
    sheet['AJ6'] = value
    sheet['AK6'] = 0.851472683750808
    sheet['A7'] = 11
    sheet['AH7'] = previous_hot_roll
    sheet['AJ7'] = previous_total
    output = BytesIO()
    workbook.save(output)
    return _set_workbook_modified(output.getvalue(), modified)


def _extract(content: bytes) -> dict:
    return extract_verified_file_fact_updates(
        file_name='各车间成品率.xlsx',
        content=content,
        business_date=date(2026, 7, 12),
        file_sha256=sha256(content).hexdigest(),
    )


def test_verified_yield_extractor_requires_matching_month() -> None:
    assert _extract(_workbook_bytes(month=6)) == {}


def test_verified_yield_extractor_requires_total_yield_header() -> None:
    assert _extract(_workbook_bytes(include_total_header=False)) == {}


def test_verified_yield_extractor_rejects_non_numeric_target_cell() -> None:
    assert _extract(_workbook_bytes(value='/')) == {}


def test_verified_yield_extractor_returns_cell_traceable_fact() -> None:
    result = _extract(_workbook_bytes())

    assert result['daily_yield_rate']['value'] == 83.81
    assert result['daily_yield_rate']['source_ref']['workbook_modified_date'] == '2026-07-12'


def test_verified_yield_extractor_returns_complete_report_yield_facts() -> None:
    result = _extract(_workbook_bytes())

    assert {field: item['value'] for field, item in result.items()} == {
        'daily_yield_rate': 83.81,
        'monthly_yield_rate': 85.15,
        'hot_roll_yield_rate': 83.54,
        'hot_roll_monthly_yield_rate': 84.46,
        'cast_roll_yield_rate': 92.02,
        'plate_coil_yield_rate': 92.02,
        'daily_yield_delta': -1.38,
        'hot_roll_yield_delta': -0.92,
    }
    assert result['daily_yield_delta']['source_ref']['value_cell'] == 'AJ6'
    assert result['daily_yield_delta']['source_ref']['previous_value_cell'] == 'AJ7'
    assert result['hot_roll_yield_delta']['source_ref']['value_cell'] == 'AH6'
    assert result['hot_roll_yield_delta']['source_ref']['previous_value_cell'] == 'AH7'


def test_verified_yield_delta_uses_displayed_two_decimal_rates() -> None:
    result = _extract(
        _workbook_bytes(
            value=0.844536413217838,
            previous_total=0.866785514841703,
            hot_roll=0.844536413217838,
            previous_hot_roll=0.844965978549187,
        )
    )

    assert result['daily_yield_delta']['value'] == -2.23
    assert result['hot_roll_yield_delta']['value'] == -0.05


def test_verified_yield_extractor_requires_complete_matching_workbook_date() -> None:
    assert _extract(_workbook_bytes(modified=None)) == {}
    assert _extract(_workbook_bytes(modified=datetime(2025, 7, 12))) == {}


def test_verified_yield_extractor_rejects_duplicate_business_date_rows() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = '成品率'
    workbook.properties.modified = datetime(2026, 7, 12)
    sheet['A1'] = '7月份各车间成品率车间指标'
    sheet['A2'] = '日期'
    sheet.merge_cells('A2:A5')
    sheet['AJ3'] = '总成品率92%'
    sheet.merge_cells('AJ3:AK4')
    sheet['AJ5'] = '日合计'
    sheet['A6'] = 12
    sheet['AJ6'] = 0.8381
    sheet['A7'] = 12
    sheet['AJ7'] = 0.9
    output = BytesIO()
    workbook.save(output)
    content = _set_workbook_modified(output.getvalue(), datetime(2026, 7, 12))

    assert _extract(content) == {}


PLAN_CONTRACT_TEXT = (
    '投料量：2050投料463吨 1850投料0吨 外加工62吨  中厚板0吨  '
    '当天合同443吨    热轧436吨   总余合同量2765吨'
)


def _extract_text(text: str = PLAN_CONTRACT_TEXT) -> dict:
    return extract_verified_text_fact_updates(
        text=text,
        business_date=date(2026, 7, 11),
        content_sha256=sha256(text.encode('utf-8')).hexdigest(),
    )


def test_verified_plan_contract_text_returns_traceable_exact_facts() -> None:
    result = _extract_text()

    assert result['daily_input_weight']['value'] == 525
    assert result['daily_input_weight']['source_ref']['components'] == {
        '2050_input': 463,
        '1850_input': 0,
        'external_processing': 62,
        'medium_plate': 0,
    }
    assert result['cold_roll_input_daily']['value'] == 525
    assert result['cold_roll_input_daily']['source_ref']['components'] == {
        '2050_input': 463,
        '1850_input': 0,
        'external_processing': 62,
    }
    assert result['remaining_contract_weight']['value'] == 2765
    assert result['remaining_contract_weight']['source_ref']['parser'] == 'plan_contract_message_v1'
    assert result['remaining_contract_weight']['source_ref']['business_date'] == '2026-07-11'
    assert result['remaining_contract_weight']['source_ref']['content_sha256'] == sha256(
        PLAN_CONTRACT_TEXT.encode('utf-8')
    ).hexdigest()
    assert result['remaining_contract_weight']['source_ref']['matched_segments']['remaining_contract'][
        'text'
    ] == '总余合同量2765吨'
    assert {field: result[field]['value'] for field in (
        'cold_2050_input_daily',
        'cold_1850_input_daily',
        'outsourced_input_daily',
        'medium_plate_input_daily',
        'daily_contract_weight',
        'daily_hot_roll_contract_weight',
    )} == {
        'cold_2050_input_daily': 463,
        'cold_1850_input_daily': 0,
        'outsourced_input_daily': 62,
        'medium_plate_input_daily': 0,
        'daily_contract_weight': 443,
        'daily_hot_roll_contract_weight': 436,
    }


def test_verified_plan_contract_text_tolerates_spacing_and_ascii_colon() -> None:
    text = (
        '投料量: 2050投料 313 吨\n1850投料 0 吨，外加工 24 吨；中厚板 106 吨 '
        '当天合同 212 吨 热轧 212 吨 总余合同量 2664 吨'
    )

    result = _extract_text(text)

    assert result['daily_input_weight']['value'] == 443
    assert result['cold_roll_input_daily']['value'] == 337
    assert result['remaining_contract_weight']['value'] == 2664


def test_verified_plan_contract_text_rejects_missing_component_or_duplicate_label() -> None:
    assert _extract_text(PLAN_CONTRACT_TEXT.replace('中厚板0吨', '')) == {}
    assert _extract_text(PLAN_CONTRACT_TEXT + ' 总余合同量2765吨') == {}


def test_verified_plan_contract_text_rejects_malformed_number_or_wrong_hash() -> None:
    assert _extract_text(PLAN_CONTRACT_TEXT.replace('2765吨', '2,7,65吨')) == {}
    assert extract_verified_text_fact_updates(
        text=PLAN_CONTRACT_TEXT,
        business_date=date(2026, 7, 11),
        content_sha256='0' * 64,
    ) == {}


def test_verified_plan_contract_text_ignores_unrelated_messages() -> None:
    assert _extract_text('今天合同很多，2050投料463吨，请大家关注') == {}


WIP_SCREENSHOT = b'\xff\xd8\xffowner-verified-wip\xff\xd9'
WIP_FULL_VALUES = {
    'wip_total': 1166.5,
    'wip_1650_2050_cold': 345.5,
    'wip_1850_cold': 103.5,
    'wip_milling': 0,
    'wip_anneal_total': 323.5,
    'wip_new_north': 231,
    'wip_new_south': 37,
    'wip_park_anneal': 55.5,
    'wip_finishing_total': 394,
    'wip_straightening': 290,
    'wip_finishing': 99.5,
    'wip_park_finishing': 4.5,
    'wip_hot_plate_shearing': 0,
    'wip_coating': 0,
}
WIP_ROW_LABELS = {
    'wip_total': '汇总',
    'wip_1650_2050_cold': '1650+2050冷轧合计',
    'wip_1850_cold': '1850冷轧合计',
    'wip_milling': '铣床合计',
    'wip_anneal_total': '在线退火合计',
    'wip_new_north': '新厂北',
    'wip_new_south': '新厂南',
    'wip_park_anneal': '园区退火',
    'wip_finishing_total': '后工序合计',
    'wip_straightening': '拉矫合计',
    'wip_finishing': '精整合计',
    'wip_park_finishing': '园区剪切',
    'wip_hot_plate_shearing': '热轧（中厚板）',
    'wip_coating': '彩涂',
}


def _extract_wip_visual(
    *,
    business_date: date = date(2026, 7, 11),
    reported_date: str = '2026-07-12',
    event_time: str = '2026-07-12T08:19:42+08:00',
    value=1877,
    content_sha256: str | None = None,
) -> dict:
    return extract_owner_verified_visual_fact_updates(
        file_name='wip.jpg',
        content=WIP_SCREENSHOT,
        business_date=business_date,
        event_time=event_time,
        file_sha256=content_sha256 or sha256(WIP_SCREENSHOT).hexdigest(),
        verified_facts={
            'wip_total': {
                'value': value,
                'unit': '吨',
                'reported_date': reported_date,
                'row_label': '汇总',
                'column_label': '在制料',
            }
        },
    )


def test_owner_verified_wip_screenshot_returns_traceable_fact() -> None:
    result = _extract_wip_visual()

    assert result['wip_total']['value'] == 1877
    assert result['wip_total']['unit'] == '吨'
    assert result['wip_total']['source_ref'] == {
        'parser': 'owner_verified_wip_screenshot_v1',
        'verification_mode': 'owner_verified_visual',
        'reported_date': '2026-07-12',
        'business_date': '2026-07-11',
        'business_date_rule': 'next_calendar_day_before_owner_daily_cutoff',
        'event_time_cutoff': '09:30',
        'row_label': '汇总',
        'column_label': '在制料',
        'file_sha256': sha256(WIP_SCREENSHOT).hexdigest(),
    }


def test_owner_verified_wip_screenshot_returns_complete_consistent_table() -> None:
    facts = {
        field: {
            'value': value,
            'unit': '吨',
            'reported_date': '2026-07-12',
            'row_label': WIP_ROW_LABELS[field],
            'column_label': '在制料',
        }
        for field, value in WIP_FULL_VALUES.items()
    }

    result = extract_owner_verified_visual_fact_updates(
        file_name='wip.jpg',
        content=WIP_SCREENSHOT,
        business_date=date(2026, 7, 11),
        event_time='2026-07-12T08:19:42+08:00',
        file_sha256=sha256(WIP_SCREENSHOT).hexdigest(),
        verified_facts=facts,
    )

    assert {field: item['value'] for field, item in result.items()} == WIP_FULL_VALUES
    assert all(item['source_ref']['parser'] == 'owner_verified_wip_screenshot_v2' for item in result.values())
    assert result['wip_new_north']['source_ref']['row_label'] == '新厂北'


def test_owner_verified_wip_screenshot_rejects_inconsistent_full_table() -> None:
    values = {**WIP_FULL_VALUES, 'wip_total': 1167}
    facts = {
        field: {
            'value': value,
            'unit': '吨',
            'reported_date': '2026-07-12',
            'row_label': WIP_ROW_LABELS[field],
            'column_label': '在制料',
        }
        for field, value in values.items()
    }

    assert extract_owner_verified_visual_fact_updates(
        file_name='wip.jpg',
        content=WIP_SCREENSHOT,
        business_date=date(2026, 7, 11),
        event_time='2026-07-12T08:19:42+08:00',
        file_sha256=sha256(WIP_SCREENSHOT).hexdigest(),
        verified_facts=facts,
    ) == {}


def test_owner_verified_wip_screenshot_uses_image_magic_for_synthetic_media_extension() -> None:
    content = b'\x89PNG\r\n\x1a\nowner-verified-wip-IEND\xaeB`\x82'

    result = extract_owner_verified_visual_fact_updates(
        file_name='dingtalk-media.jpg',
        content=content,
        business_date=date(2026, 7, 11),
        event_time='2026-07-12T08:19:42+08:00',
        file_sha256=sha256(content).hexdigest(),
        verified_facts={
            'wip_total': {
                'value': 1877,
                'unit': '吨',
                'reported_date': '2026-07-12',
                'row_label': '汇总',
                'column_label': '在制料',
            }
        },
    )

    assert result['wip_total']['value'] == 1877


def test_owner_verified_wip_screenshot_rejects_wrong_business_date_mapping() -> None:
    assert _extract_wip_visual(business_date=date(2026, 7, 12)) == {}
    assert _extract_wip_visual(reported_date='2026-07-13') == {}
    assert _extract_wip_visual(event_time='2026-07-13T08:19:42+08:00') == {}


def test_owner_verified_wip_screenshot_requires_next_morning_before_owner_cutoff() -> None:
    assert _extract_wip_visual(event_time='2026-07-12T09:29:59+08:00')['wip_total']['value'] == 1877
    assert _extract_wip_visual(event_time='2026-07-12T09:30:00+08:00') == {}
    assert _extract_wip_visual(event_time='2026-07-12T15:19:42+08:00') == {}


def test_owner_verified_wip_screenshot_rejects_wrong_hash_or_invalid_value() -> None:
    assert _extract_wip_visual(content_sha256='0' * 64) == {}
    assert _extract_wip_visual(value='not-a-number') == {}
