from __future__ import annotations

from collections import defaultdict
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


MISSING_REPORT_HEADERS = [
    '序号',
    '车间',
    '机列/岗位',
    '班次/频次',
    '应填角色',
    '责任人',
    '登录账号',
    '状态',
    '缺报口径',
]

PENDING_ASSIGNMENT_HEADERS = [
    '序号',
    '车间',
    '班次',
    '缺报人员',
    '登录账号',
    '缺报时间',
    '对应卷号',
    '问题类型',
    '缺失项',
    'MES命中数',
    'MES推断机列',
    '候选机列',
    '投入吨',
    '产出吨',
    '废料吨',
    '记录状态',
    '记录类型',
    '工单ID',
    '填报ID',
]

SUMMARY_HEADERS = [
    '车间',
    '应填缺报',
    '主操缺报',
    '电工缺报',
    '内勤缺报',
    '迟报',
    '待归属记录',
    '涉及人员/岗位',
]

MES_GAP_HEADERS = [
    '序号',
    '状态',
    '车间',
    '工序',
    '批号',
    '随行卡',
    '本地填报ID',
    'MES产出kg',
    '本地产出kg',
    'MES机列',
    '本地机列',
]

FIELD_LABELS = {
    'machine_id': '机列未填',
    'shift_id': '班次未填',
}

SHIFT_LABELS = {
    'A': '长白班',
    'D': '长白班',
    '白班': '长白班',
    '长白': '长白班',
    '长白班': '长白班',
    'B': '小夜班',
    'E': '小夜班',
    '中班': '小夜班',
    '小夜': '小夜班',
    '小夜班': '小夜班',
    'C': '大夜班',
    'N': '大夜班',
    '夜班': '大夜班',
    '大夜': '大夜班',
    '大夜班': '大夜班',
}

SHIFT_ORDER = {
    '长白班': 0,
    '小夜班': 1,
    '大夜班': 2,
    '每日一录': 3,
}

MES_GAP_STATUS_LABELS = {
    'missing_local_entry': 'MES有工序本地未填',
    'mes_batch_unmapped': '批号未映射',
    'local_entry_unassigned': '本地未归机列',
    'weight_mismatch': '重量不一致',
    'matched': '已匹配',
}


def _text(value: Any, default: str = '-') -> str:
    if value is None:
        return default
    text = str(value).strip()
    return text or default


def _number(value: Any) -> float:
    try:
        return round(float(value or 0), 3)
    except (TypeError, ValueError):
        return 0.0


def _time_text(value: Any) -> str:
    text = _text(value, '')
    if not text:
        return '-'
    return text.replace('T', ' ')[:19]


def _missing_fields_text(item: dict[str, Any]) -> str:
    labels = [FIELD_LABELS.get(str(field), str(field)) for field in item.get('missing_fields') or []]
    return '、'.join(labels) or '-'


def _person_text(item: dict[str, Any]) -> str:
    return _text(item.get('created_by_user_name') or item.get('created_by_username'))


def _candidate_text(item: dict[str, Any]) -> str:
    names = item.get('machine_candidate_names') or []
    return '、'.join(str(name) for name in names if str(name).strip()) or '-'


def _issue_type(item: dict[str, Any]) -> str:
    missing_fields = set(item.get('missing_fields') or [])
    if 'machine_id' in missing_fields and int(item.get('mes_match_count') or 0) > 0:
        return 'MES抓到但机列未填'
    if 'machine_id' in missing_fields:
        return '机列未填'
    if 'shift_id' in missing_fields:
        return '班次未填'
    return '待补齐'


def _shift_text(value: Any) -> str:
    text = _text(value)
    return SHIFT_LABELS.get(text, text)


def _row_status_text(value: Any) -> str:
    text = _text(value)
    if text == 'late':
        return '迟报'
    if text in {'not_started', 'missing'}:
        return '缺报'
    return text


def _missing_caliber(row: dict[str, Any]) -> str:
    if row.get('source_type') == 'owner_daily':
        return '每日一录未完成'
    if row.get('source_type') == 'machine_shift':
        return '机列班次未填'
    return _text(row.get('missing_caliber'))


def _role_bucket(row: dict[str, Any]) -> str:
    role_label = str(row.get('role_label') or row.get('roleLabel') or '')
    if '电工' in role_label or '能源' in role_label or '能耗' in role_label:
        return 'electrician'
    if row.get('source_type') == 'owner_daily' or '内勤' in role_label or row.get('machine_name') == '每日一录':
        return 'owner'
    return 'operator'


def _is_missing_shift_cell(cell: dict[str, Any]) -> bool:
    return cell.get('is_applicable') is not False and (
        cell.get('submission_status') == 'not_started'
        or cell.get('status_text') == '缺报'
        or cell.get('statusText') == '缺报'
    )


def _normal_missing_row(raw: dict[str, Any]) -> dict[str, Any]:
    return {
        'workshop_name': _text(raw.get('workshop_name') or raw.get('workshopName')),
        'machine_name': _text(raw.get('machine_name') or raw.get('machineName')),
        'shift_name': _shift_text(raw.get('shift_name') or raw.get('shiftName')),
        'role_label': _text(raw.get('role_label') or raw.get('roleLabel')),
        'owner_name': _text(raw.get('owner_name') or raw.get('ownerName') or raw.get('person_name') or raw.get('personName')),
        'username': _text(raw.get('username') or raw.get('created_by_username')),
        'status_text': _row_status_text(raw.get('status_text') or raw.get('statusText') or raw.get('status')),
        'source_type': _text(raw.get('source_type') or raw.get('sourceType')),
        'missing_caliber': _text(raw.get('missing_caliber') or raw.get('missingCaliber')),
    }


def _build_live_missing_rows(payload: dict[str, Any]) -> list[dict[str, Any]]:
    explicit_rows = payload.get('missing_report_rows')
    if isinstance(explicit_rows, list):
        return [_normal_missing_row(row) for row in explicit_rows if isinstance(row, dict)]

    source = payload.get('live_aggregation') if isinstance(payload.get('live_aggregation'), dict) else payload
    rows: list[dict[str, Any]] = []
    for workshop in source.get('workshops') or []:
        if not isinstance(workshop, dict):
            continue
        workshop_name = _text(workshop.get('workshop_name') or workshop.get('workshopName'))
        for machine in workshop.get('machines') or []:
            if not isinstance(machine, dict):
                continue
            machine_name = _text(machine.get('machine_name') or machine.get('machineName'))
            for shift in machine.get('shifts') or []:
                if not isinstance(shift, dict) or not _is_missing_shift_cell(shift):
                    continue
                rows.append(_normal_missing_row({
                    'workshop_name': workshop_name,
                    'machine_name': machine_name,
                    'shift_name': shift.get('shift_name') or shift.get('shiftName'),
                    'role_label': '主操',
                    'status_text': shift.get('status_text') or shift.get('statusText') or '缺报',
                    'source_type': 'machine_shift',
                    'missing_caliber': '机列班次未填',
                }))

    owner_status = source.get('owner_daily_status') or {}
    if isinstance(owner_status, dict):
        for item in owner_status.get('items') or []:
            if not isinstance(item, dict) or item.get('status') == 'submitted':
                continue
            rows.append(_normal_missing_row({
                'workshop_name': item.get('workshop_name') or item.get('workshopName') or '全厂专项',
                'machine_name': '每日一录',
                'shift_name': '每日一录',
                'role_label': item.get('role_label') or item.get('role') or '内勤岗',
                'owner_name': item.get('person_name') or item.get('name'),
                'username': item.get('username'),
                'status_text': '迟报' if item.get('status') == 'late' else '缺报',
                'source_type': 'owner_daily',
                'missing_caliber': '每日一录未完成',
            }))

    return sorted(
        rows,
        key=lambda row: (
            1 if row.get('source_type') == 'owner_daily' else 0,
            _text(row.get('workshop_name')),
            SHIFT_ORDER.get(_shift_text(row.get('shift_name')), 99),
            _text(row.get('machine_name')),
            _text(row.get('role_label')),
            _text(row.get('owner_name')),
        ),
    )


def _sort_items(items: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return sorted(
        items,
        key=lambda item: (
            _text(item.get('workshop_name')),
            _text(item.get('shift_name')),
            _person_text(item),
            _time_text(item.get('created_at')),
            _text(item.get('tracking_card_no')),
        ),
    )


def _append_title(ws, title: str, column_count: int) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=column_count)
    cell = ws.cell(row=1, column=1, value=title)
    cell.font = Font(bold=True, size=14, color='FFFFFF')
    cell.fill = PatternFill('solid', fgColor='17324D')
    cell.alignment = Alignment(horizontal='center')


def _style_table(ws, header_row: int, column_count: int) -> None:
    fill = PatternFill('solid', fgColor='D9EAF7')
    for col in range(1, column_count + 1):
        cell = ws.cell(row=header_row, column=col)
        cell.font = Font(bold=True, color='17324D')
        cell.fill = fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1).coordinate
    ws.auto_filter.ref = f'A{header_row}:{get_column_letter(column_count)}{max(ws.max_row, header_row)}'


def _fit_columns(ws, max_width: int = 42) -> None:
    for column_cells in ws.columns:
        width = 10
        for cell in column_cells:
            width = max(width, min(len(str(cell.value or '')) + 2, max_width))
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = width


def _build_missing_report_rows(items: list[dict[str, Any]]) -> list[list[Any]]:
    rows = []
    for index, item in enumerate(items, start=1):
        rows.append(
            [
                index,
                _text(item.get('workshop_name')),
                _text(item.get('machine_name')),
                _shift_text(item.get('shift_name')),
                _text(item.get('role_label')),
                _text(item.get('owner_name')),
                _text(item.get('username')),
                _text(item.get('status_text')),
                _missing_caliber(item),
            ]
        )
    return rows


def _build_pending_assignment_rows(items: list[dict[str, Any]]) -> list[list[Any]]:
    rows = []
    for index, item in enumerate(_sort_items(items), start=1):
        rows.append(
            [
                index,
                _text(item.get('workshop_name')),
                _text(item.get('shift_name')),
                _person_text(item),
                _text(item.get('created_by_username')),
                _time_text(item.get('created_at')),
                _text(item.get('tracking_card_no')),
                _issue_type(item),
                _missing_fields_text(item),
                int(item.get('mes_match_count') or 0),
                _text(item.get('mes_machine_name')),
                _candidate_text(item),
                _number(item.get('input_weight')),
                _number(item.get('output_weight')),
                _number(item.get('scrap_weight')),
                _text(item.get('entry_status')),
                _text(item.get('entry_type')),
                item.get('work_order_id') or '',
                item.get('entry_id') or '',
            ]
        )
    return rows


def _build_summary_rows(missing_rows: list[dict[str, Any]], pending_items: list[dict[str, Any]]) -> list[list[Any]]:
    buckets: dict[str, dict[str, Any]] = defaultdict(lambda: {
        'missing': 0,
        'operator': 0,
        'electrician': 0,
        'owner': 0,
        'late': 0,
        'pending_assignment': 0,
        'people': set(),
    })
    for row in missing_rows:
        workshop = _text(row.get('workshop_name'), '未标记车间')
        bucket = buckets[workshop]
        bucket['missing'] += 1
        bucket[_role_bucket(row)] += 1
        if _text(row.get('status_text')) == '迟报':
            bucket['late'] += 1
        person = _text(row.get('owner_name'))
        label = _text(row.get('role_label'))
        if person != '-':
            bucket['people'].add(person)
        elif label != '-':
            bucket['people'].add(label)

    for item in pending_items:
        workshop = _text(item.get('workshop_name'), '未标记车间')
        bucket = buckets[workshop]
        bucket['pending_assignment'] += 1
        person = _person_text(item)
        if person != '-':
            bucket['people'].add(person)

    rows: list[list[Any]] = []
    for workshop, bucket in sorted(
        buckets.items(),
        key=lambda pair: (
            -pair[1]['operator'],
            -pair[1]['electrician'],
            -pair[1]['owner'],
            -pair[1]['pending_assignment'],
            pair[0],
        ),
    ):
        rows.append(
            [
                workshop,
                bucket['missing'],
                bucket['operator'],
                bucket['electrician'],
                bucket['owner'],
                bucket['late'],
                bucket['pending_assignment'],
                '、'.join(sorted(bucket['people'])) or '-',
            ]
        )
    return rows


def _build_mes_gap_rows(items: list[dict[str, Any]]) -> list[list[Any]]:
    rows = []
    for index, item in enumerate(items, start=1):
        status = _text(item.get('status'), '')
        rows.append(
            [
                index,
                MES_GAP_STATUS_LABELS.get(status, status or '-'),
                _text(item.get('workshop_name')),
                _text(item.get('process_name')),
                _text(item.get('batch_no')),
                _text(item.get('tracking_card_no')),
                item.get('local_entry_id') or '',
                _number(item.get('mes_output_weight')),
                _number(item.get('local_output_weight')),
                _text(item.get('mes_machine_name')),
                _text(item.get('local_machine_name')),
            ]
        )
    return rows


def _append_mes_gap_sheet(workbook: Workbook, *, business_date: str, mes_fill_gaps: dict[str, Any]) -> None:
    mes_sheet = workbook.create_sheet('MES异常明细')
    _append_title(mes_sheet, f'MES异常明细 {business_date}', len(MES_GAP_HEADERS))
    mes_sheet.append([])
    mes_sheet.append(MES_GAP_HEADERS)
    rows = _build_mes_gap_rows(list(mes_fill_gaps.get('items') or []))
    if rows:
        for row in rows:
            mes_sheet.append(row)
    else:
        mes_sheet.append(['暂无MES异常'] + [''] * (len(MES_GAP_HEADERS) - 1))
    _style_table(mes_sheet, 3, len(MES_GAP_HEADERS))
    _fit_columns(mes_sheet)


def build_missing_report_workbook(payload: dict[str, Any]) -> bytes:
    business_date = _text(payload.get('business_date'), '')
    pending_assignment = payload.get('pending_assignment') if isinstance(payload.get('pending_assignment'), dict) else payload
    items = list(pending_assignment.get('items') or [])
    summary = dict(pending_assignment.get('summary') or payload.get('summary') or {})
    missing_rows = _build_live_missing_rows(payload)

    workbook = Workbook()
    detail = workbook.active
    detail.title = '缺报明细'
    _append_title(detail, f'缺报明细 {business_date}', len(MISSING_REPORT_HEADERS))
    detail.cell(row=2, column=1, value=(
        f"应填缺报{len(missing_rows)}条；"
        f"待归属{int(summary.get('entry_count') or len(items))}条；"
        f"缺机列{int(summary.get('missing_machine_count') or 0)}条"
    ))
    detail.append([])
    detail.append(MISSING_REPORT_HEADERS)
    missing_detail_rows = _build_missing_report_rows(missing_rows)
    if missing_detail_rows:
        for row in missing_detail_rows:
            detail.append(row)
    else:
        detail.append(['暂无应填缺报'] + [''] * (len(MISSING_REPORT_HEADERS) - 1))
    _style_table(detail, 4, len(MISSING_REPORT_HEADERS))
    _fit_columns(detail)

    pending_sheet = workbook.create_sheet('待归属明细')
    _append_title(pending_sheet, f'待归属明细 {business_date}', len(PENDING_ASSIGNMENT_HEADERS))
    pending_sheet.cell(row=2, column=1, value=(
        f"共{int(summary.get('entry_count') or len(items))}条；"
        f"缺机列{int(summary.get('missing_machine_count') or 0)}条；"
        f"缺班次{int(summary.get('missing_shift_count') or 0)}条"
    ))
    pending_sheet.append([])
    pending_sheet.append(PENDING_ASSIGNMENT_HEADERS)
    pending_rows = _build_pending_assignment_rows(items)
    if pending_rows:
        for row in pending_rows:
            pending_sheet.append(row)
    else:
        pending_sheet.append(['暂无待归属记录'] + [''] * (len(PENDING_ASSIGNMENT_HEADERS) - 1))
    _style_table(pending_sheet, 4, len(PENDING_ASSIGNMENT_HEADERS))
    _fit_columns(pending_sheet)

    summary_sheet = workbook.create_sheet('车间汇总')
    _append_title(summary_sheet, f'车间缺报汇总 {business_date}', len(SUMMARY_HEADERS))
    summary_sheet.append([])
    summary_sheet.append(SUMMARY_HEADERS)
    summary_rows = _build_summary_rows(missing_rows, items)
    if summary_rows:
        for row in summary_rows:
            summary_sheet.append(row)
    else:
        summary_sheet.append(['暂无缺报'] + [''] * (len(SUMMARY_HEADERS) - 1))
    _style_table(summary_sheet, 3, len(SUMMARY_HEADERS))
    _fit_columns(summary_sheet)

    mes_fill_gaps = payload.get('mes_fill_gaps')
    if isinstance(mes_fill_gaps, dict):
        _append_mes_gap_sheet(workbook, business_date=business_date, mes_fill_gaps=mes_fill_gaps)

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
