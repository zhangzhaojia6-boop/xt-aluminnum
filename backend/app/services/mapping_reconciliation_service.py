from __future__ import annotations

import json
import os
import re
from datetime import date, datetime
from dataclasses import asdict, dataclass
from decimal import Decimal
from pathlib import Path
from typing import Any, Iterable, Mapping, Sequence

from openpyxl import load_workbook
from sqlalchemy import func
from sqlalchemy.exc import OperationalError, ProgrammingError
from sqlalchemy.orm import Session

from app.core.business_time import production_business_window
from app.models.consumable import DailyConsumableLog
from app.models.energy import MachineEnergyRecord
from app.models.executive import CostDailyResult, MachineDailyCostSnapshot
from app.models.master import Equipment, Workshop
from app.models.mes import MesDailyWipSnapshot, MesStockRecord, MesWipTotalSnapshot, MesWorkshopProcessRecord
from app.models.production import MobileShiftReport, ShiftProductionData, WorkOrder, WorkOrderEntry
from app.models.shift import ShiftConfig


DEFAULT_DIMENSIONS = ('business_date', 'workshop', 'shift')
DEFAULT_DIMENSION_ALIASES = {
    'workshop': {
        '铸轧二': '铸二',
        '铸轧二车间': '铸二',
        '铸二车间': '铸二',
        '铸轧三': '铸三',
        '铸轧三车间': '铸三',
        '铸三车间': '铸三',
        '1650': '冷轧1650',
        '1650车间': '冷轧1650',
        '1650冷轧': '冷轧1650',
        '1650冷轧车间': '冷轧1650',
        '1850': '冷轧1850',
        '1850车间': '冷轧1850',
        '1850冷轧': '冷轧1850',
        '1850冷轧车间': '冷轧1850',
        '2050': '冷轧2050',
        '2050车间': '冷轧2050',
        '2050冷轧': '冷轧2050',
        '2050冷轧车间': '冷轧2050',
        '精整车间': '精整',
        '精整(剪子)': '精整',
        '拉矫车间': '拉矫',
        '拉矫下机量': '拉矫',
        '园区剪切车间': '园区剪切',
        '剪切车间': '园区剪切',
        '园区精整': '园区剪切',
        '园区圆片': '圆片料',
        '淬火': '淬火车间',
        '园区淬火': '淬火车间',
        '园区淬火车间': '淬火车间',
        '淬火+覆膜': '淬火车间',
        '新厂在线车间': '新厂在线',
        '在线车间': '新厂在线',
        '新厂北线': '新厂在线',
        '新厂南线': '新厂在线',
        '园区在线车间': '园区在线',
        '园区退火': '园区在线',
        '园区退火(北)': '园区在线',
        '园区退火(南)': '园区在线',
        '铣床车间': '铣床',
        '涂层': '彩涂',
        '辊涂(光铝)': '彩涂',
    },
}
REPOSITORY_ROOT = Path(__file__).resolve().parents[3]
DEFAULT_REFERENCE_ROOT = Path('D:/输出skill')
FALLBACK_REFERENCE_ROOT = REPOSITORY_ROOT / 'reference' / 'output-skill'
PARSEABLE_REFERENCE_EXTENSIONS = {'.txt', '.md', '.log', '.xlsx', '.xls', '.json', '.ndjson'}
IMAGE_REFERENCE_EXTENSIONS = {'.png', '.jpg', '.jpeg', '.webp', '.bmp'}
IGNORED_REFERENCE_DIR_NAMES = {'.pytest_cache', '__pycache__'}
SYSTEM_SOURCES = [
    'mes_stock_records',
    'mes_workshop_process_records',
    'shift_production_data',
    'work_order_entries',
    'daily_consumable_logs',
    'cost_daily_result',
    'machine_daily_cost_snapshots',
    'machine_energy_records',
    'mes_daily_wip_snapshots',
    'mes_wip_total_snapshots',
    'data_quality_issues',
    'data_reconciliation_items',
    'daily_reports',
    'workshops',
    'equipment',
    'shift_configs',
]
TEXT_EXTENSIONS = {'.txt', '.md', '.log'}
EXCEL_EXTENSIONS = {'.xlsx', '.xls'}
JSON_EXTENSIONS = {'.json'}
JSON_LINES_EXTENSIONS = {'.ndjson'}
SHIFT_NAMES = ('长白班', '小夜班', '大夜班', '白班', '小夜', '大夜')
DATE_RE = re.compile(r'(20\d{2})[年\-/.](\d{1,2})[月\-/.](\d{1,2})日?')
NUMBER_RE = r'([0-9]+(?:\.[0-9]+)?)'
CONSUMABLE_REFERENCE_FIELD_ALIASES = {
    'liquefied_gas_per_ton': ('液化气吨耗', '液化气单吨消耗', '液化气每吨', 'liquefied_gas_per_ton'),
    'titanium_wire_per_ton': ('钛丝吨耗', '钛丝单吨消耗', '钛丝每吨', 'titanium_wire_per_ton'),
    'steel_strip_per_ton': ('钢带吨耗', '钢带单吨消耗', '钢带每吨', 'steel_strip_per_ton'),
    'magnesium_per_ton': ('镁吨耗', '镁单吨消耗', '镁每吨', 'magnesium_per_ton'),
    'manganese_per_ton': ('锰吨耗', '锰单吨消耗', '锰每吨', 'manganese_per_ton'),
    'iron_per_ton': ('铁吨耗', '铁单吨消耗', '铁每吨', 'iron_per_ton'),
    'copper_per_ton': ('铜吨耗', '铜单吨消耗', '铜每吨', 'copper_per_ton'),
    'electricity_monthly': ('用电月累计', '电量月累计', '电耗月累计', '总电月累计', 'electricity_monthly'),
    'electricity_target': ('用电指标', '电量指标', '电耗指标', '总电指标', 'electricity_target'),
    'gas_monthly': ('用气月累计', '气耗月累计', '天然气月累计', '燃气月累计', 'gas_monthly'),
    'gas_target': ('用气指标', '气耗指标', '天然气指标', '燃气指标', 'gas_target'),
    'hot_roll_emulsion_per_ton': ('热轧乳液吨耗', '热轧乳化液吨耗', '乳液吨耗', 'hot_roll_emulsion_per_ton'),
    'diatomite_per_ton': ('硅藻土吨耗', '硅藻土单吨消耗', '硅藻土每吨', 'diatomite_per_ton'),
    'white_earth_per_ton': ('白土吨耗', '白土单吨消耗', '白土每吨', 'white_earth_per_ton'),
    'filter_cloth_daily': ('滤布日耗', '滤布当日', '滤布用量', 'filter_cloth_daily'),
    'high_temp_tape_daily': ('高温胶带日耗', '高温胶带当日', '高温胶带用量', 'high_temp_tape_daily'),
    'regen_oil_out': ('再生油出库', '再生油出', '再生油发出', 'regen_oil_out'),
    'regen_oil_in': ('再生油入库', '再生油入', '再生油回收', 'regen_oil_in'),
    'hydraulic_oil_daily': ('液压油日耗', '液压油当日', 'hydraulic_oil_daily'),
    'hydraulic_oil_monthly': ('液压油月累计', '液压油月累', 'hydraulic_oil_monthly'),
    'hydraulic_oil_target': ('液压油指标', '液压油定额', 'hydraulic_oil_target'),
    'gear_oil_daily': ('齿轮油日耗', '齿轮油当日', 'gear_oil_daily'),
    'gear_oil_monthly': ('齿轮油月累计', '齿轮油月累', 'gear_oil_monthly'),
    'gear_oil_target': ('齿轮油指标', '齿轮油定额', 'gear_oil_target'),
    'd40_per_ton': ('D40吨耗', 'd40吨耗', 'D40单吨消耗', 'd40_per_ton'),
    'steel_plate_per_ton': ('钢板吨耗', '钢板单吨消耗', '钢板每吨', 'steel_plate_per_ton'),
    'steel_buckle_per_ton': ('钢扣吨耗', '钢扣单吨消耗', '钢扣每吨', 'steel_buckle_per_ton'),
    'filter_agent_per_ton': ('飞滤剂吨耗', '飞滤剂单吨消耗', '飞滤剂每吨', 'filter_agent_per_ton'),
    'paint_per_ton': ('油漆吨耗', '油漆单吨消耗', '油漆每吨', 'paint_per_ton'),
    'ingot_block_count': ('铸锭块数', '铸锭数量', '锭块数', 'ingot_block_count'),
    'ingot_input_tons': ('铸锭投料量', '铸锭投料', '铸锭投入量', 'ingot_input_tons'),
    'ingot_output_tons': ('铸锭下机量', '铸锭产量', '铸锭产出量', 'ingot_output_tons'),
}
OWNER_DAILY_REFERENCE_FIELD_ALIASES = {
    'total_electricity_kwh': (
        '全厂用电',
        '全厂总用电',
        '全厂高压用电',
        '全厂高压总用电量',
        '高压总用电量',
        '总用电',
        'total_electricity_kwh',
    ),
    'new_plant_electricity_kwh': ('新厂用电', '新厂总用电', 'new_plant_electricity_kwh'),
    'park_electricity_kwh': ('园区用电', '园区总用电', 'park_electricity_kwh'),
    'cast_roll_gas_m3': ('铸轧用气', '铸轧天然气', 'cast_roll_gas_m3'),
    'smelting_gas_m3': ('熔炼炉用气', '熔炼炉天然气', 'smelting_gas_m3'),
    'heating_furnace_gas_m3': ('加热炉用气', '加热炉天然气', 'heating_furnace_gas_m3'),
    'boiler_gas_m3': ('锅炉用气', '锅炉天然气', 'boiler_gas_m3'),
    'total_gas_m3': ('天然气总量', '燃气总量', '全厂用气', '气耗共计', '用气共计', 'total_gas_m3'),
    'groundwater_ton': ('地下水', '地下水用量', 'groundwater_ton'),
    'tap_water_ton': ('自来水', '自来水用量', 'tap_water_ton'),
    'daily_contract_weight': ('当天接合同', '当日接合同', '当日合同', '日接合同', 'daily_contract_weight'),
    'daily_hot_roll_contract_weight': ('当日热轧合同', '热轧当日合同', '含热轧', 'daily_hot_roll_contract_weight'),
    'month_to_date_contract_weight': ('月累计合同', '合同月累计', '月累合同', 'month_to_date_contract_weight'),
    'month_to_date_hot_roll_contract_weight': (
        '月累计热轧合同',
        '热轧合同月累计',
        'month_to_date_hot_roll_contract_weight',
    ),
    'remaining_contract_weight': ('余合同量', '剩余合同量', 'remaining_contract_weight'),
    'remaining_hot_roll_contract_weight': ('余热轧合同', '剩余热轧合同', 'remaining_hot_roll_contract_weight'),
    'remaining_contract_delta_weight': (
        '余合同较昨日',
        '余合同变化',
        '剩余合同较昨日',
        'remaining_contract_delta_weight',
    ),
    'billet_inventory_weight': ('坯料总量', '坯料库存', '坯料结存', 'billet_inventory_weight'),
    'daily_input_weight': ('当日投料', '日投料', '当天冷轧投料', '冷轧投料', 'daily_input_weight'),
    'month_to_date_input_weight': ('月累计投料', '投料月累计', '月累投料', 'month_to_date_input_weight'),
}
COST_REFERENCE_FIELD_ALIASES = {
    'electricity_cost': ('电费', '用电成本', '电力成本', 'electricity_cost'),
    'natural_gas_cost': ('气费', '天然气费', '燃气费', '天然气成本', '燃气成本', 'natural_gas_cost'),
}
WIP_REFERENCE_FIELD_ALIASES = {
    'wip_total': ('当天在制料', '在制料总计', '在制料', '在制总量', 'wip_total'),
    'wip_1650_2050_cold': ('1650/2050冷轧', '1650和2050冷轧', 'wip_1650_2050_cold'),
    'wip_1850_cold': ('1850冷轧', 'wip_1850_cold'),
    'wip_milling': ('铣床', 'wip_milling'),
    'wip_anneal_total': ('退火分厂', '退火在制', 'wip_anneal_total'),
    'wip_new_north': ('新厂北线', '北线退火', 'wip_new_north'),
    'wip_new_south': ('新厂南线', '南线退火', 'wip_new_south'),
    'wip_park_anneal': ('园区退火', '园区在线退火', 'wip_park_anneal'),
    'wip_finishing_total': ('精整分厂', '精整在制', 'wip_finishing_total'),
    'wip_straightening': ('拉矫', 'wip_straightening'),
    'wip_finishing': ('精整', 'wip_finishing'),
    'wip_park_finishing': ('园区精整', 'wip_park_finishing'),
    'wip_hot_plate_shearing': ('热轧中厚板剪切', '中厚板剪切', 'wip_hot_plate_shearing'),
    'wip_coating': ('彩涂', 'wip_coating'),
}
REFERENCE_FIELD_ALIASES = {
    **CONSUMABLE_REFERENCE_FIELD_ALIASES,
    **OWNER_DAILY_REFERENCE_FIELD_ALIASES,
    **COST_REFERENCE_FIELD_ALIASES,
    **WIP_REFERENCE_FIELD_ALIASES,
}
NUMERIC_REFERENCE_FIELDS = {
    'input_tons',
    'output_tons',
    'energy_kwh',
    'scrap_tons',
    'downtime_minutes',
    'quality_issue_count',
    'yield_rate',
    'gas_m3',
    'rolling_oil_per_ton',
    'total_cost',
    'cost_per_ton',
    'throughput_cost_per_ton',
    *REFERENCE_FIELD_ALIASES.keys(),
}
DIFFERENCE_REASON_LABELS = {
    'value_diff': '数值不一致',
    'missing_system_row': '系统缺少同维度数据',
    'extra_system_row': '系统存在额外数据',
    'missing_field_value': '字段值缺失',
}


@dataclass(frozen=True, slots=True)
class MappingFieldSpec:
    metric: str
    reference_field: str
    system_field: str
    reference_unit: str | None = None
    system_unit: str | None = None
    tolerance: float = 0
    weight: float = 1


@dataclass(frozen=True, slots=True)
class MappingDifference:
    reason_code: str
    metric: str
    dimension: dict[str, Any]
    reference_value: float | str | None
    system_value: float | str | None
    diff_value: float | None
    suggested_rule: str


@dataclass(frozen=True, slots=True)
class MappingReconciliationResult:
    total_fields: int
    matched_fields: int
    overall_match_rate: float
    field_match_rates: dict[str, float]
    differences: list[MappingDifference]


def serialize_result(result: MappingReconciliationResult) -> dict[str, Any]:
    return asdict(result)


def summarize_differences(differences: Sequence[MappingDifference]) -> dict[str, Any]:
    by_reason_code: dict[str, int] = {}
    by_metric: dict[str, int] = {}
    for item in differences:
        by_reason_code[item.reason_code] = by_reason_code.get(item.reason_code, 0) + 1
        by_metric[item.metric] = by_metric.get(item.metric, 0) + 1
    return {
        'total': len(differences),
        'by_reason_code': by_reason_code,
        'by_metric': by_metric,
        'reason_breakdown': [
            {
                'reason_code': reason_code,
                'label': DIFFERENCE_REASON_LABELS.get(reason_code, reason_code),
                'count': count,
            }
            for reason_code, count in by_reason_code.items()
        ],
    }


def summarize_match_result(result: MappingReconciliationResult) -> dict[str, Any]:
    return {
        'total_fields': result.total_fields,
        'matched_fields': result.matched_fields,
        'unmatched_fields': max(result.total_fields - result.matched_fields, 0),
        'overall_match_rate': result.overall_match_rate,
        'field_breakdown': [
            {'metric': metric, 'match_rate': match_rate}
            for metric, match_rate in result.field_match_rates.items()
        ],
    }


def _reference_root() -> Path:
    configured = os.getenv('OUTPUT_SKILL_REFERENCE_ROOT')
    if configured:
        return Path(configured)
    if DEFAULT_REFERENCE_ROOT.exists():
        return DEFAULT_REFERENCE_ROOT
    return FALLBACK_REFERENCE_ROOT


def resolve_reference_file(reference_file: str | Path, *, reference_root: str | Path | None = None) -> Path:
    root = Path(reference_root) if reference_root is not None else _reference_root()
    resolved_root = root.resolve()
    incoming = Path(reference_file)
    candidate = incoming if incoming.is_absolute() else resolved_root / incoming
    resolved_candidate = candidate.resolve()
    try:
        resolved_candidate.relative_to(resolved_root)
    except ValueError as exc:
        raise ValueError('reference_file must stay inside output skill reference root') from exc
    if not resolved_candidate.exists():
        raise ValueError('reference_file not found inside output skill reference root')
    return resolved_candidate


def _reference_parse_status(path: Path) -> str:
    suffix = path.suffix.lower()
    if suffix in PARSEABLE_REFERENCE_EXTENSIONS:
        return 'parseable'
    if suffix in IMAGE_REFERENCE_EXTENSIONS:
        return 'image_pending_ocr'
    return 'unsupported'


def _is_reference_file_visible(path: Path, root: Path) -> bool:
    try:
        relative = path.relative_to(root)
    except ValueError:
        return False
    return not any(part.startswith('.') or part in IGNORED_REFERENCE_DIR_NAMES for part in relative.parts[:-1])


def list_sources(*, reference_root: str | Path | None = None, limit: int = 200) -> dict[str, Any]:
    root = Path(reference_root) if reference_root is not None else _reference_root()
    all_files: list[dict[str, Any]] = []
    if root.exists():
        candidates = (
            path
            for path in root.rglob('*')
            if path.is_file() and _is_reference_file_visible(path, root)
        )
        for item in sorted(candidates, key=lambda path: str(path)):
            parse_status = _reference_parse_status(item)
            all_files.append(
                {
                    'name': item.name,
                    'relative_path': str(item.relative_to(root)).replace('\\', '/'),
                    'extension': item.suffix.lower(),
                    'size_bytes': item.stat().st_size,
                    'parse_status': parse_status,
                }
            )
    total_files = len(all_files)
    parseable_files = sum(1 for item in all_files if item.get('parse_status') == 'parseable')
    image_pending_files = sum(1 for item in all_files if item.get('parse_status') == 'image_pending_ocr')
    unsupported_files = max(total_files - parseable_files - image_pending_files, 0)
    return {
        'reference_source': str(root),
        'available': root.exists(),
        'files': all_files[: max(int(limit), 0)],
        'file_summary': {
            'total_files': total_files,
            'parseable_files': parseable_files,
            'image_pending_files': image_pending_files,
            'unsupported_files': unsupported_files,
            'parseable_coverage_rate': round((parseable_files / total_files) * 100, 2) if total_files else 0,
        },
        'system_sources': SYSTEM_SOURCES,
    }


def _to_date_text(value: Any) -> str | None:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value or '').strip()
    if not text:
        return None
    match = DATE_RE.search(text)
    if match:
        year, month, day = (int(part) for part in match.groups())
        return date(year, month, day).isoformat()
    try:
        return date.fromisoformat(text[:10]).isoformat()
    except ValueError:
        return None


def _read_reference_text(path: Path) -> str:
    for encoding in ('utf-8-sig', 'utf-8', 'gbk'):
        try:
            return path.read_text(encoding=encoding)
        except UnicodeDecodeError:
            continue
    return path.read_text(encoding='utf-8', errors='ignore')


def _metric_tons(line: str, labels: Sequence[str]) -> float | None:
    label_pattern = '|'.join(re.escape(label) for label in labels)
    match = re.search(rf'(?:{label_pattern})\s*{NUMBER_RE}\s*(吨|t|T|kg|KG|公斤|千克)?', line)
    if not match:
        return None
    value = float(match.group(1))
    unit = (match.group(2) or '吨').lower()
    if unit in {'kg', '公斤', '千克'}:
        return value / 1000
    return value


def _metric_number(line: str, labels: Sequence[str]) -> float | None:
    label_pattern = '|'.join(re.escape(label) for label in labels)
    match = re.search(rf'(?:{label_pattern})\s*约?\s*{NUMBER_RE}', line, flags=re.IGNORECASE)
    return float(match.group(1)) if match else None


def _metric_money(line: str, labels: Sequence[str]) -> float | None:
    label_pattern = '|'.join(re.escape(label) for label in sorted(labels, key=len, reverse=True))
    match = re.search(rf'(?:{label_pattern})\s*约?\s*{NUMBER_RE}\s*(万元|万|元)?', line, flags=re.IGNORECASE)
    if not match:
        return None
    value = float(match.group(1))
    unit = match.group(2) or '元'
    if unit in {'万元', '万'}:
        return round(value * 10000, 6)
    return value


def _metric_minutes(line: str, labels: Sequence[str]) -> float | None:
    label_pattern = '|'.join(re.escape(label) for label in sorted(labels, key=len, reverse=True))
    match = re.search(rf'(?:{label_pattern})\s*{NUMBER_RE}\s*(小时|h|H|分钟|min|MIN)?', line)
    if not match:
        return None
    value = float(match.group(1))
    unit = match.group(2) or '分钟'
    if unit in {'小时', 'h', 'H'}:
        return value * 60
    return value


def _line_workshop_shift(line: str) -> tuple[str | None, str | None]:
    for shift in SHIFT_NAMES:
        if shift in line:
            workshop = line.split(shift, 1)[0].strip(' ：:，,')
            return workshop.replace(' ', '') or None, shift
    return None, None


def _add_text_source(row: dict[str, Any], path: Path) -> dict[str, Any]:
    row['source_file'] = str(path)
    row['source_type'] = 'output_skill_text'
    return row


def _factory_narrative_row(line: str, current_date: str, path: Path) -> dict[str, Any] | None:
    row: dict[str, Any] = {
        'business_date': current_date,
        'workshop': '全厂',
        'shift': '',
    }
    for field, aliases in REFERENCE_FIELD_ALIASES.items():
        value = _metric_number(line, aliases)
        if value is not None:
            row[field] = value

    if 'total_gas_m3' not in row and ('用气' in line or '天然气' in line):
        total_gas_m3 = _metric_number(line, ('天然气合计', '用气合计', '燃气合计', '共计'))
        if total_gas_m3 is not None:
            row['total_gas_m3'] = total_gas_m3

    electricity_cost = _metric_money(line, COST_REFERENCE_FIELD_ALIASES['electricity_cost'])
    natural_gas_cost = _metric_money(line, COST_REFERENCE_FIELD_ALIASES['natural_gas_cost'])
    total_cost = _metric_money(line, ('已核合计', '总成本', '成本合计', '总费用', 'total_cost'))
    cost_per_ton = _metric_number(line, ('综合吨成本', '单吨成本', '吨成本', '成本/吨', '折算', 'cost_per_ton'))
    if electricity_cost is not None:
        row['electricity_cost'] = electricity_cost
    if natural_gas_cost is not None:
        row['natural_gas_cost'] = natural_gas_cost
    if total_cost is not None:
        row['total_cost'] = total_cost
    if cost_per_ton is not None:
        row['cost_per_ton'] = cost_per_ton
    if len(row) > 3:
        return _add_text_source(row, path)
    return None


def _parse_text_rows(path: Path) -> list[dict[str, Any]]:
    content = _read_reference_text(path)
    current_date: str | None = _to_date_text(path.name)
    rows: list[dict[str, Any]] = []
    for raw_line in content.splitlines():
        line = raw_line.strip()
        if not line:
            continue
        line_date = _to_date_text(line)
        if line_date:
            current_date = line_date

        workshop, shift = _line_workshop_shift(line)
        if not current_date or not workshop or not shift:
            if current_date and (row := _factory_narrative_row(line, current_date, path)) is not None:
                rows.append(row)
            continue

        row: dict[str, Any] = {
            'business_date': current_date,
            'workshop': workshop,
            'shift': shift,
        }
        input_tons = _metric_tons(line, ('投入量', '投入重量', '投料量', '投料', '上机重量', '上机量', 'input_tons'))
        output_tons = _metric_tons(line, ('产量', '下机量', '包装产量', '入库量'))
        energy_kwh = _metric_number(line, ('能耗', '电量', '用电', '总电气', '总用电'))
        scrap_tons = _metric_tons(line, ('废料', '废品', '废料量'))
        downtime_minutes = _metric_minutes(line, ('停机时长', '停机时间', '停机'))
        quality_issue_count = _metric_number(line, ('质量异常', '质量问题', '质量门禁', '异常数'))
        yield_rate = _metric_number(line, ('成材率', '成品率', '良品率', '得率'))
        gas_m3 = _metric_number(line, ('燃气', '用气', '气量', '天然气', 'gas_m3'))
        rolling_oil_per_ton = _metric_number(line, ('轧制油吨耗', '轧制油单吨消耗', '轧制油每吨', 'rolling_oil_per_ton'))
        total_cost = _metric_number(line, ('总成本', '成本合计', '总费用', 'total_cost'))
        throughput_cost_per_ton = _metric_number(
            line,
            ('过站吨成本', '流转吨成本', '吞吐吨成本', '过工序吨成本', 'throughput_cost_per_ton'),
        )
        cost_per_ton = _metric_number(line, ('综合吨成本', '单吨成本', '吨成本', '成本/吨', 'cost_per_ton'))
        if input_tons is not None:
            row['input_tons'] = input_tons
        if output_tons is not None:
            row['output_tons'] = output_tons
        if energy_kwh is not None:
            row['energy_kwh'] = energy_kwh
        if scrap_tons is not None:
            row['scrap_tons'] = scrap_tons
        if downtime_minutes is not None:
            row['downtime_minutes'] = downtime_minutes
        if quality_issue_count is not None:
            row['quality_issue_count'] = quality_issue_count
        if yield_rate is not None:
            row['yield_rate'] = yield_rate
        if gas_m3 is not None:
            row['gas_m3'] = gas_m3
        if rolling_oil_per_ton is not None:
            row['rolling_oil_per_ton'] = rolling_oil_per_ton
        for field, aliases in REFERENCE_FIELD_ALIASES.items():
            value = _metric_number(line, aliases)
            if value is not None:
                row[field] = value
        if total_cost is not None:
            row['total_cost'] = total_cost
        if throughput_cost_per_ton is not None:
            row['throughput_cost_per_ton'] = throughput_cost_per_ton
        if cost_per_ton is not None:
            row['cost_per_ton'] = cost_per_ton
        if len(row) > 3:
            rows.append(_add_text_source(row, path))
    return rows


def _normalize_header(value: Any) -> str:
    return _normalize_text(value).lower().replace('吨', 'ton').replace('度', 'kwh')


def _excel_field(header: str) -> str | None:
    if header in {'日期', '生产日', '业务日'}:
        return 'business_date'
    if header in {'车间', '部门', '工厂'}:
        return 'workshop'
    if header == '班次':
        return 'shift'
    if header in {'机台', '机列', '设备', '设备名称', '设备名'}:
        return 'machine'
    if header in {'工序', '工艺', '当前工艺'}:
        return 'process'
    if header in {'卷号', '随行卡号', '随行卡', '卡号', '批号'}:
        return 'coil_no'
    if header in {'合同号', '合同', '合同编号'}:
        return 'contract_no'
    if header in {'客户', '客户名', '客户名称'}:
        return 'customer'
    for field, aliases in REFERENCE_FIELD_ALIASES.items():
        if field in header or any(_normalize_header(alias) in header for alias in aliases):
            return field
    if (
        '投入' in header
        or '投料' in header
        or '上机' in header
        or '来料' in header
        or 'input_tons' in header
    ):
        return 'input_tons'
    if '能耗' in header or '电量' in header or 'kwh' in header:
        return 'energy_kwh'
    if '燃气' in header or '然气' in header or '用气' in header or '气量' in header or '天然气' in header or 'gas_m3' in header:
        return 'gas_m3'
    if '废料' in header or '废品' in header:
        return 'scrap_tons'
    if '停机' in header:
        return 'downtime_minutes'
    if '质量异常' in header or '质量问题' in header or '异常数' in header or 'quality' in header:
        return 'quality_issue_count'
    if '成材率' in header or '成品率' in header or '良品率' in header or '得率' in header or 'yield' in header:
        return 'yield_rate'
    if '轧制油' in header and (
        '吨耗' in header or 'ton耗' in header or '单吨' in header or '单ton' in header or '每吨' in header or '每ton' in header
    ):
        return 'rolling_oil_per_ton'
    if 'rolling_oil_per_ton' in header:
        return 'rolling_oil_per_ton'
    if '总成本' in header or '成本合计' in header or '总费用' in header or 'total_cost' in header:
        return 'total_cost'
    if (
        '过站吨成本' in header
        or '过站ton成本' in header
        or '流转吨成本' in header
        or '流转ton成本' in header
        or '吞吐吨成本' in header
        or '吞吐ton成本' in header
        or '过工序吨成本' in header
        or '过工序ton成本' in header
        or 'throughput_cost_per_ton' in header
    ):
        return 'throughput_cost_per_ton'
    if '成本' in header and (
        'ton' in header or '单ton' in header or '每ton' in header or '元/ton' in header or 'cost_per_ton' in header
    ):
        return 'cost_per_ton'
    if 'cost_per_ton' in header:
        return 'cost_per_ton'
    if '产量' in header or '下机量' in header or '入库量' in header or '包装' in header:
        return 'output_tons'
    return None


def _normalize_reference_number(field: str, value: Any) -> float | None:
    number = _to_float(value)
    if number is None:
        return None
    if field == 'yield_rate' and 0 < number <= 1:
        return round(number * 100, 6)
    return number


REFERENCE_DIMENSION_FIELDS = {
    'business_date',
    'workshop',
    'shift',
    'machine',
    'machine_code',
    'process',
    'coil_no',
    'contract_no',
    'customer',
    'source_file',
    'source_type',
}


def _reference_text(value: Any) -> str:
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _reference_row_has_metric(row: Mapping[str, Any]) -> bool:
    return any(field not in REFERENCE_DIMENSION_FIELDS and value not in (None, '') for field, value in row.items())


def _finalize_reference_row(row: dict[str, Any], *, path: Path, source_type: str) -> dict[str, Any] | None:
    if not (row.get('business_date') and row.get('workshop') and _reference_row_has_metric(row)):
        return None
    row.setdefault('shift', '')
    row['source_file'] = str(path)
    row['source_type'] = source_type
    return row


def _default_business_date(sheet_name: Any, path: Path) -> str | None:
    return _to_date_text(sheet_name) or _to_date_text(path.name)


def _header_fields(values: Sequence[Any]) -> dict[int, str]:
    return {
        index: field
        for index, header in enumerate(values)
        if (field := _excel_field(_normalize_header(header))) is not None
    }


def _header_row_index(rows: Sequence[Sequence[Any]]) -> int | None:
    for index, values in enumerate(rows):
        fields = set(_header_fields(values).values())
        if 'workshop' in fields and _reference_row_has_metric({field: 1 for field in fields}):
            return index
    return None


def _parse_excel_rows(path: Path) -> list[dict[str, Any]]:
    if path.suffix.lower() == '.xls':
        return _parse_xls_rows(path)

    workbook = load_workbook(path, read_only=True, data_only=True)
    parsed_rows: list[dict[str, Any]] = []
    for sheet in workbook.worksheets:
        sheet_rows = [tuple(values) for values in sheet.iter_rows(values_only=True)]
        header_index = _header_row_index(sheet_rows)
        if header_index is None:
            continue
        field_by_index = _header_fields(sheet_rows[header_index])
        default_date = _default_business_date(sheet.title, path)
        current_process = ''
        for values in sheet_rows[header_index + 1 :]:
            row: dict[str, Any] = {}
            for index, value in enumerate(values):
                field = field_by_index.get(index)
                if not field or value in (None, ''):
                    continue
                if field in row:
                    continue
                if field == 'business_date':
                    row[field] = _to_date_text(value)
                elif field in NUMERIC_REFERENCE_FIELDS:
                    number = _normalize_reference_number(field, value)
                    if number is not None:
                        row[field] = number
                else:
                    row[field] = _reference_text(value)
            if not row.get('business_date') and default_date:
                row['business_date'] = default_date
            if row.get('process'):
                current_process = str(row['process'])
            elif current_process:
                row['process'] = current_process
            if finalized := _finalize_reference_row(row, path=path, source_type='output_skill_excel'):
                parsed_rows.append(finalized)
    workbook.close()
    return parsed_rows


def _parse_xls_rows(path: Path) -> list[dict[str, Any]]:
    import xlrd

    workbook = xlrd.open_workbook(str(path))
    parsed_rows: list[dict[str, Any]] = []
    for sheet in workbook.sheets():
        if sheet.nrows < 1:
            continue
        sheet_rows = [tuple(sheet.row_values(row_index)) for row_index in range(sheet.nrows)]
        header_index = _header_row_index(sheet_rows)
        if header_index is None:
            continue
        field_by_index = _header_fields(sheet_rows[header_index])
        default_date = _default_business_date(sheet.name, path)
        current_process = ''
        for row_index in range(header_index + 1, sheet.nrows):
            row: dict[str, Any] = {}
            for index, value in enumerate(sheet.row_values(row_index)):
                field = field_by_index.get(index)
                if not field or value in (None, ''):
                    continue
                if field in row:
                    continue
                if field == 'business_date':
                    cell = sheet.cell(row_index, index)
                    if cell.ctype == xlrd.XL_CELL_DATE:
                        row[field] = xlrd.xldate.xldate_as_datetime(value, workbook.datemode).date().isoformat()
                    else:
                        row[field] = _to_date_text(value)
                elif field in NUMERIC_REFERENCE_FIELDS:
                    number = _normalize_reference_number(field, value)
                    if number is not None:
                        row[field] = number
                else:
                    row[field] = _reference_text(value)
            if not row.get('business_date') and default_date:
                row['business_date'] = default_date
            if row.get('process'):
                current_process = str(row['process'])
            elif current_process:
                row['process'] = current_process
            if finalized := _finalize_reference_row(row, path=path, source_type='output_skill_excel'):
                parsed_rows.append(finalized)
    return parsed_rows


def _parse_delivery_override_rows(payload: Mapping[str, Any], path: Path) -> list[dict[str, Any]]:
    current_date = _to_date_text(path.name)
    summaries = payload.get('summaries')
    if not current_date or not isinstance(summaries, Mapping):
        return []

    rows: list[dict[str, Any]] = []
    for value in summaries.values():
        if not isinstance(value, str):
            continue
        row = _factory_narrative_row(value, current_date, path)
        if not row:
            continue
        row['source_type'] = 'output_skill_json'
        rows.append(row)
    return rows


def _json_reference_records(payload: Any) -> list[Mapping[str, Any]]:
    if isinstance(payload, list):
        return [item for item in payload if isinstance(item, Mapping)]
    if isinstance(payload, Mapping):
        for key in ('rows', 'items', 'data', 'records'):
            value = payload.get(key)
            if isinstance(value, list):
                return [item for item in value if isinstance(item, Mapping)]
        return [payload]
    return []


def _json_record_rows(records: Iterable[Mapping[str, Any]], *, path: Path, source_type: str) -> list[dict[str, Any]]:
    default_date = _to_date_text(path.name)
    parsed_rows: list[dict[str, Any]] = []
    for item in records:
        row: dict[str, Any] = {}
        for header, value in item.items():
            field = _excel_field(_normalize_header(header))
            if not field or value in (None, ''):
                continue
            if field == 'business_date':
                row[field] = _to_date_text(value)
            elif field in NUMERIC_REFERENCE_FIELDS:
                number = _normalize_reference_number(field, value)
                if number is not None:
                    row[field] = number
            else:
                row[field] = str(value).strip()
        if not row.get('business_date') and default_date:
            row['business_date'] = default_date
        if finalized := _finalize_reference_row(row, path=path, source_type=source_type):
            parsed_rows.append(finalized)
    return parsed_rows


def _parse_json_rows(path: Path) -> list[dict[str, Any]]:
    payload = json.loads(_read_reference_text(path))
    rows = _json_record_rows(_json_reference_records(payload), path=path, source_type='output_skill_json')
    if not rows and isinstance(payload, Mapping):
        rows = _parse_delivery_override_rows(payload, path)
    return rows


def _parse_json_lines_rows(path: Path) -> list[dict[str, Any]]:
    records: list[Mapping[str, Any]] = []
    for line in _read_reference_text(path).splitlines():
        text = line.strip()
        if not text:
            continue
        payload = json.loads(text)
        if isinstance(payload, Mapping):
            records.append(payload)
    return _json_record_rows(records, path=path, source_type='output_skill_json_lines')


def parse_output_skill_reference_file(file_path: str | Path) -> dict[str, Any]:
    path = Path(file_path)
    suffix = path.suffix.lower()
    if suffix in TEXT_EXTENSIONS:
        rows = _parse_text_rows(path)
        source_type = 'output_skill_text'
    elif suffix in EXCEL_EXTENSIONS:
        rows = _parse_excel_rows(path)
        source_type = 'output_skill_excel'
    elif suffix in JSON_EXTENSIONS:
        rows = _parse_json_rows(path)
        source_type = 'output_skill_json'
    elif suffix in JSON_LINES_EXTENSIONS:
        rows = _parse_json_lines_rows(path)
        source_type = 'output_skill_json_lines'
    else:
        return {
            'status': 'unsupported',
            'source_file': str(path),
            'source_type': 'unsupported',
            'rows': [],
            'issues': [{'code': 'unsupported_extension', 'extension': suffix}],
        }
    return {
        'status': 'parsed',
        'source_file': str(path),
        'source_type': source_type,
        'rows': rows,
        'issues': [],
    }


def _normalize_text(value: Any) -> str:
    return str(value or '').strip().replace(' ', '').replace('（', '(').replace('）', ')')


def _dimension_aliases(field: str, aliases: Mapping[str, Mapping[str, str]] | None) -> dict[str, str]:
    merged = {
        _normalize_text(source): _normalize_text(target)
        for source, target in DEFAULT_DIMENSION_ALIASES.get(field, {}).items()
    }
    if aliases and field in aliases:
        merged.update({_normalize_text(source): _normalize_text(target) for source, target in aliases[field].items()})
    return merged


def _normalize_dimension(
    field: str,
    value: Any,
    aliases: Mapping[str, Mapping[str, str]] | None,
) -> str:
    normalized = _normalize_text(value)
    field_aliases = _dimension_aliases(field, aliases)
    return _normalize_text(field_aliases.get(normalized, normalized))


def _dimension(row: Mapping[str, Any], dimensions: Sequence[str], aliases: Mapping[str, Mapping[str, str]] | None) -> dict[str, Any]:
    return {field: _normalize_dimension(field, row.get(field), aliases) for field in dimensions}


def _dimension_key(dimension: Mapping[str, Any], dimensions: Sequence[str]) -> tuple[str, ...]:
    return tuple(str(dimension.get(field) or '') for field in dimensions)


def _to_float(value: Any) -> float | None:
    if value is None or value == '':
        return None
    if isinstance(value, Decimal):
        return float(value)
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _to_common_unit(value: Any, unit: str | None) -> float | None:
    number = _to_float(value)
    if number is None:
        return None
    normalized_unit = _normalize_text(unit).lower()
    if normalized_unit in {'kg', '公斤', '千克'}:
        return number / 1000
    return number


def _round_rate(value: float) -> float:
    rounded = round(value, 2)
    return int(rounded) if rounded.is_integer() else rounded


def _build_index(
    rows: Iterable[Mapping[str, Any]],
    *,
    dimensions: Sequence[str],
    aliases: Mapping[str, Mapping[str, str]] | None,
) -> dict[tuple[str, ...], Mapping[str, Any]]:
    indexed: dict[tuple[str, ...], dict[str, Any]] = {}
    for row in rows:
        dimension = _dimension(row, dimensions, aliases)
        key = _dimension_key(dimension, dimensions)
        if key not in indexed:
            indexed[key] = dict(row)
            continue
        existing = indexed[key]
        for field, value in row.items():
            if existing.get(field) in (None, '') and value not in (None, ''):
                existing[field] = value
    return indexed


def _value_diff_message(metric: str) -> str:
    return f'检查 {metric} 字段口径、单位或时间范围。'


def compare_mapping_rows(
    *,
    reference_rows: Sequence[Mapping[str, Any]],
    system_rows: Sequence[Mapping[str, Any]],
    fields: Sequence[MappingFieldSpec],
    dimensions: Sequence[str] = DEFAULT_DIMENSIONS,
    dimension_aliases: Mapping[str, Mapping[str, str]] | None = None,
) -> MappingReconciliationResult:
    reference_index = _build_index(reference_rows, dimensions=dimensions, aliases=dimension_aliases)
    system_index = _build_index(system_rows, dimensions=dimensions, aliases=dimension_aliases)

    differences: list[MappingDifference] = []
    matched_fields = 0
    total_fields = 0
    metric_totals: dict[str, float] = {}
    metric_matches: dict[str, float] = {}
    total_weight = 0.0
    matched_weight = 0.0

    for key, reference_row in reference_index.items():
        reference_dimension = _dimension(reference_row, dimensions, dimension_aliases)
        system_row = system_index.get(key)
        if system_row is None:
            for field in fields:
                total_fields += 1
                total_weight += field.weight
                metric_totals[field.metric] = metric_totals.get(field.metric, 0) + field.weight
                differences.append(
                    MappingDifference(
                        reason_code='missing_system_row',
                        metric=field.metric,
                        dimension=reference_dimension,
                        reference_value=reference_row.get(field.reference_field),
                        system_value=None,
                        diff_value=None,
                        suggested_rule='系统侧缺少同日期、车间、班次的数据行。',
                    )
                )
            continue

        for field in fields:
            total_fields += 1
            total_weight += field.weight
            metric_totals[field.metric] = metric_totals.get(field.metric, 0) + field.weight
            reference_value = _to_common_unit(reference_row.get(field.reference_field), field.reference_unit)
            system_value = _to_common_unit(system_row.get(field.system_field), field.system_unit)
            if reference_value is None or system_value is None:
                differences.append(
                    MappingDifference(
                        reason_code='missing_field_value',
                        metric=field.metric,
                        dimension=reference_dimension,
                        reference_value=reference_value,
                        system_value=system_value,
                        diff_value=None,
                        suggested_rule=f'检查 {field.metric} 字段是否缺值或字段名是否映射错误。',
                    )
                )
                continue

            diff_value = round(system_value - reference_value, 6)
            if abs(diff_value) <= field.tolerance:
                matched_fields += 1
                matched_weight += field.weight
                metric_matches[field.metric] = metric_matches.get(field.metric, 0) + field.weight
                continue

            differences.append(
                MappingDifference(
                    reason_code='value_diff',
                    metric=field.metric,
                    dimension=reference_dimension,
                    reference_value=reference_value,
                    system_value=system_value,
                    diff_value=diff_value,
                    suggested_rule=_value_diff_message(field.metric),
                )
            )

    for key, system_row in system_index.items():
        if key in reference_index:
            continue
        system_dimension = _dimension(system_row, dimensions, dimension_aliases)
        differences.append(
            MappingDifference(
                reason_code='extra_system_row',
                metric='dimension',
                dimension=system_dimension,
                reference_value=None,
                system_value=None,
                diff_value=None,
                suggested_rule='系统侧有额外数据行，需确认是否日期、车间、班次别名未对齐。',
            )
        )

    field_match_rates = {
        metric: _round_rate((metric_matches.get(metric, 0) / total) * 100) if total else 0
        for metric, total in metric_totals.items()
    }
    overall_match_rate = _round_rate((matched_weight / total_weight) * 100) if total_weight else 0
    return MappingReconciliationResult(
        total_fields=total_fields,
        matched_fields=matched_fields,
        overall_match_rate=overall_match_rate,
        field_match_rates=field_match_rates,
        differences=differences,
    )


def propose_rules(differences: Sequence[MappingDifference]) -> list[dict[str, Any]]:
    missing = [item for item in differences if item.reason_code == 'missing_system_row']
    extra = [item for item in differences if item.reason_code == 'extra_system_row']
    if not missing or not extra:
        return []

    proposals: list[dict[str, Any]] = []
    missing_dimension = missing[0].dimension
    extra_dimension = extra[0].dimension
    for field in ('workshop', 'shift', 'machine', 'process'):
        reference_value = missing_dimension.get(field)
        system_value = extra_dimension.get(field)
        if reference_value and system_value and reference_value != system_value:
            proposals.append(
                {
                    'rule_type': 'alias_candidate',
                    'field': field,
                    'reference_value': reference_value,
                    'system_value': system_value,
                    'confidence': 'manual_review',
                    'dry_run': True,
                }
            )
    return proposals


def _tons_from_pair(tons: Any, kg: Any) -> float | None:
    tons_value = _to_float(tons)
    if tons_value is not None:
        return tons_value
    kg_value = _to_float(kg)
    if kg_value is None:
        return None
    return kg_value / 1000


def _shift_production_weight_tons(record: ShiftProductionData, field_name: str) -> float | None:
    value = _to_float(getattr(record, field_name, None))
    if value is None:
        return None
    if record.data_source == 'mobile_coil_agg':
        return value / 1000
    return value


def _wip_total_snapshot_weight_tons(value: Any) -> float | None:
    number = _to_float(value)
    if number is None:
        return None
    return number / 1000


def _consumable_payload_metrics(payload: Mapping[str, Any] | None) -> dict[str, Any]:
    if not payload:
        return {}
    known_metrics = {
        'packaging_inbound_output_tons',
        'electricity_daily',
        'gas_daily',
    }
    return {
        key: _to_float(value)
        for key, value in payload.items()
        if key not in known_metrics and _to_float(value) is not None
    }


def _owner_daily_payload_metrics(payload: Mapping[str, Any] | None) -> dict[str, Any]:
    if not payload:
        return {}
    return {
        key: value
        for key in OWNER_DAILY_REFERENCE_FIELD_ALIASES
        if (value := _to_float(payload.get(key))) is not None
    }


def _latest_wip_total_by_workshop(db: Session, *, business_date: date) -> dict[str, dict[str, Any]]:
    try:
        start_at, end_at = production_business_window(business_date)
        query = db.query(MesWipTotalSnapshot).filter(
            MesWipTotalSnapshot.snapshot_at >= start_at,
            MesWipTotalSnapshot.snapshot_at < end_at,
        )
        if query.limit(1).first() is None:
            return {}

        rows = (
            query.with_entities(
                MesWipTotalSnapshot.workshop_name,
                func.sum(MesWipTotalSnapshot.doing_count),
                func.sum(MesWipTotalSnapshot.doing_weight_tons),
            )
            .group_by(MesWipTotalSnapshot.workshop_name)
            .all()
        )
    except (OperationalError, ProgrammingError):
        return {}

    result: dict[str, dict[str, Any]] = {}
    for workshop, count, weight in rows:
        key = _normalize_text(workshop)
        if not key:
            continue
        result[key] = {
            'workshop': str(workshop or ''),
            'coil_count': int(count or 0),
            'wip_total': _wip_total_snapshot_weight_tons(weight) or 0.0,
        }
    return result


def _build_wip_mapping_rows(db: Session, *, business_date: date) -> list[dict[str, Any]]:
    fallback_by_workshop = _latest_wip_total_by_workshop(db, business_date=business_date)
    try:
        daily_rows = (
            db.query(
                MesDailyWipSnapshot.workshop_name,
                func.sum(MesDailyWipSnapshot.coil_count),
                func.sum(MesDailyWipSnapshot.material_weight_tons),
                func.sum(MesDailyWipSnapshot.feeding_weight_tons),
            )
            .filter(MesDailyWipSnapshot.business_date == business_date)
            .group_by(MesDailyWipSnapshot.workshop_name)
            .all()
        )
    except (OperationalError, ProgrammingError):
        daily_rows = []

    rows: list[dict[str, Any]] = []
    used_workshops: set[str] = set()
    for workshop, count, weight, feeding_weight in daily_rows:
        key = _normalize_text(workshop)
        fallback = fallback_by_workshop.get(key)
        total_weight = _to_float(weight)
        source_basis = 'mes_daily_wip_snapshots'
        coil_count = int(count or 0)
        if (total_weight is None or total_weight <= 0) and fallback and fallback['wip_total'] > 0:
            total_weight = fallback['wip_total']
            coil_count = fallback['coil_count'] or coil_count
            source_basis = 'mes_wip_total_snapshots'
        row = {
            'business_date': business_date.isoformat(),
            'workshop': str(workshop or ''),
            'shift': '',
            'process': '在制料',
            'machine': '',
            'wip_coil_count': coil_count,
            'wip_feeding_tons': _to_float(feeding_weight) or 0.0,
            'source_table': source_basis,
        }
        if total_weight is not None and total_weight > 0:
            row['wip_total'] = round(total_weight, 4)
        else:
            row['wip_source_issue'] = 'missing_material_weight_for_business_date'
        rows.append(row)
        used_workshops.add(key)

    for key, fallback in fallback_by_workshop.items():
        if key in used_workshops or fallback['wip_total'] <= 0:
            continue
        rows.append(
            {
                'business_date': business_date.isoformat(),
                'workshop': fallback['workshop'],
                'shift': '',
                'process': '在制料',
                'machine': '',
                'wip_total': round(fallback['wip_total'], 4),
                'wip_coil_count': fallback['coil_count'],
                'wip_feeding_tons': 0.0,
                'source_table': 'mes_wip_total_snapshots',
            }
        )
    return rows


def _sum_metric(rows: Sequence[Mapping[str, Any]], field: str, *, source_table: str | None = None) -> float | None:
    values = [
        number
        for row in rows
        if (source_table is None or row.get('source_table') == source_table)
        if (number := _to_float(row.get(field))) is not None
    ]
    if not values:
        return None
    return round(sum(values), 4)


def _build_factory_summary_row(rows: Sequence[Mapping[str, Any]], *, business_date: date) -> dict[str, Any] | None:
    summary: dict[str, Any] = {
        'business_date': business_date.isoformat(),
        'workshop': '全厂',
        'shift': '',
        'process': '汇总',
        'machine': '',
    }
    metric_sources = {
        'wip_total': [('wip_total', None)],
        'total_electricity_kwh': [
            ('total_electricity_kwh', None),
            ('electricity_kwh', 'machine_daily_cost_snapshots'),
            ('energy_kwh', 'machine_energy_records'),
        ],
        'total_gas_m3': [
            ('total_gas_m3', None),
            ('natural_gas_m3', 'machine_daily_cost_snapshots'),
            ('gas_m3', 'machine_energy_records'),
        ],
        'electricity_cost': [('electricity_cost', 'machine_daily_cost_snapshots')],
        'natural_gas_cost': [('natural_gas_cost', 'machine_daily_cost_snapshots')],
        'total_cost': [('total_cost', 'machine_daily_cost_snapshots')],
        'daily_contract_weight': [('daily_contract_weight', None)],
        'daily_hot_roll_contract_weight': [('daily_hot_roll_contract_weight', None)],
        'month_to_date_contract_weight': [('month_to_date_contract_weight', None)],
        'remaining_contract_weight': [('remaining_contract_weight', None)],
        'remaining_hot_roll_contract_weight': [('remaining_hot_roll_contract_weight', None)],
        'remaining_contract_delta_weight': [('remaining_contract_delta_weight', None)],
        'billet_inventory_weight': [('billet_inventory_weight', None)],
        'daily_input_weight': [('daily_input_weight', None)],
        'month_to_date_input_weight': [('month_to_date_input_weight', None)],
    }
    for target_field, sources in metric_sources.items():
        for source_field, source_table in sources:
            value = _sum_metric(rows, source_field, source_table=source_table)
            if value is not None:
                summary[target_field] = value
                break

    if not _reference_row_has_metric(summary):
        return None
    summary['source_table'] = 'mapping_reconciliation_summary'
    return summary


def build_system_mapping_rows(db: Session, *, business_date: date) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    shift_rows = (
        db.query(ShiftProductionData, Workshop, ShiftConfig, Equipment)
        .join(Workshop, ShiftProductionData.workshop_id == Workshop.id)
        .join(ShiftConfig, ShiftProductionData.shift_config_id == ShiftConfig.id)
        .outerjoin(Equipment, ShiftProductionData.equipment_id == Equipment.id)
        .filter(
            ShiftProductionData.business_date == business_date,
            ShiftProductionData.data_status != 'voided',
        )
        .order_by(ShiftProductionData.id.asc())
        .all()
    )
    for record, workshop, shift, equipment in shift_rows:
        rows.append(
            {
                'business_date': record.business_date.isoformat(),
                'workshop': workshop.name,
                'shift': shift.name,
                'process': '班次产量',
                'machine': equipment.name if equipment else '',
                'machine_code': equipment.code if equipment else '',
                'input_tons': _shift_production_weight_tons(record, 'input_weight'),
                'output_tons': _shift_production_weight_tons(record, 'output_weight'),
                'scrap_tons': _shift_production_weight_tons(record, 'scrap_weight'),
                'downtime_minutes': _to_float(record.downtime_minutes),
                'quality_issue_count': _to_float(record.issue_count),
                'energy_kwh': _to_float(record.electricity_kwh),
                'source_table': 'shift_production_data',
            }
        )

    process_records = (
        db.query(MesWorkshopProcessRecord)
        .filter(MesWorkshopProcessRecord.business_date == business_date)
        .order_by(MesWorkshopProcessRecord.id.asc())
        .all()
    )
    for record in process_records:
        rows.append(
            {
                'business_date': business_date.isoformat(),
                'workshop': record.workshop_name or '',
                'shift': '',
                'process': record.process_name or '',
                'machine': record.device_name or '',
                'coil_no': record.batch_no or '',
                'input_tons': _tons_from_pair(record.input_weight_tons, record.input_weight_kg),
                'output_tons': _tons_from_pair(record.output_weight_tons, record.output_weight_kg),
                'yield_rate': _to_float(record.yield_rate),
                'source_table': 'mes_workshop_process_records',
            }
        )

    stock_records = (
        db.query(MesStockRecord)
        .filter(MesStockRecord.business_date == business_date)
        .order_by(MesStockRecord.id.asc())
        .all()
    )
    for record in stock_records:
        rows.append(
            {
                'business_date': business_date.isoformat(),
                'workshop': '成品库',
                'shift': '',
                'process': '入库',
                'machine': '',
                'coil_no': record.batch_no or '',
                'contract_no': record.contract_no or '',
                'customer': record.customer_alias or '',
                'output_tons': _tons_from_pair(record.net_weight_tons, record.net_weight_kg),
                'status': record.status_name or '',
                'source_table': 'mes_stock_records',
            }
        )

    energy_records = (
        db.query(MachineEnergyRecord, MobileShiftReport, Workshop, ShiftConfig, Equipment)
        .join(MobileShiftReport, MachineEnergyRecord.shift_report_id == MobileShiftReport.id)
        .join(Workshop, MobileShiftReport.workshop_id == Workshop.id)
        .join(ShiftConfig, MobileShiftReport.shift_config_id == ShiftConfig.id)
        .outerjoin(Equipment, MachineEnergyRecord.machine_id == Equipment.id)
        .filter(MobileShiftReport.business_date == business_date)
        .order_by(MachineEnergyRecord.id.asc())
        .all()
    )
    for energy, report, workshop, shift, equipment in energy_records:
        rows.append(
            {
                'business_date': report.business_date.isoformat(),
                'workshop': workshop.name,
                'shift': shift.name,
                'process': '能耗',
                'machine': energy.machine_name or (equipment.name if equipment else ''),
                'machine_code': energy.machine_code or (equipment.code if equipment else ''),
                'energy_kwh': _to_float(energy.energy_kwh),
                'gas_m3': _to_float(energy.gas_m3),
                'source_table': 'machine_energy_records',
            }
        )

    consumable_rows = (
        db.query(DailyConsumableLog, Workshop)
        .join(Workshop, DailyConsumableLog.workshop_id == Workshop.id)
        .filter(DailyConsumableLog.business_date == business_date)
        .order_by(DailyConsumableLog.id.asc())
        .all()
    )
    for log, workshop in consumable_rows:
        payload = log.payload or {}
        consumable_metrics = _consumable_payload_metrics(payload)
        rows.append(
            {
                'business_date': log.business_date.isoformat(),
                'workshop': workshop.name,
                'shift': '',
                'process': '内勤辅材',
                'machine': '',
                'output_tons': _to_float(payload.get('packaging_inbound_output_tons')),
                'energy_kwh': _to_float(payload.get('electricity_daily')),
                'gas_m3': _to_float(payload.get('gas_daily')),
                **consumable_metrics,
                'consumable_payload': consumable_metrics,
                'source_table': 'daily_consumable_logs',
            }
        )

    owner_daily_rows = (
        db.query(WorkOrderEntry, WorkOrder, Workshop, ShiftConfig, Equipment)
        .join(WorkOrder, WorkOrderEntry.work_order_id == WorkOrder.id)
        .join(Workshop, WorkOrderEntry.workshop_id == Workshop.id)
        .outerjoin(ShiftConfig, WorkOrderEntry.shift_id == ShiftConfig.id)
        .outerjoin(Equipment, WorkOrderEntry.machine_id == Equipment.id)
        .filter(
            WorkOrderEntry.business_date == business_date,
            WorkOrderEntry.entry_type == 'owner_daily',
            WorkOrderEntry.entry_status != 'voided',
        )
        .order_by(WorkOrderEntry.id.asc())
        .all()
    )
    for entry, work_order, workshop, shift, equipment in owner_daily_rows:
        owner_daily_metrics = _owner_daily_payload_metrics(entry.extra_payload)
        if not owner_daily_metrics:
            continue
        rows.append(
            {
                'business_date': entry.business_date.isoformat(),
                'workshop': workshop.name,
                'shift': shift.name if shift else '',
                'process': '每日一录',
                'machine': equipment.name if equipment else '',
                'machine_code': equipment.code if equipment else '',
                'contract_no': work_order.contract_no or '',
                'customer': work_order.customer_name or '',
                **owner_daily_metrics,
                'owner_daily_payload': owner_daily_metrics,
                'source_table': 'work_order_entries',
            }
        )

    cost_rows = (
        db.query(CostDailyResult, Workshop)
        .outerjoin(Workshop, CostDailyResult.workshop_code == Workshop.code)
        .filter(CostDailyResult.business_date == business_date)
        .order_by(CostDailyResult.id.asc())
        .all()
    )
    for cost, workshop in cost_rows:
        rows.append(
            {
                'business_date': cost.business_date.isoformat(),
                'workshop': workshop.name if workshop else cost.workshop_code,
                'shift': '',
                'process': '成本策略',
                'machine': '',
                'strategy_code': cost.strategy_code,
                'cost_caliber': cost.caliber,
                'total_cost': _to_float(cost.total_cost),
                'cost_per_ton': _to_float(cost.output_ton_cost),
                'throughput_cost_per_ton': _to_float(cost.throughput_ton_cost),
                'breakdown_count': cost.breakdown_count,
                'process_count': cost.process_count,
                'source_table': 'cost_daily_result',
            }
        )
    machine_cost_rows = (
        db.query(MachineDailyCostSnapshot, Workshop, Equipment)
        .join(Workshop, MachineDailyCostSnapshot.workshop_id == Workshop.id)
        .outerjoin(Equipment, MachineDailyCostSnapshot.machine_line_id == Equipment.id)
        .filter(MachineDailyCostSnapshot.business_date == business_date)
        .order_by(MachineDailyCostSnapshot.id.asc())
        .all()
    )
    for cost, workshop, equipment in machine_cost_rows:
        rows.append(
            {
                'business_date': cost.business_date.isoformat(),
                'workshop': workshop.name,
                'shift': '',
                'process': '机列日成本',
                'machine': equipment.name if equipment else '',
                'machine_code': equipment.code if equipment else '',
                'electricity_kwh': _to_float(cost.electricity_kwh),
                'electricity_cost': _to_float(cost.electricity_cost),
                'natural_gas_m3': _to_float(cost.natural_gas_m3),
                'natural_gas_cost': _to_float(cost.natural_gas_cost),
                'total_cost': _to_float(cost.total_cost),
                'is_estimated': cost.is_estimated,
                'source_table': 'machine_daily_cost_snapshots',
            }
        )
    rows.extend(_build_wip_mapping_rows(db, business_date=business_date))
    if factory_summary := _build_factory_summary_row(rows, business_date=business_date):
        rows.append(factory_summary)
    return rows
