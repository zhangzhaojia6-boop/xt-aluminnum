from __future__ import annotations

import csv
import hashlib
from dataclasses import dataclass
from datetime import date, datetime, time
from io import BytesIO, StringIO
from pathlib import Path
from typing import Literal

import openpyxl
import xlrd


DingTalkFileTextStatus = Literal['text_captured', 'unsupported_file_type', 'parse_failed', 'too_large']

_TEXT_SUFFIXES = {'.txt', '.csv', '.tsv', '.md'}
_XLSX_SUFFIXES = {'.xlsx', '.xlsm'}


@dataclass(frozen=True)
class DingTalkFileText:
    status: DingTalkFileTextStatus
    text: str
    detail: str
    content_hash: str


def extract_dingtalk_file_text(file_name: str, content: bytes, max_bytes: int) -> DingTalkFileText:
    content_hash = hashlib.sha256(content).hexdigest()
    suffix = Path(str(file_name or '')).suffix.lower()

    if len(content) > max_bytes:
        return DingTalkFileText(
            status='too_large',
            text='',
            detail=f'file_size_exceeds_limit bytes={len(content)} max_bytes={max_bytes}',
            content_hash=content_hash,
        )

    if suffix in _TEXT_SUFFIXES:
        return _extract_plain_text(suffix=suffix, content=content, content_hash=content_hash)
    if suffix in _XLSX_SUFFIXES:
        return _extract_xlsx_text(content=content, content_hash=content_hash)
    if suffix == '.xls':
        return _extract_xls_text(content=content, content_hash=content_hash)

    return DingTalkFileText(
        status='unsupported_file_type',
        text='',
        detail=f'unsupported_suffix={suffix or "none"}',
        content_hash=content_hash,
    )


def _extract_plain_text(*, suffix: str, content: bytes, content_hash: str) -> DingTalkFileText:
    decoded, encoding = _decode_text(content)
    if decoded is None:
        return DingTalkFileText(
            status='parse_failed',
            text='',
            detail='text_decode_failed',
            content_hash=content_hash,
        )

    if suffix == '.csv':
        text = _normalize_delimited_text(decoded, delimiter=',')
    elif suffix == '.tsv':
        text = _normalize_delimited_text(decoded, delimiter='\t')
    else:
        text = _normalize_free_text(decoded)

    return DingTalkFileText(
        status='text_captured',
        text=text,
        detail=f'encoding={encoding}',
        content_hash=content_hash,
    )


def _extract_xlsx_text(*, content: bytes, content_hash: str) -> DingTalkFileText:
    try:
        workbook = openpyxl.load_workbook(BytesIO(content), read_only=True, data_only=True)
        try:
            text = _workbook_to_text(workbook)
        finally:
            workbook.close()
    except Exception as exc:  # pragma: no cover - exact parser errors vary by openpyxl version
        return DingTalkFileText(
            status='parse_failed',
            text='',
            detail=f'xlsx_parse_failed error={exc.__class__.__name__}',
            content_hash=content_hash,
        )

    return DingTalkFileText(
        status='text_captured',
        text=text,
        detail='xlsx_parsed',
        content_hash=content_hash,
    )


def _extract_xls_text(*, content: bytes, content_hash: str) -> DingTalkFileText:
    try:
        workbook = xlrd.open_workbook(file_contents=content)
        lines: list[str] = []
        for sheet in workbook.sheets():
            sheet_lines = _xls_sheet_to_lines(sheet, getattr(workbook, 'datemode', 0))
            if sheet_lines:
                lines.append(f'[{_normalize_inline(sheet.name)}]')
                lines.extend(sheet_lines)
    except Exception as exc:  # pragma: no cover - exact parser errors vary by xlrd version
        return DingTalkFileText(
            status='parse_failed',
            text='',
            detail=f'xls_parse_failed error={exc.__class__.__name__}',
            content_hash=content_hash,
        )

    return DingTalkFileText(
        status='text_captured',
        text='\n'.join(lines),
        detail='xls_parsed',
        content_hash=content_hash,
    )


def _decode_text(content: bytes) -> tuple[str | None, str]:
    for encoding in ('utf-8-sig', 'utf-8', 'gb18030'):
        try:
            return content.decode(encoding), encoding
        except UnicodeDecodeError:
            continue
    return None, ''


def _workbook_to_text(workbook) -> str:
    lines: list[str] = []
    for worksheet in workbook.worksheets:
        sheet_lines: list[str] = []
        for row in worksheet.iter_rows(values_only=True):
            row_text = _cells_to_line(row)
            if row_text:
                sheet_lines.append(row_text)
        if sheet_lines:
            lines.append(f'[{_normalize_inline(worksheet.title)}]')
            lines.extend(sheet_lines)
    return '\n'.join(lines)


def _xls_sheet_to_lines(sheet, datemode: int) -> list[str]:
    lines: list[str] = []
    for row_index in range(sheet.nrows):
        cells: list[str] = []
        for column_index in range(sheet.ncols):
            cells.append(_cell_to_text(_xls_cell_value(sheet, row_index, column_index, datemode)))
        row_text = _cells_to_line(cells)
        if row_text:
            lines.append(row_text)
    return lines


def _xls_cell_value(sheet, row_index: int, column_index: int, datemode: int):
    try:
        cell = sheet.cell(row_index, column_index)
    except AttributeError:
        return sheet.cell_value(row_index, column_index)
    if cell.ctype == xlrd.XL_CELL_DATE:
        try:
            return xlrd.xldate.xldate_as_datetime(cell.value, datemode)
        except Exception:
            return cell.value
    return cell.value


def _cells_to_line(row) -> str:
    cells = [_cell_to_text(value) for value in row]
    cells = [cell for cell in cells if cell]
    return '\t'.join(cells)


def _cell_to_text(value) -> str:
    if value is None:
        return ''
    if isinstance(value, bool):
        return 'TRUE' if value else 'FALSE'
    if isinstance(value, datetime):
        return value.isoformat(sep=' ')
    if isinstance(value, (date, time)):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return _normalize_inline(str(value))


def _normalize_free_text(text: str) -> str:
    lines = [_normalize_inline(line) for line in _split_lines(text)]
    return '\n'.join(line for line in lines if line)


def _normalize_delimited_text(text: str, *, delimiter: str) -> str:
    rows = csv.reader(StringIO(text), delimiter=delimiter)
    lines = [_cells_to_line(row) for row in rows]
    return '\n'.join(line for line in lines if line)


def _normalize_inline(value: str) -> str:
    return ' '.join(str(value).replace('\u3000', ' ').split())


def _split_lines(text: str) -> list[str]:
    return text.replace('\r\n', '\n').replace('\r', '\n').split('\n')
