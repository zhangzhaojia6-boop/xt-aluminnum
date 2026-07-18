from __future__ import annotations

from collections.abc import Mapping
from datetime import date, datetime, timedelta
import hashlib
from io import BytesIO
import math
from pathlib import Path
import re
from typing import Any
import unicodedata

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet

from app.core.business_time import OWNER_DAILY_CUTOFF, local_now


PLAN_CONTRACT_PARSER = 'plan_contract_message_v1'
WIP_SCREENSHOT_PARSER = 'owner_verified_wip_screenshot_v1'
WIP_SCREENSHOT_TABLE_PARSER = 'owner_verified_wip_screenshot_v2'
WIP_VISUAL_FIELD_ROWS = {
    'wip_total': '汇总',
    'wip_1650_2050_cold': '1650+2050冷轧合计',
    'wip_1850_cold': '1850冷轧合计',
    'wip_milling': '铣床合计',
    'wip_anneal_total': '在线退火合计',
    'wip_new_north': '新厂北',
    'wip_new_south': '新厂南',
    'wip_park_anneal': '园区退火',
    'wip_finishing_total': '后工序合计',
    'wip_straightening': '拉矫合计',
    'wip_finishing': '精整合计',
    'wip_park_finishing': '园区剪切',
    'wip_hot_plate_shearing': '热轧（中厚板）',
    'wip_coating': '彩涂',
}
WIP_TOTAL_COMPONENT_FIELDS = (
    'wip_1650_2050_cold',
    'wip_1850_cold',
    'wip_milling',
    'wip_anneal_total',
    'wip_finishing_total',
    'wip_hot_plate_shearing',
    'wip_coating',
)
_NUMBER_PATTERN = r'(?P<value>(?:\d{1,3}(?:,\d{3})+|\d+)(?:\.\d+)?)'
_PLAN_CONTRACT_PATTERNS = {
    '2050_input': re.compile(rf'2050\s*投料\s*{_NUMBER_PATTERN}\s*吨'),
    '1850_input': re.compile(rf'1850\s*投料\s*{_NUMBER_PATTERN}\s*吨'),
    'external_processing': re.compile(rf'外加工\s*{_NUMBER_PATTERN}\s*吨'),
    'medium_plate': re.compile(rf'中厚板\s*{_NUMBER_PATTERN}\s*吨'),
    'daily_contract': re.compile(rf'当天合同\s*{_NUMBER_PATTERN}\s*吨'),
    'hot_rolling_contract': re.compile(rf'热轧\s*{_NUMBER_PATTERN}\s*吨'),
    'remaining_contract': re.compile(rf'总余合同量\s*{_NUMBER_PATTERN}\s*吨'),
}


def extract_owner_verified_visual_fact_updates(
    *,
    file_name: str,
    content: bytes,
    business_date: date,
    event_time: str | datetime,
    file_sha256: str,
    verified_facts: Mapping[str, Any],
) -> dict[str, dict[str, Any]]:
    clean_hash = str(file_sha256 or '').strip().lower()
    if hashlib.sha256(content).hexdigest() != clean_hash:
        return {}
    if not _is_supported_image(file_name, content):
        return {}
    field_names = set(verified_facts)
    if field_names == {'wip_total'}:
        return _extract_owner_verified_wip_total(
            business_date=business_date,
            event_time=event_time,
            clean_hash=clean_hash,
            verified_facts=verified_facts,
        )
    if field_names != set(WIP_VISUAL_FIELD_ROWS):
        return {}

    lineage = _verified_visual_lineage(
        business_date=business_date,
        event_time=event_time,
        verified_facts=verified_facts,
    )
    if lineage is None:
        return {}
    reported_date, local_event_time = lineage

    values: dict[str, int | float] = {}
    for field_name, expected_row_label in WIP_VISUAL_FIELD_ROWS.items():
        item = verified_facts.get(field_name)
        if not isinstance(item, Mapping):
            return {}
        if _normalize_label(item.get('row_label')) != _normalize_label(expected_row_label):
            return {}
        if _normalize_label(item.get('column_label')) != _normalize_label('在制料'):
            return {}
        if str(item.get('unit') or '').strip() != '吨':
            return {}
        value = _finite_nonnegative_number(item.get('value'))
        if value is None:
            return {}
        values[field_name] = value

    if not _wip_values_are_consistent(values):
        return {}

    return {
        field_name: {
            'value': value,
            'unit': '吨',
            'confidence': 0.99,
            'reason': '钉钉在制料截图人工确权字段',
            'source_ref': {
                'parser': WIP_SCREENSHOT_TABLE_PARSER,
                'verification_mode': 'owner_verified_visual',
                'reported_date': reported_date.isoformat(),
                'business_date': business_date.isoformat(),
                'business_date_rule': 'next_calendar_day_before_owner_daily_cutoff',
                'event_time': local_event_time.isoformat(),
                'event_time_cutoff': OWNER_DAILY_CUTOFF.isoformat(timespec='minutes'),
                'row_label': WIP_VISUAL_FIELD_ROWS[field_name],
                'column_label': '在制料',
                'file_sha256': clean_hash,
            },
        }
        for field_name, value in values.items()
    }


def _extract_owner_verified_wip_total(
    *,
    business_date: date,
    event_time: str | datetime,
    clean_hash: str,
    verified_facts: Mapping[str, Any],
) -> dict[str, dict[str, Any]]:
    lineage = _verified_visual_lineage(
        business_date=business_date,
        event_time=event_time,
        verified_facts=verified_facts,
    )
    if lineage is None:
        return {}
    reported_date, _local_event_time = lineage

    item = verified_facts.get('wip_total')
    if not isinstance(item, Mapping):
        return {}
    if str(item.get('unit') or '').strip() != '吨':
        return {}
    if str(item.get('row_label') or '').strip() != '汇总':
        return {}
    if str(item.get('column_label') or '').strip() != '在制料':
        return {}
    value = _finite_nonnegative_number(item.get('value'))
    if value is None:
        return {}

    return {
        'wip_total': {
            'value': value,
            'unit': '吨',
            'confidence': 0.99,
            'reason': '钉钉在制料截图人工确权汇总值',
            'source_ref': {
                'parser': WIP_SCREENSHOT_PARSER,
                'verification_mode': 'owner_verified_visual',
                'reported_date': reported_date.isoformat(),
                'business_date': business_date.isoformat(),
                'business_date_rule': 'next_calendar_day_before_owner_daily_cutoff',
                'event_time_cutoff': OWNER_DAILY_CUTOFF.isoformat(timespec='minutes'),
                'row_label': '汇总',
                'column_label': '在制料',
                'file_sha256': clean_hash,
            },
        }
    }


def _verified_visual_lineage(
    *,
    business_date: date,
    event_time: str | datetime,
    verified_facts: Mapping[str, Any],
) -> tuple[date, datetime] | None:
    reported_dates = {
        _iso_date(item.get('reported_date'))
        for item in verified_facts.values()
        if isinstance(item, Mapping)
    }
    if len(reported_dates) != 1 or None in reported_dates:
        return None
    reported_date = next(iter(reported_dates))
    parsed_event_time = _iso_datetime(event_time)
    if parsed_event_time is None:
        return None
    local_event_time = local_now(parsed_event_time)
    if reported_date != business_date + timedelta(days=1):
        return None
    if local_event_time.date() != reported_date:
        return None
    if local_event_time.time().replace(tzinfo=None) >= OWNER_DAILY_CUTOFF:
        return None
    return reported_date, local_event_time


def _wip_values_are_consistent(values: Mapping[str, int | float]) -> bool:
    total = sum(float(values[field]) for field in WIP_TOTAL_COMPONENT_FIELDS)
    anneal = sum(float(values[field]) for field in ('wip_new_north', 'wip_new_south', 'wip_park_anneal'))
    finishing = sum(
        float(values[field])
        for field in ('wip_straightening', 'wip_finishing', 'wip_park_finishing')
    )
    return (
        math.isclose(float(values['wip_total']), total, abs_tol=0.01)
        and math.isclose(float(values['wip_anneal_total']), anneal, abs_tol=0.01)
        and math.isclose(float(values['wip_finishing_total']), finishing, abs_tol=0.01)
    )


def _normalize_label(value: Any) -> str:
    return re.sub(r'\s+', '', unicodedata.normalize('NFKC', str(value or '')))


def extract_verified_file_fact_updates(
    *,
    file_name: str,
    content: bytes,
    business_date: date,
    file_sha256: str,
) -> dict[str, dict[str, Any]]:
    if Path(file_name).suffix.lower() not in {'.xlsx', '.xlsm'}:
        return {}
    try:
        workbook = load_workbook(BytesIO(content), data_only=True, read_only=False)
    except Exception:  # noqa: BLE001
        return {}
    modified = workbook.properties.modified
    modified_date = modified.date() if isinstance(modified, datetime) else modified
    if not isinstance(modified_date, date) or modified_date != business_date:
        return {}

    for sheet in workbook.worksheets:
        source_ref = _company_daily_yield_source(sheet, business_date=business_date)
        if source_ref is None:
            continue
        value = _yield_percent(sheet[source_ref['value_cell']].value)
        if value is None:
            continue
        updates = {
            'daily_yield_rate': {
                'value': round(value, 2),
                'unit': '%',
                'confidence': 0.99,
                'reason': '钉钉成品率工作簿确定性单元格映射',
                'source_ref': {
                    'parser': 'company_daily_yield_v1',
                    'sheet_name': sheet.title,
                    'workbook_modified_date': modified_date.isoformat(),
                    **source_ref,
                    'file_sha256': file_sha256,
                },
            }
        }
        updates.update(
            _additional_company_yield_updates(
                sheet,
                business_date=business_date,
                modified_date=modified_date,
                file_sha256=file_sha256,
                daily_source=source_ref,
            )
        )
        return updates
    return {}


def extract_verified_text_fact_updates(
    *,
    text: str,
    business_date: date,
    content_sha256: str,
) -> dict[str, dict[str, Any]]:
    clean_hash = str(content_sha256 or '').strip().lower()
    if hashlib.sha256(text.encode('utf-8')).hexdigest() != clean_hash:
        return {}
    parsed = parse_plan_contract_message(text)
    if parsed is None:
        return {}

    base_source_ref = {
        'parser': PLAN_CONTRACT_PARSER,
        'business_date': business_date.isoformat(),
        'content_sha256': clean_hash,
        'matched_segments': parsed['matched_segments'],
    }
    updates = {
        'daily_input_weight': {
            'value': parsed['daily_input_weight'],
            'unit': '吨',
            'confidence': 0.99,
            'reason': '钉钉计划科合同消息确定性分项求和',
            'source_ref': {
                **base_source_ref,
                'components': parsed['components'],
            },
        },
        'cold_roll_input_daily': {
            'value': parsed['cold_roll_input_weight'],
            'unit': '吨',
            'confidence': 0.99,
            'reason': '钉钉计划科合同消息冷轧投料分项求和',
            'source_ref': {
                **base_source_ref,
                'components': parsed['cold_roll_components'],
            },
        },
        'remaining_contract_weight': {
            'value': parsed['remaining_contract_weight'],
            'unit': '吨',
            'confidence': 0.99,
            'reason': '钉钉计划科合同消息明确标签值',
            'source_ref': {
                **base_source_ref,
                'context_values': parsed['context_values'],
            },
        },
    }
    component_fields = {
        'cold_2050_input_daily': '2050_input',
        'cold_1850_input_daily': '1850_input',
        'outsourced_input_daily': 'external_processing',
        'medium_plate_input_daily': 'medium_plate',
    }
    for field_name, component_name in component_fields.items():
        updates[field_name] = {
            'value': parsed['components'][component_name],
            'unit': '吨',
            'confidence': 0.99,
            'reason': '钉钉计划科合同消息明确投料分项',
            'source_ref': {
                **base_source_ref,
                'component': component_name,
                'matched_segment': parsed['matched_segments'][component_name],
            },
        }
    context_fields = {
        'daily_contract_weight': 'daily_contract',
        'daily_hot_roll_contract_weight': 'hot_rolling_contract',
    }
    for field_name, context_name in context_fields.items():
        updates[field_name] = {
            'value': parsed['context_values'][context_name],
            'unit': '吨',
            'confidence': 0.99,
            'reason': '钉钉计划科合同消息明确合同字段',
            'source_ref': {
                **base_source_ref,
                'context_field': context_name,
                'matched_segment': parsed['matched_segments'][context_name],
            },
        }
    return updates


def parse_plan_contract_message(text: str) -> dict[str, Any] | None:
    normalized = unicodedata.normalize('NFKC', str(text or ''))
    if len(re.findall(r'投料量\s*:', normalized)) != 1:
        return None

    values: dict[str, int | float] = {}
    matched_segments: dict[str, dict[str, Any]] = {}
    for key, pattern in _PLAN_CONTRACT_PATTERNS.items():
        matches = list(pattern.finditer(normalized))
        if len(matches) != 1:
            return None
        match = matches[0]
        values[key] = _plain_number(match.group('value'))
        matched_segments[key] = {
            'text': match.group(0),
            'start': match.start(),
            'end': match.end(),
        }

    components = {
        key: values[key]
        for key in ('2050_input', '1850_input', 'external_processing', 'medium_plate')
    }
    cold_roll_components = {
        key: components[key]
        for key in ('2050_input', '1850_input', 'external_processing')
    }
    return {
        'daily_input_weight': sum(components.values()),
        'cold_roll_input_weight': sum(cold_roll_components.values()),
        'remaining_contract_weight': values['remaining_contract'],
        'components': components,
        'cold_roll_components': cold_roll_components,
        'context_values': {
            'daily_contract': values['daily_contract'],
            'hot_rolling_contract': values['hot_rolling_contract'],
        },
        'matched_segments': matched_segments,
    }


def _plain_number(value: str) -> int | float:
    number = float(value.replace(',', ''))
    return int(number) if number.is_integer() else number


def _is_supported_image(file_name: str, content: bytes) -> bool:
    suffix = Path(file_name).suffix.lower()
    if suffix not in {'.jpg', '.jpeg', '.png'}:
        return False
    is_jpeg = content.startswith(b'\xff\xd8\xff') and content.endswith(b'\xff\xd9')
    is_png = content.startswith(b'\x89PNG\r\n\x1a\n') and content.endswith(b'IEND\xaeB`\x82')
    return is_jpeg or is_png


def _iso_date(value: Any) -> date | None:
    try:
        return date.fromisoformat(str(value or '').strip())
    except ValueError:
        return None


def _iso_datetime(value: Any) -> datetime | None:
    if isinstance(value, datetime):
        return value
    try:
        return datetime.fromisoformat(str(value or '').strip().replace('Z', '+00:00'))
    except ValueError:
        return None


def _finite_nonnegative_number(value: Any) -> int | float | None:
    if isinstance(value, bool) or value in (None, ''):
        return None
    try:
        number = float(str(value).replace(',', ''))
    except ValueError:
        return None
    if not math.isfinite(number) or number < 0:
        return None
    return int(number) if number.is_integer() else number


def _company_daily_yield_source(sheet: Worksheet, *, business_date: date) -> dict[str, str] | None:
    if not _sheet_period_matches(sheet, business_date=business_date):
        return None
    date_header = _find_header(sheet, '日期')
    total_header = _find_header(sheet, '总成品率')
    if date_header is None or total_header is None:
        return None

    total_min_col, total_max_col, total_max_row = _merged_bounds(sheet, total_header)
    daily_header = _find_text_in_range(
        sheet,
        text='日合计',
        min_row=total_max_row + 1,
        max_row=min(total_max_row + 4, sheet.max_row),
        min_col=total_min_col,
        max_col=total_max_col,
    )
    if daily_header is None:
        return None

    _, _, date_header_max_row = _merged_bounds(sheet, date_header)
    first_data_row = max(date_header_max_row, daily_header.row) + 1
    matching_rows = [
        row_number
        for row_number in range(first_data_row, sheet.max_row + 1)
        if _cell_matches_business_date(
            sheet.cell(row=row_number, column=date_header.column).value,
            business_date=business_date,
        )
    ]
    if len(matching_rows) != 1:
        return None
    row_number = matching_rows[0]
    date_cell = sheet.cell(row=row_number, column=date_header.column)
    value_cell = sheet.cell(row=row_number, column=daily_header.column)
    if _yield_percent(value_cell.value) is None:
        return None
    return {
        'date_cell': date_cell.coordinate,
        'value_cell': value_cell.coordinate,
        'header_cell': total_header.coordinate,
    }


def _additional_company_yield_updates(
    sheet: Worksheet,
    *,
    business_date: date,
    modified_date: date,
    file_sha256: str,
    daily_source: Mapping[str, str],
) -> dict[str, dict[str, Any]]:
    specs = (
        ('monthly_yield_rate', '总成品率', '月累计'),
        ('hot_roll_yield_rate', '热轧普卷板', '日合计'),
        ('hot_roll_monthly_yield_rate', '热轧普卷板', '月累计'),
        ('cast_roll_yield_rate', '铸轧（幕墙卷+普卷板）', '月累计'),
        ('plate_coil_yield_rate', '铸轧普卷板', '月累计'),
    )
    updates: dict[str, dict[str, Any]] = {}
    sources: dict[str, dict[str, str]] = {'daily_yield_rate': dict(daily_source)}
    for field_name, header_text, period_label in specs:
        source_ref = _company_yield_metric_source(
            sheet,
            business_date=business_date,
            header_text=header_text,
            period_label=period_label,
        )
        if source_ref is None:
            continue
        value = _yield_percent(sheet[source_ref['value_cell']].value)
        if value is None:
            continue
        sources[field_name] = source_ref
        updates[field_name] = {
            'value': round(value, 2),
            'unit': '%',
            'confidence': 0.99,
            'reason': '钉钉成品率工作簿确定性单元格映射',
            'source_ref': _yield_source_ref(
                sheet,
                modified_date=modified_date,
                file_sha256=file_sha256,
                source_ref=source_ref,
            ),
        }

    for field_name, current_field in (
        ('daily_yield_delta', 'daily_yield_rate'),
        ('hot_roll_yield_delta', 'hot_roll_yield_rate'),
    ):
        current_source = sources.get(current_field)
        if current_source is None:
            continue
        delta = _yield_day_over_day_delta(
            sheet,
            business_date=business_date,
            current_source=current_source,
        )
        if delta is None:
            continue
        value, previous_date_cell, previous_value_cell = delta
        updates[field_name] = {
            'value': value,
            'unit': '%',
            'confidence': 0.99,
            'reason': '钉钉成品率工作簿当日与昨日确定性单元格差值',
            'source_ref': {
                **_yield_source_ref(
                    sheet,
                    modified_date=modified_date,
                    file_sha256=file_sha256,
                    source_ref=current_source,
                ),
                'previous_date_cell': previous_date_cell,
                'previous_value_cell': previous_value_cell,
                'comparison': 'current_day_minus_previous_calendar_day',
            },
        }
    return updates


def _company_yield_metric_source(
    sheet: Worksheet,
    *,
    business_date: date,
    header_text: str,
    period_label: str,
) -> dict[str, str] | None:
    date_header = _find_header(sheet, '日期')
    metric_header = _find_header(sheet, header_text)
    if date_header is None or metric_header is None:
        return None
    metric_min_col, metric_max_col, metric_max_row = _merged_bounds(sheet, metric_header)
    period_header = _find_text_in_range(
        sheet,
        text=period_label,
        min_row=metric_max_row + 1,
        max_row=min(metric_max_row + 4, sheet.max_row),
        min_col=metric_min_col,
        max_col=metric_max_col,
    )
    if period_header is None:
        return None
    _, _, date_header_max_row = _merged_bounds(sheet, date_header)
    first_data_row = max(date_header_max_row, period_header.row) + 1
    matching_rows = [
        row_number
        for row_number in range(first_data_row, sheet.max_row + 1)
        if _cell_matches_business_date(
            sheet.cell(row=row_number, column=date_header.column).value,
            business_date=business_date,
        )
    ]
    if len(matching_rows) != 1:
        return None
    row_number = matching_rows[0]
    value_cell = sheet.cell(row=row_number, column=period_header.column)
    if _yield_percent(value_cell.value) is None:
        return None
    return {
        'date_cell': sheet.cell(row=row_number, column=date_header.column).coordinate,
        'value_cell': value_cell.coordinate,
        'header_cell': metric_header.coordinate,
    }


def _yield_day_over_day_delta(
    sheet: Worksheet,
    *,
    business_date: date,
    current_source: Mapping[str, str],
) -> tuple[float, str, str] | None:
    current_date_cell = sheet[str(current_source['date_cell'])]
    current_value_cell = sheet[str(current_source['value_cell'])]
    previous_date = business_date - timedelta(days=1)
    previous_rows = [
        row_number
        for row_number in range(1, sheet.max_row + 1)
        if _cell_matches_business_date(
            sheet.cell(row=row_number, column=current_date_cell.column).value,
            business_date=previous_date,
        )
    ]
    if len(previous_rows) != 1:
        return None
    previous_date_cell = sheet.cell(row=previous_rows[0], column=current_date_cell.column)
    previous_value_cell = sheet.cell(row=previous_rows[0], column=current_value_cell.column)
    current_value = _yield_percent(current_value_cell.value)
    previous_value = _yield_percent(previous_value_cell.value)
    if current_value is None or previous_value is None:
        return None
    return (
        round(round(current_value, 2) - round(previous_value, 2), 2),
        previous_date_cell.coordinate,
        previous_value_cell.coordinate,
    )


def _yield_source_ref(
    sheet: Worksheet,
    *,
    modified_date: date,
    file_sha256: str,
    source_ref: Mapping[str, str],
) -> dict[str, str]:
    return {
        'parser': 'company_daily_yield_v1',
        'sheet_name': sheet.title,
        'workbook_modified_date': modified_date.isoformat(),
        **dict(source_ref),
        'file_sha256': file_sha256,
    }


def _sheet_period_matches(sheet: Worksheet, *, business_date: date) -> bool:
    header_text = ' '.join(
        str(sheet.cell(row=row, column=column).value or '').strip()
        for row in range(1, min(sheet.max_row, 5) + 1)
        for column in range(1, min(sheet.max_column, 20) + 1)
    )
    month_match = re.search(r'(?<!\d)(\d{1,2})\s*月份?', header_text)
    if month_match is None or int(month_match.group(1)) != business_date.month:
        return False
    year_match = re.search(r'(20\d{2})\s*年', header_text)
    return year_match is None or int(year_match.group(1)) == business_date.year


def _find_header(sheet: Worksheet, text: str) -> Cell | None:
    return _find_text_in_range(
        sheet,
        text=text,
        min_row=1,
        max_row=min(sheet.max_row, 10),
        min_col=1,
        max_col=sheet.max_column,
    )


def _find_text_in_range(
    sheet: Worksheet,
    *,
    text: str,
    min_row: int,
    max_row: int,
    min_col: int,
    max_col: int,
) -> Cell | None:
    matches: list[Cell] = []
    for row in range(min_row, max_row + 1):
        for column in range(min_col, max_col + 1):
            cell = sheet.cell(row=row, column=column)
            if text in str(cell.value or '').replace(' ', ''):
                matches.append(cell)
    return matches[0] if len(matches) == 1 else None


def _merged_bounds(sheet: Worksheet, cell: Cell) -> tuple[int, int, int]:
    for merged in sheet.merged_cells.ranges:
        if cell.coordinate in merged:
            return merged.min_col, merged.max_col, merged.max_row
    return cell.column, cell.column, cell.row


def _cell_matches_business_date(value: Any, *, business_date: date) -> bool:
    if isinstance(value, datetime):
        return value.date() == business_date
    if isinstance(value, date):
        return value == business_date
    if isinstance(value, bool):
        return False
    if isinstance(value, (int, float)):
        return int(value) == business_date.day and float(value).is_integer()
    clean = str(value or '').strip()
    return clean in {str(business_date.day), business_date.isoformat()}


def _yield_percent(value: Any) -> float | None:
    if isinstance(value, bool) or value in (None, ''):
        return None
    if isinstance(value, (int, float)):
        number = float(value)
    else:
        clean = str(value).strip().replace(',', '')
        if clean.endswith('%'):
            clean = clean[:-1].strip()
        try:
            number = float(clean)
        except ValueError:
            return None
    if 0 < number <= 1:
        number *= 100
    if not 0 < number <= 100:
        return None
    return number
